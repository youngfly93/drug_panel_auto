"""Production-only knowledge release gate and multidimensional coverage report."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence

import yaml
from openpyxl import load_workbook

from reportgen.knowledge.governance import (
    effective_governance,
    load_and_validate_overlay,
    validate_knowledge_rows,
)
from reportgen.knowledge.quality import (
    profile_panel_runtime_content,
    profile_panel_targeted_drug_contracts,
)
from reportgen.panels.loader import PanelPackageLoader


DEFAULT_PANELS = ("crc_301_msi", "crc_358_msi")


def _discover_panel_statuses(project_root: Path) -> dict[str, str]:
    statuses: dict[str, str] = {}
    for path in sorted((project_root / "panels").glob("*/panel.yaml")):
        raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        panel_id = str(raw.get("panel_id") or path.parent.name)
        statuses[panel_id] = str(raw.get("status") or "unknown")
    return statuses


def _clinical_release_registry(project_root: Path) -> dict[str, Any]:
    path = project_root / "data/knowledge_bases/review/clinical_release_registry.yaml"
    if not path.exists():
        return {"path": str(path.relative_to(project_root)), "policy": {}, "panels": {}}
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return {
        "path": str(path.relative_to(project_root)),
        "policy": raw.get("policy") or {},
        "panels": raw.get("panels") or {},
    }


def _secondary_review_receipt(
    project_root: Path,
    registry: Mapping[str, Any],
    panel_id: str,
) -> dict[str, Any]:
    """Validate that report-group approval is bound to the current file bytes."""
    record = (registry.get("panels") or {}).get(panel_id) or {}
    declared_path = str(record.get("secondary_review_receipt") or "").strip()
    issues: list[dict[str, Any]] = []
    if not declared_path:
        return {
            "status": "FAIL",
            "path": "",
            "receipt_id": "",
            "reviewed_at": "",
            "reviewer_group": "",
            "artifacts": [],
            "issues": [
                {
                    "code": "MISSING_SECONDARY_REVIEW_RECEIPT",
                    "message": f"{panel_id} has no secondary review receipt",
                }
            ],
        }

    path = (project_root / declared_path).resolve()
    try:
        path.relative_to(project_root)
    except ValueError:
        issues.append(
            {
                "code": "INVALID_SECONDARY_REVIEW_RECEIPT_PATH",
                "message": declared_path,
            }
        )
    if not path.exists() or not path.is_file():
        issues.append(
            {
                "code": "MISSING_SECONDARY_REVIEW_RECEIPT",
                "message": declared_path,
            }
        )
        return {
            "status": "FAIL",
            "path": declared_path,
            "receipt_id": "",
            "reviewed_at": "",
            "reviewer_group": "",
            "artifacts": [],
            "issues": issues,
        }

    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    receipt_id = str(raw.get("receipt_id") or "").strip()
    reviewed_at = str(raw.get("reviewed_at") or "").strip()
    reviewer_group = str((raw.get("reviewer") or {}).get("group") or "").strip()
    expected_receipt_id = str(record.get("secondary_review_receipt_id") or "").strip()
    if str(raw.get("schema_version") or "") != "1.0":
        issues.append(
            {
                "code": "INVALID_SECONDARY_REVIEW_RECEIPT_SCHEMA",
                "message": declared_path,
            }
        )
    if str(raw.get("status") or "").strip().lower() != "approved":
        issues.append(
            {
                "code": "SECONDARY_REVIEW_NOT_APPROVED",
                "message": declared_path,
            }
        )
    if str(raw.get("decision") or "").strip().lower() != "approved_for_runtime":
        issues.append(
            {
                "code": "INVALID_SECONDARY_REVIEW_DECISION",
                "message": declared_path,
            }
        )
    if not receipt_id or (expected_receipt_id and receipt_id != expected_receipt_id):
        issues.append(
            {
                "code": "SECONDARY_REVIEW_RECEIPT_ID_MISMATCH",
                "message": f"expected={expected_receipt_id} actual={receipt_id}",
            }
        )
    if not reviewed_at or not reviewer_group:
        issues.append(
            {
                "code": "INCOMPLETE_SECONDARY_REVIEWER_METADATA",
                "message": declared_path,
            }
        )
    panels = {
        str(value).strip()
        for value in (raw.get("scope") or {}).get("panels") or []
        if str(value).strip()
    }
    if panel_id not in panels:
        issues.append(
            {
                "code": "SECONDARY_REVIEW_PANEL_SCOPE_MISMATCH",
                "message": f"{panel_id} is not covered by {receipt_id}",
            }
        )

    artifact_results: list[dict[str, Any]] = []
    artifacts = raw.get("artifacts") or []
    if not isinstance(artifacts, list) or not artifacts:
        issues.append(
            {
                "code": "EMPTY_SECONDARY_REVIEW_ARTIFACTS",
                "message": declared_path,
            }
        )
        artifacts = []
    for artifact in artifacts:
        artifact = artifact if isinstance(artifact, Mapping) else {}
        relative = str(artifact.get("path") or "").strip()
        expected = str(artifact.get("sha256") or "").strip().lower()
        target = (project_root / relative).resolve()
        valid_path = bool(relative)
        try:
            target.relative_to(project_root)
        except ValueError:
            valid_path = False
        exists = valid_path and target.exists() and target.is_file()
        actual = _sha256(target) if exists else ""
        matches = bool(re.fullmatch(r"[0-9a-f]{64}", expected)) and actual == expected
        artifact_results.append(
            {
                "path": relative,
                "expected_sha256": expected,
                "actual_sha256": actual,
                "sha256_matches": matches,
            }
        )
        if not matches:
            issues.append(
                {
                    "code": "SECONDARY_REVIEW_ARTIFACT_MISMATCH",
                    "message": relative or "<missing path>",
                }
            )
    declared_bundle = str(
        (raw.get("candidate_identity") or {}).get("value") or ""
    ).strip().lower()
    bundle_digest = hashlib.sha256()
    for artifact in sorted(artifact_results, key=lambda item: item["path"]):
        bundle_digest.update(artifact["path"].encode("utf-8"))
        bundle_digest.update(b"\0")
        bundle_digest.update(artifact["actual_sha256"].encode("ascii"))
        bundle_digest.update(b"\n")
    actual_bundle = bundle_digest.hexdigest()
    if not re.fullmatch(r"[0-9a-f]{64}", declared_bundle) or (
        declared_bundle != actual_bundle
    ):
        issues.append(
            {
                "code": "SECONDARY_REVIEW_BUNDLE_MISMATCH",
                "message": (
                    f"expected={declared_bundle} actual={actual_bundle}"
                ),
            }
        )
    return {
        "status": "PASS" if not issues else "FAIL",
        "path": declared_path,
        "receipt_id": receipt_id,
        "reviewed_at": reviewed_at,
        "reviewer_group": reviewer_group,
        "expected_bundle_sha256": declared_bundle,
        "actual_bundle_sha256": actual_bundle,
        "artifacts": artifact_results,
        "issues": issues,
    }


def _clinical_readiness(
    registry: Mapping[str, Any],
    panel_id: str,
    status_counts: Counter[str],
    secondary_receipt: Mapping[str, Any],
) -> dict[str, Any]:
    policy = registry.get("policy") or {}
    record = (registry.get("panels") or {}).get(panel_id) or {}
    uat = record.get("uat") or {}
    reviewed = int(uat.get("reviewed_reports", 0) or 0)
    passed = int(uat.get("passed_reports", 0) or 0)
    observed_rate = round(100.0 * passed / reviewed, 2) if reviewed else None
    declared_rate = uat.get("pass_rate_percent")
    rate = observed_rate if observed_rate is not None else declared_rate
    threshold = float(policy.get("uat_pass_rate_threshold_percent", 90.0) or 90.0)
    minimum = int(policy.get("minimum_reviewed_reports", 10) or 10)
    pending_rows = status_counts.get("provisional_runtime", 0) + status_counts.get(
        "needs_review", 0
    )
    reasons: list[str] = []
    secondary_status = str(record.get("secondary_review_status") or "not_recorded")
    if secondary_status not in {"completed", "approved", "report_group_approved"}:
        reasons.append("secondary_review_registry_not_completed")
    if secondary_receipt.get("status") != "PASS":
        reasons.append("secondary_review_receipt_invalid")
    if pending_rows:
        reasons.append("pending_report_group_secondary_review")
    if reviewed < minimum:
        reasons.append("insufficient_uat_reports")
    if rate is None or float(rate) < threshold:
        reasons.append("uat_pass_rate_below_threshold_or_unknown")
    return {
        "status": "READY" if not reasons else "BLOCKED",
        "secondary_review": {
            "owner": str(record.get("secondary_review_owner") or "report_group"),
            "status": secondary_status,
            "pending_runtime_rows": pending_rows,
            "receipt": dict(secondary_receipt),
        },
        "uat": {
            "status": str(uat.get("status") or "not_recorded"),
            "reviewed_reports": reviewed,
            "passed_reports": passed,
            "pass_rate_percent": rate,
            "required_pass_rate_percent": threshold,
            "minimum_reviewed_reports": minimum,
        },
        "blocking_reasons": reasons,
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _percent(numerator: int, denominator: int) -> float:
    return round(100.0 * numerator / denominator, 2) if denominator else 100.0


def _validate_manifest(project_root: Path) -> dict[str, Any]:
    path = project_root / "data/knowledge_bases/processed/knowledge_base_manifest.yaml"
    issues: list[dict[str, Any]] = []
    if not path.exists():
        return {
            "status": "FAIL",
            "path": str(path),
            "artifacts": [],
            "issues": [
                {
                    "code": "MISSING_BASE_MANIFEST",
                    "message": "base knowledge manifest is missing",
                }
            ],
        }
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if str(raw.get("schema_version") or "") != "1.0":
        issues.append(
            {
                "code": "INVALID_BASE_MANIFEST_SCHEMA",
                "message": "schema_version must be 1.0",
            }
        )
    artifacts: list[dict[str, Any]] = []
    for artifact_id, item in (raw.get("artifacts") or {}).items():
        if not isinstance(item, Mapping):
            issues.append(
                {
                    "code": "INVALID_BASE_ARTIFACT",
                    "message": f"{artifact_id} must be a mapping",
                }
            )
            continue
        declared_path = Path(str(item.get("path") or ""))
        resolved = (project_root / declared_path).resolve()
        try:
            resolved.relative_to(project_root.resolve())
        except ValueError:
            issues.append(
                {
                    "code": "BASE_ARTIFACT_PATH_ESCAPE",
                    "message": str(declared_path),
                }
            )
            continue
        expected = str(item.get("sha256") or "").lower()
        actual = _sha256(resolved) if resolved.exists() else ""
        source_refs = item.get("source_refs") or []
        artifact = {
            "artifact_id": str(artifact_id),
            "path": str(declared_path),
            "exists": resolved.exists(),
            "sha256_matches": bool(expected and actual == expected),
            "source_type": str(item.get("source_type") or ""),
            "source_refs": len(source_refs) if isinstance(source_refs, list) else 0,
            "evidence_as_of": str(item.get("evidence_as_of") or ""),
        }
        artifacts.append(artifact)
        if not artifact["exists"]:
            issues.append(
                {
                    "code": "MISSING_BASE_ARTIFACT",
                    "message": str(declared_path),
                }
            )
        elif not artifact["sha256_matches"]:
            issues.append(
                {
                    "code": "BASE_ARTIFACT_HASH_MISMATCH",
                    "message": str(declared_path),
                    "expected": expected,
                    "actual": actual,
                }
            )
        if not artifact["source_type"] or not artifact["source_refs"]:
            issues.append(
                {
                    "code": "INCOMPLETE_BASE_PROVENANCE",
                    "message": str(declared_path),
                }
            )
    return {
        "status": "PASS" if not issues else "FAIL",
        "path": str(path.relative_to(project_root)),
        "review_status": str((raw.get("policy") or {}).get("review_status") or ""),
        "artifacts": artifacts,
        "issues": issues,
    }


def _base_genes(project_root: Path) -> set[str]:
    path = project_root / "data/knowledge_bases/processed/gene_knowledge_db.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next(
        (name for name in workbook.sheetnames if "基因变异解析" in name), ""
    )
    if not sheet_name:
        return set()
    rows = workbook[sheet_name].iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows, ())]
    try:
        gene_index = headers.index("基因名称")
    except ValueError:
        gene_index = 0
    return {
        str(row[gene_index] or "").strip().upper()
        for row in rows
        if gene_index < len(row)
        and str(row[gene_index] or "").strip()
        and str(row[gene_index] or "").strip() != "基因"
    }


def _overlay_paths(package: Any) -> list[Path]:
    raw = getattr(package, "raw", None) or {}
    values: list[str] = []
    primary = raw.get("reviewed_part3_overlay")
    if primary:
        values.append(str(primary))
    additions = raw.get("reviewed_part3_overlay_additions") or []
    if isinstance(additions, str):
        additions = [additions]
    values.extend(str(value) for value in additions if str(value).strip())
    return [package._resolve_path(value).resolve() for value in values]


def _declared_genes(package: Any) -> set[str]:
    path = package.resolve_rule_file("knowledge_coverage")
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return {
        str(value).strip().upper()
        for value in raw.get("reportable_genes") or []
        if str(value).strip()
    }


def _validate_drug_rules(package: Any) -> dict[str, Any]:
    path = package.resolve_rule_file("drugs")
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    governance = raw.get("governance") or {}
    issues: list[dict[str, Any]] = []
    if str(governance.get("schema_version") or "") != "1.0":
        issues.append(
            {
                "code": "INVALID_DRUG_GOVERNANCE_SCHEMA",
                "message": f"{package.panel_id} drugs governance must be 1.0",
            }
        )
    policy = raw.get("targeted_drug_rules") or {}
    rows: list[dict[str, Any]] = []
    gene_overrides = policy.get("overrides") or {}
    for gene, rule in gene_overrides.items():
        if isinstance(rule, Mapping):
            rows.append({"gene": gene, **dict(rule)})
            issues.append(
                {
                    "code": "UNSCOPED_GENE_LEVEL_TARGETED_DRUG_OVERRIDE",
                    "message": (
                        f"{package.panel_id} gene-level targeted-drug override "
                        f"for {str(gene).upper()} has no variant/event selector"
                    ),
                    "gene": str(gene).upper(),
                }
            )
    rows.extend(
        dict(row)
        for row in policy.get("reviewed_variant_overrides") or []
        if isinstance(row, Mapping)
    )
    targeted = validate_knowledge_rows(
        raw,
        rows,
        panel_id=package.panel_id,
        kind="targeted_drug",
    )
    # A source-only inheritance declaration legitimately has no local targeted
    # rows or targeted defaults; its source panel is validated independently.
    source_panel_id = str(policy.get("source_panel_id") or "")
    if rows:
        issues.extend(targeted["issues"])
    elif source_panel_id:
        targeted["status"] = "INHERITED"

    approved_rows = [
        row for row in raw.get("approved_drug_rows") or [] if isinstance(row, Mapping)
    ]
    approved_statuses: Counter[str] = Counter()
    approved_source_rows = 0
    for index, row in enumerate(approved_rows, start=1):
        effective = effective_governance(raw, row, "approved_drug")
        approved_statuses[effective["status"]] += 1
        approved_source_rows += int(bool(effective["source_refs"]))
        if not row.get("drug") or not str(row.get("indication") or "").strip():
            issues.append(
                {
                    "code": "INCOMPLETE_APPROVED_DRUG_ROW",
                    "message": f"approved drug row {index} is incomplete",
                }
            )
        if not effective["source_refs"]:
            issues.append(
                {
                    "code": "MISSING_APPROVED_DRUG_SOURCE",
                    "message": f"approved drug row {index} has no source",
                }
            )
    return {
        "path": str(path),
        "status": "PASS" if not issues else "FAIL",
        "version": str(raw.get("version") or ""),
        "updated": str(raw.get("updated") or ""),
        "source_panel_id": source_panel_id,
        "targeted_rules": targeted,
        "approved_drug_rows": {
            "total_rows": len(approved_rows),
            "status_counts": dict(approved_statuses),
            "structured_source_rows": approved_source_rows,
            "structured_source_percent": _percent(
                approved_source_rows, len(approved_rows)
            ),
        },
        "issues": issues,
    }


def _panel_report(
    project_root: Path,
    panel_id: str,
    base_genes: set[str],
    clinical_registry: Mapping[str, Any],
) -> dict[str, Any]:
    package = PanelPackageLoader(project_root=project_root).load(panel_id)
    declared = _declared_genes(package)
    paths = _overlay_paths(package)
    overlay_reports: list[dict[str, Any]] = []
    runtime_overlay_genes: set[str] = set()
    all_gene_rows = 0
    all_drug_rows = 0
    gene_level_rows = 0
    variant_level_rows = 0
    event_level_drug_rows = 0
    status_counts: Counter[str] = Counter()
    structured_source_rows = 0
    evidence_level_rows = 0
    cancer_scope_rows = 0
    secondary_complete_rows = 0
    governance_total = 0
    issues: list[dict[str, Any]] = []

    for path in paths:
        origin_data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        origin_panel = str((origin_data.get("source") or {}).get("panel") or panel_id)
        overlay_report = load_and_validate_overlay(path, origin_panel)
        overlay_reports.append(overlay_report)
        issues.extend(overlay_report["issues"])
        for kind, section in (("gene", "gene_sections"), ("drug", "drug_sections")):
            rows = [row for row in origin_data.get(section) or [] if isinstance(row, Mapping)]
            for row in rows:
                effective = effective_governance(origin_data, row, kind)
                governance_total += 1
                status_counts[effective["status"]] += 1
                structured_source_rows += int(bool(effective["source_refs"]))
                evidence_level_rows += int(bool(effective["evidence_level"]))
                cancer_scope_rows += int(bool(effective["cancer_scope"]))
                secondary_complete_rows += int(
                    effective["secondary_review_status"]
                    in {"completed", "approved", "report_group_approved"}
                )
                if kind == "gene":
                    all_gene_rows += 1
                    variant = bool(str(row.get("c_hgvs") or "").strip() or str(row.get("p_hgvs") or "").strip())
                    variant_level_rows += int(variant)
                    gene_level_rows += int(not variant)
                    if effective["runtime_eligible"]:
                        runtime_overlay_genes.add(str(row.get("gene") or "").strip().upper())
                else:
                    all_drug_rows += 1
                    event_level_drug_rows += int(bool(str(row.get("applicability") or "").strip()))

    runtime_gene_union = base_genes | runtime_overlay_genes
    missing = declared - runtime_gene_union
    runtime_content = profile_panel_runtime_content(project_root, package, declared)
    drug_rules = _validate_drug_rules(package)
    issues.extend(drug_rules["issues"])
    drug_analysis_contracts = profile_panel_targeted_drug_contracts(
        project_root, package
    )
    issues.extend(drug_analysis_contracts["issues"])
    clinical_status_counts = Counter(status_counts)
    targeted_status_counts = (
        drug_rules.get("targeted_rules", {}).get("status_counts") or {}
    )
    if drug_rules.get("source_panel_id") and not targeted_status_counts:
        source_package = PanelPackageLoader(project_root=project_root).load(
            drug_rules["source_panel_id"]
        )
        targeted_status_counts = (
            _validate_drug_rules(source_package)
            .get("targeted_rules", {})
            .get("status_counts", {})
        )
    clinical_status_counts.update(targeted_status_counts)
    if missing:
        issues.append(
            {
                "code": "RUNTIME_GENE_COVERAGE_GAP",
                "message": f"{len(missing)} declared genes have no runtime explanation",
                "genes": sorted(missing),
            }
        )
    if runtime_content["missing_intro_genes"]:
        issues.append(
            {
                "code": "RUNTIME_GENE_INTRO_GAP",
                "message": (
                    f"{len(runtime_content['missing_intro_genes'])} declared genes "
                    "render an empty gene introduction"
                ),
                "genes": runtime_content["missing_intro_genes"],
            }
        )
    if runtime_content["missing_analysis_genes"]:
        issues.append(
            {
                "code": "RUNTIME_MUTATION_ANALYSIS_GAP",
                "message": (
                    f"{len(runtime_content['missing_analysis_genes'])} declared genes "
                    "render an empty mutation analysis for a previously unseen variant"
                ),
                "genes": runtime_content["missing_analysis_genes"],
            }
        )
    if runtime_content["missing_fixed_domain_genes"]:
        issues.append(
            {
                "code": "RUNTIME_FIXED_DOMAIN_GAP",
                "message": (
                    f"{len(runtime_content['missing_fixed_domain_genes'])} declared genes "
                    "render without gene-level fixed protein/domain content"
                ),
                "genes": runtime_content["missing_fixed_domain_genes"],
            }
        )
    if runtime_content["duplicate_fixed_domain_genes"]:
        issues.append(
            {
                "code": "RUNTIME_FIXED_DOMAIN_DUPLICATE",
                "message": (
                    f"{len(runtime_content['duplicate_fixed_domain_genes'])} declared genes "
                    "render more than one gene-level protein/domain statement"
                ),
                "genes": runtime_content["duplicate_fixed_domain_genes"],
            }
        )
    citation_integrity = runtime_content["citation_integrity"]
    if citation_integrity["unresolved_pmids"]:
        issues.append(
            {
                "code": "UNRESOLVED_RUNTIME_PMID",
                "message": (
                    f"{len(citation_integrity['unresolved_pmids'])} cited PMIDs "
                    "have no structured reference title"
                ),
                "identifiers": citation_integrity["unresolved_pmids"],
            }
        )
    if citation_integrity["unresolved_trials"]:
        issues.append(
            {
                "code": "UNRESOLVED_RUNTIME_TRIAL_REFERENCE",
                "message": (
                    f"{len(citation_integrity['unresolved_trials'])} cited trials "
                    "have no structured reference record"
                ),
                "identifiers": citation_integrity["unresolved_trials"],
            }
        )
    if status_counts.get("not_recorded", 0):
        issues.append(
            {
                "code": "UNSTANDARDIZED_REVIEW_STATUS",
                "message": f"{status_counts['not_recorded']} overlay rows are not recorded",
            }
        )

    modern_review_rows = (
        status_counts.get("approved_for_runtime", 0)
        + status_counts.get("provisional_runtime", 0)
        + status_counts.get("needs_review", 0)
        + status_counts.get("rejected", 0)
        + status_counts.get("superseded", 0)
    )
    secondary_receipt = _secondary_review_receipt(
        project_root, clinical_registry, panel_id
    )
    clinical_readiness = _clinical_readiness(
        clinical_registry,
        panel_id,
        clinical_status_counts,
        secondary_receipt,
    )
    if runtime_content["generic_fallback_count"]:
        clinical_readiness["status"] = "BLOCKED"
        clinical_readiness["blocking_reasons"].append(
            "generic_gene_fallback_requires_content_review"
        )
        clinical_readiness["content_depth"] = {
            "generic_fallback_count": runtime_content["generic_fallback_count"],
            "specific_explanation_percent": runtime_content[
                "specific_explanation_percent"
            ],
        }
    return {
        "panel_id": panel_id,
        "status": "PASS" if not issues else "FAIL",
        "overlay_files": overlay_reports,
        "drug_rules": drug_rules,
        "drug_analysis_contracts": drug_analysis_contracts,
        "runtime_content_quality": runtime_content,
        "clinical_release_readiness": clinical_readiness,
        "multidimensional_coverage": {
            "gene_explanation": {
                "total_genes": len(declared),
                "runtime_covered_genes": len(declared & runtime_gene_union),
                "percent": _percent(len(declared & runtime_gene_union), len(declared)),
                "missing_genes": sorted(missing),
            },
            "review_governance": {
                "total_overlay_rows": governance_total,
                "status_counts": dict(status_counts),
                "standardized_rows": governance_total - status_counts.get("not_recorded", 0),
                "standardized_percent": _percent(
                    governance_total - status_counts.get("not_recorded", 0),
                    governance_total,
                ),
                "modern_review_rows": modern_review_rows,
                "modern_review_percent": _percent(modern_review_rows, governance_total),
                "secondary_review_complete_rows": secondary_complete_rows,
                "secondary_review_complete_percent": _percent(
                    secondary_complete_rows, governance_total
                ),
            },
            "source_provenance": {
                "structured_source_rows": structured_source_rows,
                "structured_source_percent": _percent(
                    structured_source_rows, governance_total
                ),
                "evidence_level_rows": evidence_level_rows,
                "evidence_level_percent": _percent(
                    evidence_level_rows, governance_total
                ),
                "cancer_scope_rows": cancer_scope_rows,
                "cancer_scope_percent": _percent(cancer_scope_rows, governance_total),
            },
            "specificity": {
                "gene_rows": all_gene_rows,
                "gene_level_rows": gene_level_rows,
                "variant_level_rows": variant_level_rows,
                "drug_rows": all_drug_rows,
                "event_scoped_drug_rows": event_level_drug_rows,
            },
        },
        "issues": issues,
    }


def run_knowledge_release_gate(
    project_root: str | Path,
    *,
    panel_ids: Optional[Sequence[str]] = None,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    root = Path(project_root).resolve()
    manifest = _validate_manifest(root)
    base_genes = _base_genes(root)
    panel_statuses = _discover_panel_statuses(root)
    selected_panel_ids = tuple(panel_ids or ()) or tuple(
        panel_id
        for panel_id, status in panel_statuses.items()
        if status == "active"
    )
    clinical_registry = _clinical_release_registry(root)
    panels = [
        _panel_report(root, panel_id, base_genes, clinical_registry)
        for panel_id in selected_panel_ids
    ]
    issues = list(manifest["issues"])
    for panel in panels:
        issues.extend(panel["issues"])
    result = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "status": "PASS" if not issues else "FAIL",
        "base_manifest": manifest,
        "panels": panels,
        "clinical_release_registry": {
            "path": clinical_registry["path"],
            "policy": clinical_registry["policy"],
        },
        "non_blocking_panel_readiness": [
            {
                "panel_id": panel_id,
                "status": status,
                "release_blocking": False,
                "readiness": "NOT_PRODUCTION_ACTIVE",
                "warning": (
                    "draft/pilot panel is excluded from the production knowledge gate; "
                    "declare a panel-specific coverage denominator and reviewed overlay "
                    "before activation"
                ),
            }
            for panel_id, status in panel_statuses.items()
            if status != "active"
        ],
        "summary": {
            "panels_checked": len(panels),
            "panels_passed": sum(panel["status"] == "PASS" for panel in panels),
            "issues": len(issues),
        },
        "issues": issues,
    }
    if output_path:
        target = Path(output_path)
        if not target.is_absolute():
            target = root / target
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        result["report_file"] = str(target)
    return result
