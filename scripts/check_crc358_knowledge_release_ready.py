#!/usr/bin/env python3
"""Check whether the CRC358 knowledge base is ready for production release.

The check is intentionally strict: pending medical-review rows, structural
issues, PII risks, incomplete drug rows, or failed context retests make the
knowledge base not release-ready.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_crc358_knowledge_buildout import candidate_text_matches_context


DEFAULT_PROD_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_APPROVED_OVERLAY = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/reviewed_part3_knowledge.approved_from_review_batch9.yaml"
)
DEFAULT_REVIEW_WORKBOOK = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/CRC358_batch9_待医学审核合入包_20260614.xlsx"
)
DEFAULT_APPROVED_SUMMARY = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/approved_review_apply_summary_batch9.json"
)
DEFAULT_CONTEXT_RETEST = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/CRC358_batch9_pending_candidate_context_retest_20260614.xlsx"
)
DEFAULT_OUTPUT = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/CRC358_batch10_release_readiness_20260614.json"
)

APPROVED = {"通过", "修改后通过"}
PENDING = {"", "待医学审核", "待审核"}
FULL_REVIEW_SHEETS = ("新增gene完整审核", "新增drug完整审核")
PII_PATTERNS = {
    "sample_id": re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    "report_no": re.compile(r"报告编号"),
    "name_label": re.compile(r"姓名[:：]"),
    "sender": re.compile(r"送检者"),
    "date": re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
}
SECTION_LEAK_PATTERNS = {
    "reading_notes": re.compile(r"3\.\s*阅读说明"),
    "reference_instruction": re.compile(r"文中参考文献及临床试验编号说明"),
    "glossary_instruction": re.compile(r"医学及生物学常见名词说明"),
}


def clean(value: Any) -> str:
    return str(value or "").strip()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def gene_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (clean(row.get("gene")).upper(), clean(row.get("c_hgvs")), clean(row.get("p_hgvs")))


def drug_key(row: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        clean(row.get("gene")).upper(),
        clean(row.get("c_hgvs")),
        clean(row.get("p_hgvs")),
        clean(row.get("type") or "benefit"),
        clean(row.get("applicability")),
        clean(row.get("drug_name")),
        clean(row.get("header")),
    )


def pii_hits(text: str) -> list[str]:
    return [name for name, pattern in PII_PATTERNS.items() if pattern.search(text)]


def section_leak_hits(text: str) -> list[str]:
    return [name for name, pattern in SECTION_LEAK_PATTERNS.items() if pattern.search(text)]


def gene_context_mismatch(row: dict[str, Any]) -> list[str]:
    gene = clean(row.get("gene")).upper()
    hits: list[str] = []
    intro = clean(row.get("intro"))
    mutation_analysis = clean(row.get("mutation_analysis"))
    if intro and not candidate_text_matches_context(gene, "gene_intro", intro):
        hits.append("intro_first_gene_mismatch")
    if mutation_analysis and not candidate_text_matches_context(gene, "mutation_analysis", mutation_analysis):
        hits.append("mutation_analysis_first_gene_mismatch")
    return hits


def row_text(row: dict[str, Any]) -> str:
    fields = (
        "gene",
        "c_hgvs",
        "p_hgvs",
        "header",
        "drug_name",
        "intro",
        "mutation_analysis",
        "relation",
        "clinical",
    )
    return "\n".join(clean(row.get(field)) for field in fields)


def overlay_stats(path: Path) -> dict[str, Any]:
    data = load_yaml(path)
    gene_rows = data.get("gene_sections") or []
    drug_rows = data.get("drug_sections") or []
    gene_keys = [gene_key(row) for row in gene_rows]
    drug_keys = [drug_key(row) for row in drug_rows]
    duplicate_gene_keys = [key for key, count in Counter(gene_keys).items() if count > 1]
    duplicate_drug_keys = [key for key, count in Counter(drug_keys).items() if count > 1]
    gene_pii = [
        {"gene": clean(row.get("gene")).upper(), "key": list(gene_key(row)), "hits": pii_hits(row_text(row))}
        for row in gene_rows
        if pii_hits(row_text(row))
    ]
    gene_section_leaks = [
        {"gene": clean(row.get("gene")).upper(), "key": list(gene_key(row)), "hits": section_leak_hits(row_text(row))}
        for row in gene_rows
        if section_leak_hits(row_text(row))
    ]
    drug_section_leaks = [
        {"gene": clean(row.get("gene")).upper(), "key": list(drug_key(row)), "hits": section_leak_hits(row_text(row))}
        for row in drug_rows
        if section_leak_hits(row_text(row))
    ]
    drug_pii = [
        {"gene": clean(row.get("gene")).upper(), "key": list(drug_key(row)), "hits": pii_hits(row_text(row))}
        for row in drug_rows
        if pii_hits(row_text(row))
    ]
    mutation_without_c = [
        {"gene": clean(row.get("gene")).upper(), "p_hgvs": clean(row.get("p_hgvs"))}
        for row in gene_rows
        if clean(row.get("mutation_analysis")) and not clean(row.get("c_hgvs"))
    ]
    gene_context_mismatches = [
        {"gene": clean(row.get("gene")).upper(), "key": list(gene_key(row)), "hits": gene_context_mismatch(row)}
        for row in gene_rows
        if gene_context_mismatch(row)
    ]
    incomplete_drugs = [
        {
            "gene": clean(row.get("gene")).upper(),
            "c_hgvs": clean(row.get("c_hgvs")),
            "p_hgvs": clean(row.get("p_hgvs")),
            "drug_name": clean(row.get("drug_name")),
        }
        for row in drug_rows
        if not clean(row.get("drug_name")) or not clean(row.get("relation")) or not clean(row.get("clinical"))
    ]
    return {
        "path": str(path),
        "exists": path.exists(),
        "gene_sections": len(gene_rows),
        "drug_sections": len(drug_rows),
        "duplicate_gene_keys": len(duplicate_gene_keys),
        "duplicate_drug_keys": len(duplicate_drug_keys),
        "mutation_analysis_without_c_hgvs": len(mutation_without_c),
        "gene_context_mismatches": len(gene_context_mismatches),
        "incomplete_drug_sections": len(incomplete_drugs),
        "pii_hits": len(gene_pii) + len(drug_pii),
        "section_leak_hits": len(gene_section_leaks) + len(drug_section_leaks),
        "details": {
            "duplicate_gene_keys": [list(key) for key in duplicate_gene_keys[:20]],
            "duplicate_drug_keys": [list(key) for key in duplicate_drug_keys[:20]],
            "mutation_analysis_without_c_hgvs": mutation_without_c[:20],
            "gene_context_mismatches": gene_context_mismatches[:20],
            "incomplete_drug_sections": incomplete_drugs[:20],
            "pii_hits": (gene_pii + drug_pii)[:20],
            "section_leak_hits": (gene_section_leaks + drug_section_leaks)[:20],
        },
    }


def review_status_counts(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [sheet for sheet in FULL_REVIEW_SHEETS if sheet in wb.sheetnames]
    if not sheet_names:
        sheet_names = ["新增候选"] if "新增候选" in wb.sheetnames else []
    counts: Counter[str] = Counter()
    total = 0
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows or len(raw_rows[0]) == 1 and clean(raw_rows[0][0]) == "无数据":
            continue
        headers = [clean(value) for value in raw_rows[0]]
        if "review_status" not in headers:
            continue
        idx = headers.index("review_status")
        for raw in raw_rows[1:]:
            status = clean(raw[idx] if idx < len(raw) else "")
            counts[status or "空白"] += 1
            total += 1
    approved_rows = sum(counts.get(status, 0) for status in APPROVED)
    pending_rows = counts.get("空白", 0) + sum(counts.get(status, 0) for status in PENDING if status)
    rejected_rows = total - approved_rows - pending_rows
    return {
        "path": str(path),
        "exists": path.exists(),
        "sheets_checked": sheet_names,
        "total_review_rows": total,
        "approved_rows": approved_rows,
        "pending_rows": pending_rows,
        "rejected_or_deferred_rows": rejected_rows,
        "status_counts": dict(counts),
    }


def context_retest_stats(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    summary: dict[str, Any] = {}
    if "汇总" in wb.sheetnames:
        for key, value in wb["汇总"].iter_rows(values_only=True):
            summary[clean(key)] = value
    sample_rows = 0
    pass_rows = 0
    fail_rows = 0
    if "样本对比" in wb.sheetnames:
        rows = list(wb["样本对比"].iter_rows(values_only=True))
        if rows:
            headers = [clean(value) for value in rows[0]]
            status_idx = headers.index("status") if "status" in headers else -1
            for raw in rows[1:]:
                sample_rows += 1
                status = clean(raw[status_idx] if status_idx >= 0 and status_idx < len(raw) else "")
                if status == "PASS":
                    pass_rows += 1
                else:
                    fail_rows += 1
    return {
        "path": str(path),
        "exists": path.exists(),
        "summary": summary,
        "sample_rows": sample_rows,
        "pass_rows": pass_rows,
        "fail_rows": fail_rows,
    }


def evaluate(report: dict[str, Any]) -> list[dict[str, Any]]:
    review = report["review_workbook"]
    approved_summary = report["approved_summary"]
    approved_overlay = report["approved_overlay"]
    context = report["context_retest"]
    approved_skipped = approved_summary.get("approved_skipped")
    if approved_skipped is None:
        approved_skipped = approved_summary.get("skipped", 0)
    approved_consumed = approved_summary.get("added", 0) + approved_skipped
    checks = [
        {
            "name": "review_workbook_has_no_pending_rows",
            "passed": review["total_review_rows"] > 0 and review["pending_rows"] == 0,
            "detail": f"pending_rows={review['pending_rows']} total={review['total_review_rows']}",
        },
        {
            "name": "approved_apply_has_no_issues",
            "passed": approved_summary.get("issues") == 0,
            "detail": f"issues={approved_summary.get('issues')}",
        },
        {
            "name": "approved_apply_added_all_approved_rows",
            "passed": review["approved_rows"] == approved_consumed
            if review["pending_rows"] == 0
            else False,
            "detail": (
                f"approved_rows={review['approved_rows']} added={approved_summary.get('added')} "
                f"approved_skipped={approved_skipped} non_approved_skipped={approved_summary.get('non_approved_skipped', 'NA')}"
            ),
        },
        {
            "name": "approved_overlay_has_no_duplicate_keys",
            "passed": approved_overlay["duplicate_gene_keys"] == 0 and approved_overlay["duplicate_drug_keys"] == 0,
            "detail": f"gene_dup={approved_overlay['duplicate_gene_keys']} drug_dup={approved_overlay['duplicate_drug_keys']}",
        },
        {
            "name": "approved_overlay_has_no_pii_hits",
            "passed": approved_overlay["pii_hits"] == 0,
            "detail": f"pii_hits={approved_overlay['pii_hits']}",
        },
        {
            "name": "approved_overlay_has_no_section_leaks",
            "passed": approved_overlay["section_leak_hits"] == 0,
            "detail": f"section_leak_hits={approved_overlay['section_leak_hits']}",
        },
        {
            "name": "approved_overlay_has_no_gene_context_mismatches",
            "passed": approved_overlay["gene_context_mismatches"] == 0,
            "detail": f"gene_context_mismatches={approved_overlay['gene_context_mismatches']}",
        },
        {
            "name": "approved_overlay_has_complete_structured_rows",
            "passed": approved_overlay["mutation_analysis_without_c_hgvs"] == 0
            and approved_overlay["incomplete_drug_sections"] == 0,
            "detail": (
                f"mutation_without_c={approved_overlay['mutation_analysis_without_c_hgvs']} "
                f"incomplete_drugs={approved_overlay['incomplete_drug_sections']}"
            ),
        },
        {
            "name": "context_retest_all_samples_pass",
            "passed": context["sample_rows"] > 0 and context["fail_rows"] == 0 and context["pass_rows"] == context["sample_rows"],
            "detail": f"pass={context['pass_rows']} fail={context['fail_rows']} total={context['sample_rows']}",
        },
    ]
    return checks


def build_report(args: argparse.Namespace) -> dict[str, Any]:
    approved_summary = load_json(args.approved_summary)
    report = {
        "status": "not_evaluated",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "prod_overlay": overlay_stats(args.prod_overlay),
        "approved_overlay": overlay_stats(args.approved_overlay),
        "review_workbook": review_status_counts(args.review_workbook),
        "approved_summary": approved_summary,
        "context_retest": context_retest_stats(args.context_retest),
    }
    checks = evaluate(report)
    report["checks"] = checks
    report["status"] = "release_ready" if all(check["passed"] for check in checks) else "not_release_ready"
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--prod-overlay", type=Path, default=DEFAULT_PROD_OVERLAY)
    parser.add_argument("--approved-overlay", type=Path, default=DEFAULT_APPROVED_OVERLAY)
    parser.add_argument("--review-workbook", type=Path, default=DEFAULT_REVIEW_WORKBOOK)
    parser.add_argument("--approved-summary", type=Path, default=DEFAULT_APPROVED_SUMMARY)
    parser.add_argument("--context-retest", type=Path, default=DEFAULT_CONTEXT_RETEST)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--strict", action="store_true", help="exit non-zero when not release-ready")
    args = parser.parse_args()

    report = build_report(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"status={report['status']}")
    for check in report["checks"]:
        mark = "PASS" if check["passed"] else "FAIL"
        print(f"{mark} {check['name']} {check['detail']}")
    print(f"output={args.output}")
    return 1 if args.strict and report["status"] != "release_ready" else 0


if __name__ == "__main__":
    raise SystemExit(main())
