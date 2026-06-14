#!/usr/bin/env python3
"""Mine all-cancer final reports for CRC358 base-missing gene-intro support.

The generated overlay is review-only. It only includes gene-level introductions
with at least two historical final-report sources. Mutation-analysis paragraphs
from non-CRC reports are exported as reference material, not as CRC overlay rows.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_crc358_knowledge_buildout import (
    docx_paragraphs,
    extract_candidates_from_paragraphs,
    is_docx_candidate,
    stable_file_hash,
)
from scripts.prepare_crc358_gene_gap_review import has_pii, strip_section_leak


DEFAULT_CORPUS = Path("各癌种基因报告近年汇总")
DEFAULT_GAP_REVIEW = Path(
    "tmp/knowledge_buildout_after_batch7_gene_normalize_20260614/"
    "CRC358_batch6_基因简介与变异解析补库审核包_20260614.xlsx"
)
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch8_cross_cancer_gene_support_20260614")


def clean(value: Any) -> str:
    return str(value or "").strip()


def compact(text: str) -> str:
    return re.sub(r"\s+", "", clean(text))


def read_base_missing_genes(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "基础库仍缺" not in wb.sheetnames:
        raise ValueError("workbook missing sheet: 基础库仍缺")
    ws = wb["基础库仍缺"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    if "gene" not in headers:
        raise ValueError("sheet 基础库仍缺 missing column: gene")
    gene_idx = headers.index("gene")
    return sorted({clean(row[gene_idx]).upper() for row in rows[1:] if clean(row[gene_idx])})


def source_family(path: Path, corpus: Path) -> str:
    try:
        rel = path.relative_to(corpus)
        return rel.parts[0] if len(rel.parts) > 1 else ""
    except ValueError:
        return path.parent.name


def scan_all_cancer_candidates(corpus: Path, target_genes: set[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    inventory: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    for path in sorted(corpus.rglob("*.docx")):
        if not is_docx_candidate(path):
            continue
        family = source_family(path, corpus)
        source_hash = stable_file_hash(path)
        row = {
            "source_hash": source_hash,
            "source_family": family,
            "is_final": "终版" in path.name,
            "docx_valid": False,
            "paragraph_count": 0,
            "candidate_count": 0,
            "parse_error": "",
        }
        try:
            paragraphs = docx_paragraphs(path)
            row["docx_valid"] = True
            row["paragraph_count"] = len(paragraphs)
            if row["is_final"]:
                extracted = extract_candidates_from_paragraphs(paragraphs, source_hash, "all_cancer_final")
                for cand in extracted:
                    if cand.gene.upper() not in target_genes:
                        continue
                    if cand.content_type not in {"gene_intro", "mutation_analysis"}:
                        continue
                    text = strip_section_leak(cand.candidate_text)
                    if not text or has_pii(text):
                        continue
                    candidates.append(
                        {
                            "gene": cand.gene.upper(),
                            "content_type": cand.content_type,
                            "candidate_text": text,
                            "c_hgvs": cand.c_hgvs,
                            "p_hgvs": cand.p_hgvs,
                            "source_hash": source_hash,
                            "source_family": family,
                        }
                    )
                row["candidate_count"] = sum(1 for item in candidates if item["source_hash"] == source_hash)
        except Exception as exc:  # pragma: no cover - defensive per-document logging
            row["parse_error"] = f"{type(exc).__name__}: {exc}"
        inventory.append(row)
    return inventory, candidates


def build_intro_review(candidates: list[dict[str, Any]], *, min_sources: int = 2) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    intro_groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    mutation_reference: list[dict[str, Any]] = []
    for row in candidates:
        if row.get("content_type") == "gene_intro":
            intro_groups[(row["gene"], compact(row["candidate_text"]))].append(row)
        elif row.get("content_type") == "mutation_analysis":
            mutation_reference.append(
                {
                    "gene": row.get("gene", ""),
                    "c_hgvs": row.get("c_hgvs", ""),
                    "p_hgvs": row.get("p_hgvs", ""),
                    "source_family": row.get("source_family", ""),
                    "source_hash": row.get("source_hash", ""),
                    "candidate_text": row.get("candidate_text", ""),
                    "usage_policy": "参考材料；非CRC来源不得直接作为CRC reviewed变异解析",
                }
            )

    intro_rows: list[dict[str, Any]] = []
    single_source_rows: list[dict[str, Any]] = []
    for (gene, _), rows in sorted(intro_groups.items()):
        source_hashes = sorted({clean(row.get("source_hash")) for row in rows if clean(row.get("source_hash"))})
        families = sorted({clean(row.get("source_family")) for row in rows if clean(row.get("source_family"))})
        best_text = max((clean(row.get("candidate_text")) for row in rows), key=len)
        out = {
            "gene": gene,
            "source_count": len(source_hashes),
            "source_family_count": len(families),
            "source_families": "；".join(families),
            "source_hashes": "；".join(source_hashes),
            "recommended_intro": best_text,
            "review_status": "待医学审核",
            "review_notes": "",
        }
        if len(source_hashes) >= min_sources:
            out["machine_suggestion"] = "多份历史终版支持；可进入基因简介 pending overlay 供审核"
            intro_rows.append(out)
        else:
            out["machine_suggestion"] = "单来源历史终版；仅作补证据参考"
            single_source_rows.append(out)

    # Keep one best supported intro per gene for overlay preview.
    by_gene: dict[str, dict[str, Any]] = {}
    for row in intro_rows:
        current = by_gene.get(row["gene"])
        if current is None or (
            int(row["source_count"]),
            int(row["source_family_count"]),
            len(row["recommended_intro"]),
        ) > (
            int(current["source_count"]),
            int(current["source_family_count"]),
            len(current["recommended_intro"]),
        ):
            by_gene[row["gene"]] = row
    return sorted(by_gene.values(), key=lambda row: row["gene"]), single_source_rows, mutation_reference


def build_pending_overlay(intro_rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "source_type": "all_cancer_final_report_gene_intro_support",
            "status": "pending_medical_review",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "merge_policy": "review-only artifact; only repeated gene introductions are included",
        },
        "gene_sections": [
            {
                "gene": row["gene"],
                "c_hgvs": "",
                "p_hgvs": "",
                "intro": row["recommended_intro"],
                "mutation_analysis": "",
            }
            for row in intro_rows
        ],
    }


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_review_workbook(
    path: Path,
    target_genes: list[str],
    inventory: list[dict[str, Any]],
    intro_rows: list[dict[str, Any]],
    single_source_rows: list[dict[str, Any]],
    mutation_reference: list[dict[str, Any]],
    overlay: dict[str, Any],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(
        ws,
        [
            {"项目": "定位", "说明": "全癌种历史终版补证据包；不是生产入库文件。"},
            {"项目": "目标基因", "说明": "、".join(target_genes)},
            {"项目": "扫描终版DOCX", "说明": sum(1 for row in inventory if row.get("is_final"))},
            {"项目": "多来源简介待审", "说明": len(intro_rows)},
            {"项目": "单来源简介参考", "说明": len(single_source_rows)},
            {"项目": "变异解析参考", "说明": len(mutation_reference)},
            {"项目": "安全边界", "说明": "pending overlay 仅包含多来源 gene_intro；mutation_analysis 仅作参考。"},
        ],
    )
    ws = wb.create_sheet("多来源简介待审")
    write_sheet(ws, intro_rows)
    ws = wb.create_sheet("单来源简介参考")
    write_sheet(ws, single_source_rows)
    ws = wb.create_sheet("变异解析参考")
    write_sheet(ws, mutation_reference)
    ws = wb.create_sheet("待审overlay预览")
    write_sheet(ws, overlay.get("gene_sections") or [])
    ws = wb.create_sheet("扫描清单")
    write_sheet(ws, inventory)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--gap-review", type=Path, default=DEFAULT_GAP_REVIEW)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    target_genes = read_base_missing_genes(args.gap_review)
    inventory, candidates = scan_all_cancer_candidates(args.corpus, set(target_genes))
    intro_rows, single_source_rows, mutation_reference = build_intro_review(candidates)
    overlay = build_pending_overlay(intro_rows)

    review_xlsx = args.out_dir / "CRC358_batch8_全癌种历史终版基因简介补证据审核包_20260614.xlsx"
    overlay_yaml = args.out_dir / "reviewed_part3_cross_cancer_gene_intro_pending_review_batch8.yaml"
    write_review_workbook(
        review_xlsx,
        target_genes,
        inventory,
        intro_rows,
        single_source_rows,
        mutation_reference,
        overlay,
    )
    dump_yaml(overlay_yaml, overlay)

    print(f"target_genes={len(target_genes)} scanned_docx={len(inventory)} candidates={len(candidates)}")
    print(f"intro_review_genes={len(intro_rows)} single_source_intro={len(single_source_rows)} mutation_reference={len(mutation_reference)}")
    print(f"review_xlsx={review_xlsx}")
    print(f"pending_overlay={overlay_yaml}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
