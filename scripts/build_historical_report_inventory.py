#!/usr/bin/env python3
"""Build a de-identified inventory for historical final report corpora.

The workbook is a planning artifact for knowledge-base buildout. It deliberately
does not write patient names, sample IDs, full filenames, or raw report text.
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CORPUS = PROJECT_ROOT / "各癌种基因报告近年汇总"
DEFAULT_OUTPUT = PROJECT_ROOT / "tmp/knowledge_buildout/report_inventory.xlsx"

PRODUCT_TOKEN_RE = re.compile(
    r"(基因|分子分型|HRD|hrd|精准治疗|靶向|PD[-_\s]?L1|pd[-_\s]?l1|MSI|msi|TMB|tmb)"
)
SAMPLE_TOKEN_RE = re.compile(r"^(?:MLJY|mljy|[Ll][ZzWw]\d{4,}|[A-Z]{2,}\d{5,}.*)$")
PANEL_SIZE_RE = re.compile(r"(\d{1,4})\s*基因")
RESOURCE_PREFIX = "._"


@dataclass(frozen=True)
class ReportRecord:
    source_id: str
    content_hash: str
    cancer_dir: str
    extension: str
    product_family: str
    panel_size: int | None
    has_msi: bool
    has_tmb: bool
    has_pdl1: bool
    is_resource_file: bool
    parse_status: str
    paragraph_count: int
    table_count: int
    file_size_bytes: int
    relative_depth: int


def _text(value: object) -> str:
    return str(value or "").strip()


def normalize_product_family(value: str) -> str:
    text = _text(value)
    text = text.replace("＋", "+")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"PD[-_]?L1", "pd-l1", text, flags=re.IGNORECASE)
    text = re.sub(r"MSI", "msi", text, flags=re.IGNORECASE)
    text = re.sub(r"TMB", "tmb", text, flags=re.IGNORECASE)
    text = text.strip("-_ ")
    return text or "未识别产品族"


def infer_product_family(filename: str) -> str:
    """Infer report product family from a de-identified filename pattern.

    Historical names usually follow:
    patient-diagnosis-product-mljy-sample-final.docx
    Only the product token is returned; patient/sample tokens are discarded.
    """
    stem = Path(filename).stem
    if stem.startswith(RESOURCE_PREFIX):
        stem = stem[len(RESOURCE_PREFIX) :]
    stem = re.sub(r"终版\d*$|终版$|补充报告|修改版|已审核", "", stem, flags=re.IGNORECASE)
    # Protect PD-L1 before splitting filename parts on hyphen-like separators.
    stem = re.sub(r"PD[-_\s]?L1", "PDL1", stem, flags=re.IGNORECASE)
    parts = [
        part.strip()
        for part in re.split(r"[-－—–]", stem)
        if part and part.strip()
    ]
    candidates: list[str] = []
    for part in parts:
        clean = part.strip()
        if SAMPLE_TOKEN_RE.match(clean):
            continue
        if PRODUCT_TOKEN_RE.search(clean):
            candidates.append(clean)
    if candidates:
        return normalize_product_family(candidates[-1])

    return "未识别产品族"


def _stable_id(relative_path: Path) -> str:
    digest = hashlib.sha256(str(relative_path).encode("utf-8")).hexdigest()
    return f"rpt_{digest[:16]}"


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()[:16]


def _inspect_docx(path: Path) -> tuple[str, int, int]:
    try:
        with zipfile.ZipFile(path) as zf:
            try:
                xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
            except KeyError:
                return "missing_document_xml", 0, 0
    except zipfile.BadZipFile:
        return "bad_zip", 0, 0
    except Exception as exc:
        return f"read_error:{type(exc).__name__}", 0, 0
    return "parseable", xml.count("<w:p"), xml.count("<w:tbl")


def build_inventory(corpus_dir: Path) -> list[ReportRecord]:
    records: list[ReportRecord] = []
    corpus_dir = corpus_dir.resolve()
    for path in sorted(corpus_dir.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(corpus_dir)
        cancer_dir = rel.parts[0] if len(rel.parts) > 1 else "根目录"
        ext = path.suffix.lower().lstrip(".") or "no_ext"
        is_resource = path.name.startswith(RESOURCE_PREFIX)
        product_family = infer_product_family(path.name)
        panel_size_match = PANEL_SIZE_RE.search(product_family)
        panel_size = int(panel_size_match.group(1)) if panel_size_match else None
        product_lower = product_family.lower()

        if is_resource:
            parse_status = "resource_file"
            paragraph_count = 0
            table_count = 0
        elif ext == "docx":
            parse_status, paragraph_count, table_count = _inspect_docx(path)
        elif ext == "doc":
            parse_status = "legacy_doc_unsupported"
            paragraph_count = 0
            table_count = 0
        else:
            parse_status = "non_report_file"
            paragraph_count = 0
            table_count = 0

        records.append(
            ReportRecord(
                source_id=_stable_id(rel),
                content_hash=_file_hash(path),
                cancer_dir=cancer_dir,
                extension=ext,
                product_family=product_family,
                panel_size=panel_size,
                has_msi="msi" in product_lower,
                has_tmb="tmb" in product_lower,
                has_pdl1="pd-l1" in product_lower or "pdl1" in product_lower,
                is_resource_file=is_resource,
                parse_status=parse_status,
                paragraph_count=paragraph_count,
                table_count=table_count,
                file_size_bytes=path.stat().st_size,
                relative_depth=len(rel.parts),
            )
        )
    return records


def _append_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, 1):
        width = len(str(header)) + 2
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                width = max(width, min(60, len(str(item.value or "")) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10), 60)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _counter_rows(counter: Counter[str]) -> list[list[Any]]:
    return [[key, count] for key, count in counter.most_common()]


def write_workbook(records: list[ReportRecord], output: Path, corpus_dir: Path) -> dict[str, Any]:
    wb = Workbook()
    wb.remove(wb.active)

    total_files = len(records)
    valid_docx = [r for r in records if r.extension == "docx" and not r.is_resource_file]
    parseable_docx = [r for r in valid_docx if r.parse_status == "parseable"]
    resource_files = [r for r in records if r.is_resource_file]
    legacy_doc = [r for r in records if r.extension == "doc" and not r.is_resource_file]
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    ws = wb.create_sheet("汇总")
    _append_table(
        ws,
        ["指标", "值"],
        [
            ["generated_at_utc", generated_at],
            ["corpus_label", corpus_dir.name],
            ["total_files", total_files],
            ["valid_docx_non_resource", len(valid_docx)],
            ["parseable_docx", len(parseable_docx)],
            ["resource_files_ignored", len(resource_files)],
            ["legacy_doc_unsupported", len(legacy_doc)],
            ["cancer_dirs", len({r.cancer_dir for r in records if r.cancer_dir != "根目录"})],
            ["product_families", len({r.product_family for r in valid_docx})],
            ["privacy_note", "不含真实患者姓名、样本号、完整文件名或报告原文"],
        ],
    )

    cancer_rows: list[list[Any]] = []
    by_cancer: dict[str, list[ReportRecord]] = defaultdict(list)
    for record in records:
        by_cancer[record.cancer_dir].append(record)
    for cancer, items in sorted(
        by_cancer.items(),
        key=lambda item: (
            -sum(1 for r in item[1] if r.extension == "docx" and not r.is_resource_file),
            item[0],
        ),
    ):
        cancer_valid_docx = [r for r in items if r.extension == "docx" and not r.is_resource_file]
        cancer_rows.append(
            [
                cancer,
                len(cancer_valid_docx),
                sum(1 for r in cancer_valid_docx if r.parse_status == "parseable"),
                sum(1 for r in items if r.is_resource_file),
                sum(1 for r in items if r.extension == "doc" and not r.is_resource_file),
                len({r.product_family for r in cancer_valid_docx}),
                round(sum(r.file_size_bytes for r in cancer_valid_docx) / 1024 / 1024, 2),
            ]
        )
    _append_table(
        wb.create_sheet("癌种目录分布"),
        ["癌种目录", "有效DOCX", "可解析DOCX", "资源文件", "旧DOC", "产品族数", "有效DOCX大小_MB"],
        cancer_rows,
    )

    product_group: dict[tuple[str, str], list[ReportRecord]] = defaultdict(list)
    for record in valid_docx:
        product_group[(record.cancer_dir, record.product_family)].append(record)
    product_rows: list[list[Any]] = []
    for (cancer, product), items in sorted(
        product_group.items(), key=lambda item: (-len(item[1]), item[0][0], item[0][1])
    ):
        sizes = [r.panel_size for r in items if r.panel_size is not None]
        product_rows.append(
            [
                cancer,
                product,
                len(items),
                sum(1 for r in items if r.parse_status == "parseable"),
                sizes[0] if sizes else "",
                "是" if any(r.has_msi for r in items) else "否",
                "是" if any(r.has_tmb for r in items) else "否",
                "是" if any(r.has_pdl1 for r in items) else "否",
                "大panel" if sizes and sizes[0] >= 300 else ("小panel" if sizes else "未知"),
            ]
        )
    _append_table(
        wb.create_sheet("产品族分布"),
        ["癌种目录", "产品族", "有效DOCX", "可解析DOCX", "基因数", "含MSI", "含TMB", "含PD-L1", "规模分层"],
        product_rows,
    )

    parse_rows = _counter_rows(Counter(r.parse_status for r in records))
    _append_table(wb.create_sheet("解析状态"), ["解析状态", "文件数"], parse_rows)

    duplicate_rows: list[list[Any]] = []
    by_hash: dict[str, list[ReportRecord]] = defaultdict(list)
    for record in valid_docx:
        by_hash[record.content_hash].append(record)
    for content_hash, items in sorted(by_hash.items(), key=lambda item: (-len(item[1]), item[0])):
        if len(items) < 2:
            continue
        duplicate_rows.append(
            [
                content_hash,
                len(items),
                "、".join(sorted({r.cancer_dir for r in items})),
                "、".join(sorted({r.product_family for r in items})[:5]),
            ]
        )
    _append_table(
        wb.create_sheet("疑似重复"),
        ["content_hash", "文件数", "癌种目录", "产品族_前5"],
        duplicate_rows or [["无", 0, "", ""]],
    )

    index_rows = [
        [
            r.source_id,
            r.content_hash,
            r.cancer_dir,
            r.extension,
            r.product_family,
            r.panel_size or "",
            "是" if r.has_msi else "否",
            "是" if r.has_tmb else "否",
            "是" if r.has_pdl1 else "否",
            "是" if r.is_resource_file else "否",
            r.parse_status,
            r.paragraph_count,
            r.table_count,
            r.file_size_bytes,
            r.relative_depth,
        ]
        for r in records
    ]
    _append_table(
        wb.create_sheet("去标识逐报告索引"),
        [
            "source_id",
            "content_hash",
            "癌种目录",
            "扩展名",
            "产品族",
            "基因数",
            "含MSI",
            "含TMB",
            "含PD-L1",
            "资源文件",
            "解析状态",
            "段落数",
            "表格数",
            "文件大小_bytes",
            "相对路径层级",
        ],
        index_rows,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "output": str(output),
        "total_files": total_files,
        "valid_docx_non_resource": len(valid_docx),
        "parseable_docx": len(parseable_docx),
        "resource_files_ignored": len(resource_files),
        "legacy_doc_unsupported": len(legacy_doc),
        "product_families": len({r.product_family for r in valid_docx}),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.corpus.exists():
        parser.error(f"corpus not found: {args.corpus}")
    records = build_inventory(args.corpus)
    result = write_workbook(records, args.output, args.corpus)
    for key, value in result.items():
        print(f"{key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
