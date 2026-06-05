#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Harvest reviewed CRC legacy knowledge candidates for Part3 overlay review.

The legacy workbook is treated as a candidate source only. This script does not
modify production YAML; report-team approval is required before promotion.
"""

from __future__ import annotations

import argparse
import os
import re
from collections import Counter
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MATRIX = PROJECT_ROOT / "CRC358知识库覆盖矩阵_20260605.xlsx"
DEFAULT_REVIEWED_YAML = (
    PROJECT_ROOT / "panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml"
)
DEFAULT_OUTPUT = PROJECT_ROOT / "CRC358旧肠癌知识库补库候选_20260605.xlsx"


GENE_RE = re.compile(r"^[A-Z0-9][A-Z0-9.-]{1,20}$")


def _text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ").strip()
    return "" if value.lower() == "nan" else value


def _is_gene(value: Any) -> bool:
    text = _text(value).upper()
    if not GENE_RE.match(text):
        return False
    return text not in {"GENE", "NULL", "NONE"}


def _load_current_reviewed(path: Path) -> tuple[set[str], set[str]]:
    if not path.exists():
        return set(), set()
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    gene_level: set[str] = set()
    site_level: set[str] = set()
    for item in data.get("gene_sections") or []:
        gene = _text(item.get("gene")).upper()
        if not gene:
            continue
        if _text(item.get("c_hgvs")) or _text(item.get("p_hgvs")):
            site_level.add(gene)
        else:
            gene_level.add(gene)
    return gene_level, site_level


def _load_matrix_levels(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    if "CRC358覆盖矩阵" not in wb.sheetnames:
        return {}
    ws = wb["CRC358覆盖矩阵"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        gene_idx = headers.index("基因")
        level_idx = headers.index("Part3覆盖级别")
    except ValueError:
        return {}
    levels: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gene = _text(row[gene_idx]).upper()
        if gene:
            levels[gene] = _text(row[level_idx])
    return levels


def _load_legacy_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "基因变异解析"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"legacy workbook missing sheet: {sheet_name}")
    ws = wb[sheet_name]

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        gene = _text(row[0]).upper() if row else ""
        if not _is_gene(gene):
            continue

        intro = _text(row[1] if len(row) > 1 else "")
        structure = _text(row[4] if len(row) > 4 else "")
        function_hint = _text(row[5] if len(row) > 5 else "")
        crc_analysis = _text(row[6] if len(row) > 6 else "")
        tail_default = _text(row[9] if len(row) > 9 else "")

        if not any([intro, structure, function_hint, crc_analysis]):
            continue

        rows.append(
            {
                "gene": gene,
                "legacy_intro": intro,
                "legacy_structure_or_variant_desc": structure,
                "legacy_function_hint": function_hint,
                "legacy_crc_analysis": crc_analysis,
                "legacy_tail_default": tail_default,
            }
        )
    return rows


def _suggest_action(
    gene: str,
    level: str,
    gene_level: set[str],
    site_level: set[str],
) -> str:
    if gene in gene_level:
        return "当前已有基因级内容；可对照旧库检查是否需要人工更新"
    if gene in site_level:
        return "当前已有位点级内容；旧库可作为泛化补充候选"
    if level.startswith("E"):
        return "优先补基因级兜底内容"
    if level.startswith("D"):
        return "建议用旧库内容升级基因级内容"
    if level.startswith("C"):
        return "已有基础库拼接；可用旧库提升成审核版"
    return "低优先级维护"


def _risk_rank(level: str, gene: str, gene_level: set[str], site_level: set[str]) -> int:
    if gene not in gene_level and gene not in site_level:
        if level.startswith("E"):
            return 1
        if level.startswith("D"):
            return 2
        if level.startswith("C"):
            return 3
    return 4


def _write_sheet_header(ws, headers: list[str], fill: str = "00B8C4") -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = 12
        for cell in column_cells:
            value = _text(cell.value)
            if value:
                width = max(width, min(max_width, len(value) + 2))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_candidates(
    legacy_db: Path,
    matrix: Path,
    reviewed_yaml: Path,
    output: Path,
) -> dict[str, Any]:
    legacy_rows = _load_legacy_rows(legacy_db)
    matrix_levels = _load_matrix_levels(matrix)
    gene_level, site_level = _load_current_reviewed(reviewed_yaml)

    enriched: list[dict[str, Any]] = []
    for item in legacy_rows:
        gene = item["gene"]
        level = matrix_levels.get(gene, "不在CRC358矩阵")
        enriched.append(
            {
                "优先级排序": _risk_rank(level, gene, gene_level, site_level),
                "基因": gene,
                "当前Part3覆盖级别": level,
                "当前是否已有基因级审核": "是" if gene in gene_level else "否",
                "当前是否已有位点级审核": "是" if gene in site_level else "否",
                "建议动作": _suggest_action(gene, level, gene_level, site_level),
                "旧库基因简介": item["legacy_intro"],
                "旧库结构域/变异说明": item["legacy_structure_or_variant_desc"],
                "旧库功能提示": item["legacy_function_hint"],
                "旧库肠癌解析": item["legacy_crc_analysis"],
                "旧库兜底尾句": item["legacy_tail_default"],
                "审核结论": "",
                "审核备注": "",
            }
        )

    enriched.sort(key=lambda x: (x["优先级排序"], x["基因"]))

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    counts = Counter(item["当前Part3覆盖级别"] for item in enriched)
    summary_rows = [
        ("指标", "结果"),
        ("旧库来源", str(legacy_db)),
        ("候选基因数", len(enriched)),
        ("旧库中已具备基因级审核内容", sum(1 for x in enriched if x["当前是否已有基因级审核"] == "是")),
        ("旧库中仅有位点级覆盖", sum(1 for x in enriched if x["当前是否已有位点级审核"] == "是" and x["当前是否已有基因级审核"] == "否")),
        ("旧库可帮助升级的基因数", sum(1 for x in enriched if x["优先级排序"] < 4)),
        ("结论", "旧库不是358全量库，但可作为重点基因二轮补库候选；需报告组审核后再入库。"),
    ]
    for row in summary_rows:
        ws_summary.append(row)
    for level, count in counts.most_common():
        ws_summary.append((level, count))
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"].font = Font(bold=True)
    _autosize(ws_summary, max_width=90)

    ws = wb.create_sheet("旧库补库候选")
    headers = list(enriched[0].keys()) if enriched else [
        "优先级排序",
        "基因",
        "当前Part3覆盖级别",
        "建议动作",
        "审核结论",
        "审核备注",
    ]
    _write_sheet_header(ws, headers)
    for item in enriched:
        ws.append([item.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)

    ws_note = wb.create_sheet("审核说明")
    notes = [
        ("字段", "说明"),
        ("审核结论", "报告组填写：通过 / 修改后通过 / 不通过。只有通过内容才可后续入 production YAML。"),
        ("旧库基因简介", "可作为 Part3 基因简介候选。"),
        ("旧库结构域/变异说明", "多为基因级结构域介绍，不应原样替代具体位点解析；需要结合变异类型自动拼接或人工改写。"),
        ("旧库功能提示", "包含 OncoKB/JAXCKB 类泛化提示，需确认适用范围，不能无条件套所有 III 类位点。"),
        ("旧库肠癌解析", "肠癌相关背景证据，适合补充基因级解析。"),
        ("安全口径", "本表只做候选采集，不修改报告生成逻辑，也不代表医学审核通过。"),
    ]
    for row in notes:
        ws_note.append(row)
    _autosize(ws_note, max_width=100)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "legacy_gene_candidates": len(enriched),
        "upgradable_candidates": sum(1 for x in enriched if x["优先级排序"] < 4),
        "coverage_counts": dict(counts),
        "output": str(output),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--legacy-db",
        type=Path,
        default=None,
        help="旧肠癌知识库 Excel 路径；也可用 CRC_LEGACY_DB_PATH 环境变量提供。",
    )
    parser.add_argument("--matrix", type=Path, default=DEFAULT_MATRIX)
    parser.add_argument("--reviewed-yaml", type=Path, default=DEFAULT_REVIEWED_YAML)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    legacy_db = args.legacy_db
    if legacy_db is None and os.environ.get("CRC_LEGACY_DB_PATH"):
        legacy_db = Path(os.environ["CRC_LEGACY_DB_PATH"])
    if legacy_db is None:
        parser.error("必须提供 --legacy-db 或设置 CRC_LEGACY_DB_PATH")

    result = build_candidates(
        legacy_db=legacy_db,
        matrix=args.matrix,
        reviewed_yaml=args.reviewed_yaml,
        output=args.output,
    )
    print(yaml.safe_dump(result, allow_unicode=True, sort_keys=False))


if __name__ == "__main__":
    main()
