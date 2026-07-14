#!/usr/bin/env python3
"""Prepare a CRC358 knowledge-base step-1 optimization review package.

This script produces review artifacts only. It does not modify the production
CRC358 reviewed Part3 overlay. The package is intended to close clear coverage
gaps first, while giving the report/medical reviewers a short list of
low-information entries to approve, revise, or leave unchanged.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_OUT_DIR = Path("tmp/knowledge_optimization_step1_20260621")

MISSING_GENE_CANDIDATES: list[dict[str, str]] = [
    {
        "gene": "CHD2",
        "intro": (
            "CHD2基因编码染色质结构调控相关的ATP依赖性染色质重塑蛋白，"
            "属于CHD家族，含有chromodomain和SNF2相关ATPase/helicase结构域，"
            "参与染色质开放状态、转录调控和基因组稳定性维持。"
        ),
        "mutation_analysis": (
            "CHD2变异可能影响染色质重塑和转录调控过程。当前CHD2在结直肠癌中的"
            "位点级用药或预后证据有限，建议作为染色质重塑相关长尾基因谨慎解释；"
            "若检出无义、移码、剪接等可能功能缺失型变异，应结合变异等级、等位基因频率、"
            "共突变背景和临床资料综合判断。"
        ),
        "source_note": "NCBI Gene 1106；MedlinePlus Genetics CHD2；候选中文表述需医学审核。",
        "source_urls": "https://www.ncbi.nlm.nih.gov/datasets/gene/1106/ | https://medlineplus.gov/genetics/gene/chd2/",
    },
    {
        "gene": "HIST1H3B",
        "intro": (
            "HIST1H3B（现行命名中常对应H3C2）编码复制依赖性组蛋白H3家族成员，"
            "是核小体核心组分，参与DNA包装为染色质，并影响转录、DNA复制、DNA修复和染色体稳定性。"
        ),
        "mutation_analysis": (
            "HIST1H3B变异可能影响组蛋白H3相关染色质结构和表观遗传调控。当前该基因在"
            "结直肠癌中的明确位点级临床证据有限，需结合变异是否位于已知组蛋白功能位点、"
            "变异等级及肿瘤分子背景综合判断。"
        ),
        "source_note": "NCBI Gene 8358/H3C2；Human Protein Atlas/UniProt function summary；候选中文表述需医学审核。",
        "source_urls": "https://www.ncbi.nlm.nih.gov/gene/8358 | https://www.proteinatlas.org/ENSG00000274267-HIST1H3B",
    },
    {
        "gene": "HLA-DPA1",
        "intro": (
            "HLA-DPA1编码MHC II类DP分子的α链，与DPβ链形成膜结合异二聚体，"
            "主要在抗原递呈细胞中表达，参与外源性抗原肽递呈和CD4+T细胞免疫应答。"
        ),
        "mutation_analysis": (
            "HLA-DPA1异常可能与抗原递呈能力和肿瘤免疫微环境相关。单个HLA-DPA1体细胞变异"
            "通常不足以直接给出用药结论，建议结合MSI/TMB、HLA表达或缺失、免疫相关共变异及"
            "证据等级综合评估。"
        ),
        "source_note": "NCBI Gene/Datasets 3113；RefSeq/HLA class II function summary；候选中文表述需医学审核。",
        "source_urls": "https://www.ncbi.nlm.nih.gov/datasets/gene/3113/ | https://www.ncbi.nlm.nih.gov/gene/3113",
    },
    {
        "gene": "WDR90",
        "intro": (
            "WDR90编码WD重复结构域蛋白，公开资料提示其与中心粒结构、纤毛发生和微管相关结构维持有关。"
            "WD重复蛋白常作为蛋白复合体装配或信号调控的支架蛋白。"
        ),
        "mutation_analysis": (
            "WDR90在结直肠癌中的明确临床证据有限。检出WDR90变异时，建议以长尾基因谨慎解释，"
            "重点结合变异类型、证据等级、是否影响关键结构域，以及是否存在其他明确驱动或用药相关变异"
            "综合判断。"
        ),
        "source_note": "UniProt Q96KV7；GeneCards/Harmonizome function summaries；候选中文表述需医学审核。",
        "source_urls": "https://www.uniprot.org/uniprotkb/Q96KV7/entry | https://www.genecards.org/card/WDR90",
    },
    {
        "gene": "ZNF703",
        "intro": (
            "ZNF703编码锌指蛋白703，属于转录调控相关蛋白，公开资料提示其与细胞黏附、"
            "迁移和增殖调控有关，并在部分肿瘤中作为扩增或过表达相关基因受到关注。"
        ),
        "mutation_analysis": (
            "ZNF703异常可能影响转录调控、细胞增殖迁移及PI3K/AKT/mTOR等相关信号背景。"
            "当前ZNF703在结直肠癌中的明确位点级临床证据有限，建议结合变异形式"
            "（扩增、错义、截短等）、证据等级和癌种背景综合判断。"
        ),
        "source_note": "UniProt Q9H7S9；ZNF703 cancer review；候选中文表述需医学审核。",
        "source_urls": "https://www.uniprot.org/uniprotkb/Q9H7S9/entry | https://pmc.ncbi.nlm.nih.gov/articles/PMC10559930/",
    },
]

LOW_INFO_PATTERNS = [
    r"具体临床意义需结合",
    r"具体临床价值需结合",
    r"潜在意义",
    r"可能影响",
    r"可能与疾病",
    r"综合判断",
    r"肿瘤发生发展",
    r"证据等级",
    r"癌种背景",
    r"当前研究",
    r"相关治疗策略",
]

GENE_LEVEL_REFINEMENTS: dict[str, str] = {
    "FANCA": (
        "FANCA属于Fanconi贫血/BRCA DNA链间交联修复通路核心成员，参与FA核心复合体形成、"
        "FANCD2/FANCI单泛素化及同源重组修复过程。FANCA功能缺失型变异可能提示DNA损伤修复缺陷；"
        "但单个错义变异是否具有功能影响需结合变异等级、功能证据和肿瘤类型判断，不能直接等同于"
        "明确PARP抑制剂获益标志。"
    ),
    "FANCD2": (
        "FANCD2是Fanconi贫血通路的关键效应蛋白，单泛素化后的FANCD2参与DNA链间交联损伤修复、"
        "复制叉稳定和同源重组修复调控。FANCD2功能缺失型变异可能提示DDR通路异常；普通错义变异"
        "需结合位点、功能证据和证据等级评估，避免过度外推为明确用药依据。"
    ),
    "RAD50": (
        "RAD50与MRE11、NBN组成MRN复合体，参与DNA双链断裂识别、末端处理和ATM检查点激活。"
        "RAD50功能缺失型变异可能导致DNA双链断裂修复能力下降，并可作为DDR通路评估的一部分；"
        "其在结直肠癌中的用药意义需结合HRD背景、变异致病性和其他HRR基因状态共同判断。"
    ),
    "BAP1": (
        "BAP1编码去泛素化酶，参与染色质调控、DNA损伤应答、细胞周期和细胞死亡相关过程。"
        "BAP1失活在部分肿瘤中具有抑癌基因意义，胚系异常还与BAP1肿瘤易感综合征相关；"
        "在结直肠癌报告中建议按DDR/染色质调控相关基因解释，并结合变异类型、结构域位置和证据等级判断。"
    ),
    "NOTCH1": (
        "NOTCH1是Notch信号通路受体，参与细胞命运决定、分化、增殖和肿瘤微环境调控。"
        "NOTCH1在不同肿瘤中可表现为促癌或抑癌相关作用；因此检出NOTCH1变异时应结合变异是否导致"
        "功能获得或功能缺失、所在结构域、证据等级和癌种背景评估，避免仅按基因名给出结论。"
    ),
    "TSC1": (
        "TSC1编码Hamartin蛋白，与TSC2形成复合物抑制Rheb-mTORC1信号。无义、移码、关键剪接等"
        "功能缺失型TSC1变异可能导致mTOR通路异常激活，并与mTOR通路抑制剂相关评估有关；"
        "普通错义变异需结合功能证据和变异等级判断，不应默认触发同等用药解释。"
    ),
    "FANCM": (
        "FANCM参与Fanconi贫血通路激活、复制叉稳定和DNA链间交联损伤应答。FANCM功能缺失型变异"
        "可作为DDR异常的候选证据；但其在结直肠癌中的直接用药证据有限，需结合HRD背景、变异类型"
        "和其他HRR基因状态综合评估。"
    ),
    "HRAS": (
        "HRAS属于RAS小GTP酶家族，激活性热点变异可持续激活MAPK/PI3K等下游通路。"
        "HRAS在结直肠癌中相对少见；若检出明确激活热点，可提示RAS通路激活并影响抗EGFR治疗相关判断，"
        "非热点或证据不足位点需按变异等级和药物提示表谨慎解释。"
    ),
    "PALB2": (
        "PALB2连接BRCA1/BRCA2并参与同源重组修复复合体装配。PALB2功能缺失型变异可能提示HRR缺陷，"
        "与PARP抑制剂或铂类敏感性评估相关；在结直肠癌中仍需结合胚系/体细胞来源、致病性等级、HRD背景"
        "和临床适应证综合判断。"
    ),
    "PBRM1": (
        "PBRM1是PBAF型SWI/SNF染色质重塑复合体成员，参与转录调控、染色质可及性和DNA损伤应答。"
        "PBRM1失活在肾透明细胞癌等肿瘤中研究较多；在结直肠癌中应作为染色质重塑相关变异解释，"
        "免疫治疗或预后意义需结合癌种证据和变异功能判断。"
    ),
    "PTCH1": (
        "PTCH1是Hedgehog通路受体并对SMO具有抑制作用，功能缺失可导致Hedgehog信号异常激活。"
        "PTCH1在部分肿瘤综合征和特定癌种中具有明确意义；在结直肠癌中检出时需结合变异是否功能缺失、"
        "是否伴随Hedgehog通路其他异常及证据等级判断。"
    ),
    "RAD51C": (
        "RAD51C属于RAD51旁系同源重组修复基因，参与DNA双链断裂修复和复制叉保护。"
        "RAD51C功能缺失型变异可能提示HRR缺陷；在CRC报告中应结合变异致病性、HRD背景、胚系/体细胞来源"
        "和药物适应证判断其临床意义。"
    ),
    "RAD51D": (
        "RAD51D参与同源重组修复和RAD51复合体功能维持。RAD51D功能缺失型变异可能提示HRR缺陷；"
        "在CRC报告中需结合变异致病性、HRD背景、胚系/体细胞来源和药物适应证判断，避免把任意错义变异"
        "直接解释为明确获益证据。"
    ),
    "WRN": (
        "WRN编码RecQ家族DNA/RNA解旋酶，参与DNA复制、重组、修复、端粒维持和复制压力应答。"
        "WRN缺陷与基因组不稳定相关，MSI肿瘤中的WRN依赖性也有研究关注；在CRC报告中应结合MSI状态、"
        "变异类型和功能证据综合判断。"
    ),
}

PII_PATTERNS = {
    "sample_id": re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    "report_no": re.compile(r"报告编号"),
    "name_label": re.compile(r"姓名[:：]"),
    "date": re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
}


def clean(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def row_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (
        clean(row.get("gene")).upper(),
        clean(row.get("c_hgvs")),
        clean(row.get("p_hgvs")),
    )


def score_low_info(row: dict[str, Any]) -> tuple[str, list[str]]:
    text = f"{clean(row.get('intro'))}\n{clean(row.get('mutation_analysis'))}"
    hits = [pattern for pattern in LOW_INFO_PATTERNS if re.search(pattern, text)]
    if len(hits) >= 5:
        priority = "P1"
    elif len(hits) >= 4 or (len(hits) >= 3 and clean(row.get("c_hgvs"))):
        priority = "P2"
    elif len(hits) >= 3:
        priority = "P3"
    else:
        priority = ""
    return priority, hits


def build_low_info_rows(gene_sections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in gene_sections:
        priority, hits = score_low_info(row)
        if priority not in {"P1", "P2"}:
            continue
        gene = clean(row.get("gene")).upper()
        c_hgvs = clean(row.get("c_hgvs"))
        current_analysis = clean(row.get("mutation_analysis"))
        candidate = GENE_LEVEL_REFINEMENTS.get(gene, "")
        if c_hgvs:
            action = "位点级条目：先核对，不建议仅因兜底语句命中而自动替换"
            candidate = ""
        elif candidate:
            action = "可作为替换候选；需医学审核后再合入"
        else:
            action = "进入优先审核清单；需人工补充更具体的证据或确认维持现状"
        rows.append(
            {
                "priority": priority,
                "hit_count": len(hits),
                "hit_patterns": "；".join(hits),
                "gene": gene,
                "c_hgvs": c_hgvs,
                "p_hgvs": clean(row.get("p_hgvs")),
                "current_intro": clean(row.get("intro")),
                "current_mutation_analysis": current_analysis,
                "candidate_mutation_analysis": candidate,
                "recommended_action": action,
                "review_status": "待医学审核",
                "reviewed_mutation_analysis": "",
                "review_notes": "",
            }
        )
    rows.sort(key=lambda item: (item["priority"], -int(item["hit_count"]), item["gene"], item["c_hgvs"]))
    return rows


def build_phgvs_rows(gene_sections: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in gene_sections:
        if clean(row.get("c_hgvs")) and not clean(row.get("p_hgvs")):
            rows.append(
                {
                    "gene": clean(row.get("gene")).upper(),
                    "c_hgvs": clean(row.get("c_hgvs")),
                    "p_hgvs": clean(row.get("p_hgvs")),
                    "current_mutation_analysis": clean(row.get("mutation_analysis")),
                    "recommendation": (
                        "建议保留p_hgvs空值或由人工确认；剪接位点不应自动伪造蛋白改变。"
                        "报告展示可使用c.HGVS和“剪接位点变异”说明。"
                    ),
                    "review_status": "待医学审核",
                    "review_notes": "",
                }
            )
    rows.sort(key=lambda item: (item["gene"], item["c_hgvs"]))
    return rows


def build_candidate_yaml(
    *,
    missing_gene_rows: list[dict[str, str]],
    low_info_rows: list[dict[str, Any]],
    phgvs_rows: list[dict[str, str]],
    generated_at: str,
) -> dict[str, Any]:
    return {
        "source": {
            "panel": "crc_358_msi",
            "status": "pending_medical_review",
            "candidate_type": "step1_knowledge_optimization_review_package",
            "generated_at": generated_at,
            "important_note": (
                "This file is a review artifact only. Do not deploy directly. "
                "Approved add/replace decisions must be applied through the normal review gate."
            ),
        },
        "add_gene_sections": [
            {
                "gene": row["gene"],
                "intro": row["intro"],
                "mutation_analysis": row["mutation_analysis"],
                "review_status": "待医学审核",
                "source_note": row["source_note"],
            }
            for row in missing_gene_rows
        ],
        "replace_mutation_analysis_candidates": [
            {
                "gene": row["gene"],
                "c_hgvs": row["c_hgvs"],
                "p_hgvs": row["p_hgvs"],
                "priority": row["priority"],
                "current_mutation_analysis": row["current_mutation_analysis"],
                "candidate_mutation_analysis": row["candidate_mutation_analysis"],
                "review_status": "待医学审核",
            }
            for row in low_info_rows
            if row.get("candidate_mutation_analysis")
        ],
        "phgvs_review_items": phgvs_rows,
    }


def write_sheet(ws: Any, rows: list[dict[str, Any]], *, freeze: str = "A2") -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 60)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(
    path: Path,
    *,
    summary_rows: list[dict[str, Any]],
    missing_gene_rows: list[dict[str, Any]],
    low_info_rows: list[dict[str, Any]],
    phgvs_rows: list[dict[str, Any]],
    yaml_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(ws, summary_rows)
    ws2 = wb.create_sheet("新增5基因候选")
    write_sheet(ws2, missing_gene_rows)
    ws3 = wb.create_sheet("P1P2低信息优化")
    write_sheet(ws3, low_info_rows)
    ws4 = wb.create_sheet("pHGVS空值核对")
    write_sheet(ws4, phgvs_rows)
    ws5 = wb.create_sheet("候选YAML说明")
    write_sheet(
        ws5,
        [
            {
                "项目": "候选YAML路径",
                "内容": str(yaml_path),
                "说明": "仅为审核材料，不可直接覆盖生产知识库。",
            },
            {
                "项目": "合入原则",
                "内容": "新增基因可审核通过后添加；替换候选必须由报告组/医学审核明确通过。",
                "说明": "pHGVS空值项建议先核对，不自动生成蛋白改变。",
            },
        ],
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def scan_pii_text(text: str) -> list[str]:
    return [name for name, pattern in PII_PATTERNS.items() if pattern.search(text)]


def validate_outputs(paths: list[Path]) -> dict[str, Any]:
    validation: dict[str, Any] = {"files": {}, "pii_hits": [], "formula_cells": []}
    for path in paths:
        if path.suffix.lower() == ".xlsx":
            wb = load_workbook(path, data_only=False, read_only=True)
            sheet_counts = {}
            for ws in wb.worksheets:
                sheet_counts[ws.title] = ws.max_row - 1 if ws.max_row else 0
                for row in ws.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            validation["formula_cells"].append(f"{path}:{ws.title}!{cell.coordinate}")
                        if value:
                            hits = scan_pii_text(str(value))
                            for hit in hits:
                                validation["pii_hits"].append(f"{hit}:{path}:{ws.title}!{cell.coordinate}")
            validation["files"][str(path)] = {"sheets": sheet_counts}
        else:
            text = path.read_text(encoding="utf-8")
            for hit in scan_pii_text(text):
                validation["pii_hits"].append(f"{hit}:{path}")
            validation["files"][str(path)] = {"bytes": path.stat().st_size}
    return validation


def build_package(*, overlay_path: Path, out_dir: Path) -> dict[str, Any]:
    generated_at = datetime.now(timezone.utc).isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)
    overlay = load_yaml(overlay_path)
    gene_sections = overlay.get("gene_sections") or []
    existing_keys = {row_key(row) for row in gene_sections}
    missing_gene_rows = [
        dict(row, review_status="待医学审核", review_notes="")
        for row in MISSING_GENE_CANDIDATES
        if (row["gene"], "", "") not in existing_keys
    ]
    low_info_rows = build_low_info_rows(gene_sections)
    phgvs_rows = build_phgvs_rows(gene_sections)

    yaml_path = out_dir / "reviewed_part3_knowledge.step1_candidates.yaml"
    candidate_yaml = build_candidate_yaml(
        missing_gene_rows=missing_gene_rows,
        low_info_rows=low_info_rows,
        phgvs_rows=phgvs_rows,
        generated_at=generated_at,
    )
    dump_yaml(yaml_path, candidate_yaml)

    xlsx_path = out_dir / "CRC358_step1_知识库优化候选审核表_20260621.xlsx"
    summary_rows = [
        {"项目": "生成时间", "结果": generated_at, "说明": "UTC时间"},
        {"项目": "生产overlay", "结果": str(overlay_path), "说明": "本脚本只读，不覆盖"},
        {"项目": "当前gene_sections", "结果": len(gene_sections), "说明": "生产库现状"},
        {"项目": "新增基因候选", "结果": len(missing_gene_rows), "说明": "覆盖矩阵E缺失优先补齐"},
        {"项目": "P1/P2低信息条目", "结果": len(low_info_rows), "说明": "不等于错误，优先审核是否精修"},
        {"项目": "pHGVS空值核对", "结果": len(phgvs_rows), "说明": "剪接位点不要自动伪造蛋白改变"},
        {"项目": "合入状态", "结果": "未合入/未部署", "说明": "审核通过后再进入正式release gate"},
    ]
    write_workbook(
        xlsx_path,
        summary_rows=summary_rows,
        missing_gene_rows=missing_gene_rows,
        low_info_rows=low_info_rows,
        phgvs_rows=phgvs_rows,
        yaml_path=yaml_path,
    )

    summary = {
        "generated_at": generated_at,
        "overlay": str(overlay_path),
        "output_dir": str(out_dir),
        "workbook": str(xlsx_path),
        "candidate_yaml": str(yaml_path),
        "counts": {
            "gene_sections": len(gene_sections),
            "missing_gene_candidates": len(missing_gene_rows),
            "low_info_p1_p2": len(low_info_rows),
            "phgvs_blank_with_chgvs": len(phgvs_rows),
            "replace_candidates": len(candidate_yaml["replace_mutation_analysis_candidates"]),
        },
        "low_info_priority_counts": dict(Counter(row["priority"] for row in low_info_rows)),
        "missing_genes": [row["gene"] for row in missing_gene_rows],
        "phgvs_review_genes": [f"{row['gene']} {row['c_hgvs']}" for row in phgvs_rows],
    }
    summary_path = out_dir / "step1_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    readme_path = out_dir / "README.md"
    readme_path.write_text(
        "\n".join(
            [
                "# CRC358 知识库第一步优化候选包",
                "",
                "状态：待医学审核，未合入生产，未部署。",
                "",
                "本包包含：",
                "",
                f"- 审核表：`{xlsx_path.name}`",
                f"- 候选YAML：`{yaml_path.name}`",
                f"- 摘要JSON：`{summary_path.name}`",
                "",
                "使用原则：新增5个缺失基因可作为第一批补库候选；P1/P2低信息条目需要报告组判断是否替换；pHGVS空值项先核对，不自动伪造蛋白改变。",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    validation = validate_outputs([xlsx_path, yaml_path, summary_path, readme_path])
    validation_path = out_dir / "validation.json"
    validation_path.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    summary["validation"] = validation
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--overlay", type=Path, default=DEFAULT_OVERLAY)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    summary = build_package(overlay_path=args.overlay, out_dir=args.out_dir)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if summary["validation"]["pii_hits"] or summary["validation"]["formula_cells"]:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
