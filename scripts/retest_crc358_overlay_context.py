#!/usr/bin/env python3
"""Retest CRC358 Part 3 context coverage with a draft reviewed overlay.

This script does not render DOCX files. It runs the same Excel mapping and
CRC358 enhancer path that builds ``gene_knowledge_sections`` and
``drug_analysis_sections``, then compares production reviewed overlay vs a
merged draft overlay.
"""

from __future__ import annotations

import argparse
import hashlib
import re
import shutil
import sys
import zipfile
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportgen.core.enhancer_registry import get_enhancer, get_panel_registry
from reportgen.core.report_generator import ReportGenerator
from reportgen.knowledge import GeneKnowledgeProvider


DEFAULT_ZIP = Path("肠癌358变异表.zip")
DEFAULT_INPUT_DIR = Path("tmp/overlay_retest/inputs")
DEFAULT_OUTPUT_DIR = Path("tmp/overlay_retest")
DEFAULT_BASE_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_DRAFT_OVERLAY = Path("tmp/knowledge_buildout/reviewed_part3_knowledge_merged_machine_preapproved_v0.1.yaml")
DEFAULT_OUTPUT_XLSX = Path("tmp/knowledge_buildout/CRC358_merged_overlay_context_retest_v0.1.xlsx")
REPORT_DATE = "2026-06-11"


HGVS_C_RE = re.compile(r"c\.[A-Za-z0-9_+\-*?>=.]+(?:delins|del|dup|ins)?[A-Za-z0-9_+\-*?>=.]*")
HGVS_P_RE = re.compile(r"p\.[A-Za-z0-9_+\-*?>=.]+")


def compact_hash(value: str) -> str:
    text = re.sub(r"\s+", "", value or "")
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:16]


def preview(value: str, limit: int = 120) -> str:
    text = re.sub(r"\s+", " ", value or "").strip()
    return text[:limit]


def sample_id_from_path(path: Path) -> str:
    match = re.search(r"(LZ\d+)", path.name, re.I)
    return match.group(1).upper() if match else path.stem


def extract_zip_inputs(zip_path: Path, input_dir: Path) -> list[Path]:
    input_dir.mkdir(parents=True, exist_ok=True)
    for old in input_dir.glob("*"):
        if old.is_file():
            old.unlink()
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir() or not info.filename.lower().endswith(".xlsx"):
                continue
            if Path(info.filename).name.startswith("._"):
                continue
            match = re.search(r"(LZ\d+)\.xlsx$", info.filename, re.I)
            if not match:
                continue
            out = input_dir / f"{match.group(1).upper()}.xlsx"
            with zf.open(info) as src, out.open("wb") as dst:
                shutil.copyfileobj(src, dst)
    for sidecar in input_dir.glob("._*"):
        sidecar.unlink()
    return sorted(input_dir.glob("*.xlsx"))


def clinical_overrides(sample_id: str) -> dict[str, Any]:
    return {
        "patient_name": f"测试样本{sample_id}",
        "sample_id": sample_id,
        "report_number": f"MLJY-{sample_id}",
        "gender": "男",
        "age": 60,
        "sample_type": "组织",
        "receive_date": REPORT_DATE,
        "report_date": REPORT_DATE,
        "project_name": "结直肠癌358基因+MSI",
        "cancer_type": "结直肠癌",
        "clinical_diagnosis": "结直肠癌",
    }


def inject_overrides(excel_data: Any, mapping: dict[str, Any], overrides: dict[str, Any]) -> None:
    for key, value in overrides.items():
        synonyms = (mapping.get(key) or {}).get("synonyms") or [key]
        excel_data.add_single_value(synonyms[0], value)


def load_provider(overlay_path: Path) -> GeneKnowledgeProvider:
    provider = GeneKnowledgeProvider(
        {
            "enabled": True,
            "gene_knowledge_db": {
                "enabled": True,
                "path": "missing.xlsx",
                "reviewed_part3_overlay_path": str(overlay_path),
            },
            "gene_transcript_db": {},
        }
    )
    return provider


def build_context_tables(excel_path: Path, overlay_path: Path) -> dict[str, Any]:
    generator = ReportGenerator(config_dir="config", log_level="ERROR")
    excel_data = generator.excel_reader.read(str(excel_path), include_tables=True)
    mapping = generator.config_loader.load_mapping_config().get("single_values", {})
    sample_id = sample_id_from_path(excel_path)
    inject_overrides(excel_data, mapping, clinical_overrides(sample_id))

    report_data = generator.field_mapper.map(excel_data)
    report_data = generator.data_cleaner.validate_and_clean(report_data)
    report_data.set_field("project_name", "结直肠癌358基因+MSI")
    report_data.set_field("report_content", generator.config_loader.get_setting("report_content", {}) or {})

    registration = get_panel_registry().get("crc_358_msi")
    panel_package = registration.package if registration is not None else None
    provider = load_provider(overlay_path)
    report_data = get_enhancer("crc_358_msi").enhance(
        report_data,
        excel_data,
        field_mapper=generator.field_mapper,
        gene_knowledge_provider=provider,
        base_path=".",
        project_type="crc_358_msi",
        panel_package=panel_package,
    )

    return {
        "gene_knowledge_sections": report_data.get_table("gene_knowledge_sections") or [],
        "drug_analysis_sections": report_data.get_table("drug_analysis_sections") or [],
        "variants": report_data.get_table("variants") or [],
        "summary_variants": report_data.get_table("summary_variants") or [],
        "validation_errors": list(report_data.validation_errors or []),
    }


def section_key(row: dict[str, Any]) -> str:
    header = str(row.get("header") or "")
    gene = str(row.get("gene") or "")
    c_hgvs = row.get("c_hgvs") or ""
    p_hgvs = row.get("p_hgvs") or ""
    if not c_hgvs:
        m = HGVS_C_RE.search(header)
        c_hgvs = m.group(0) if m else ""
    if not p_hgvs:
        m = HGVS_P_RE.search(header)
        p_hgvs = m.group(0) if m else ""
    return "|".join([gene, str(c_hgvs), str(p_hgvs), header])


def drug_key(row: dict[str, Any]) -> str:
    return "|".join(
        [
            str(row.get("gene") or ""),
            str(row.get("c_hgvs") or ""),
            str(row.get("p_hgvs") or ""),
            str(row.get("drug_type") or ""),
            str(row.get("drug_name") or ""),
            str(row.get("header") or ""),
        ]
    )


def indexed(rows: list[dict[str, Any]], kind: str) -> dict[str, dict[str, Any]]:
    key_fn = drug_key if kind == "drug" else section_key
    return {key_fn(row): row for row in rows}


def compare_gene_sections(sample_id: str, base_rows: list[dict[str, Any]], draft_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    base = indexed(base_rows, "gene")
    draft = indexed(draft_rows, "gene")
    changes: list[dict[str, Any]] = []
    for key in sorted(set(base) | set(draft)):
        b = base.get(key, {})
        d = draft.get(key, {})
        for field in ("intro", "mutation_analysis"):
            b_text = str(b.get(field) or "")
            d_text = str(d.get(field) or "")
            if compact_hash(b_text) != compact_hash(d_text):
                changes.append(
                    {
                        "sample_id": sample_id,
                        "change_type": "added" if not b_text and d_text else "changed" if b_text and d_text else "removed",
                        "field": field,
                        "gene": d.get("gene") or b.get("gene") or "",
                        "header": d.get("header") or b.get("header") or "",
                        "baseline_preview": preview(b_text),
                        "draft_preview": preview(d_text),
                    }
                )
    return changes


def compare_drug_sections(sample_id: str, base_rows: list[dict[str, Any]], draft_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    base = indexed(base_rows, "drug")
    draft = indexed(draft_rows, "drug")
    changes: list[dict[str, Any]] = []
    for key in sorted(set(base) | set(draft)):
        b = base.get(key, {})
        d = draft.get(key, {})
        for field in ("relation", "clinical"):
            b_text = str(b.get(field) or "")
            d_text = str(d.get(field) or "")
            if compact_hash(b_text) != compact_hash(d_text):
                changes.append(
                    {
                        "sample_id": sample_id,
                        "change_type": "added" if not b_text and d_text else "changed" if b_text and d_text else "removed",
                        "field": field,
                        "gene": d.get("gene") or b.get("gene") or "",
                        "variant": d.get("variant") or b.get("variant") or "",
                        "drug_type": d.get("drug_type") or b.get("drug_type") or "",
                        "drug_name": d.get("drug_name") or b.get("drug_name") or "",
                        "baseline_preview": preview(b_text),
                        "draft_preview": preview(d_text),
                    }
                )
    return changes


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 10), 60)
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                width = max(width, min(len(str(item.value or "")) + 2, 60))
                item.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def run_retest(inputs: list[Path], base_overlay: Path, draft_overlay: Path, limit: int | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    sample_rows: list[dict[str, Any]] = []
    gene_changes: list[dict[str, Any]] = []
    drug_changes: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []

    selected = inputs[:limit] if limit else inputs
    for excel in selected:
        sample_id = sample_id_from_path(excel)
        try:
            base = build_context_tables(excel, base_overlay)
            draft = build_context_tables(excel, draft_overlay)
            g_changes = compare_gene_sections(sample_id, base["gene_knowledge_sections"], draft["gene_knowledge_sections"])
            d_changes = compare_drug_sections(sample_id, base["drug_analysis_sections"], draft["drug_analysis_sections"])
            gene_changes.extend(g_changes)
            drug_changes.extend(d_changes)
            sample_rows.append(
                {
                    "sample_id": sample_id,
                    "status": "PASS",
                    "baseline_gene_sections": len(base["gene_knowledge_sections"]),
                    "draft_gene_sections": len(draft["gene_knowledge_sections"]),
                    "gene_section_changes": len(g_changes),
                    "baseline_drug_sections": len(base["drug_analysis_sections"]),
                    "draft_drug_sections": len(draft["drug_analysis_sections"]),
                    "drug_section_changes": len(d_changes),
                    "baseline_variants": len(base["variants"]),
                    "summary_variants": len(base["summary_variants"]),
                    "validation_errors": len(base["validation_errors"]) + len(draft["validation_errors"]),
                }
            )
        except Exception as exc:  # pragma: no cover - operational script
            sample_rows.append(
                {
                    "sample_id": sample_id,
                    "status": "FAIL",
                    "baseline_gene_sections": 0,
                    "draft_gene_sections": 0,
                    "gene_section_changes": 0,
                    "baseline_drug_sections": 0,
                    "draft_drug_sections": 0,
                    "drug_section_changes": 0,
                    "baseline_variants": 0,
                    "summary_variants": 0,
                    "validation_errors": 0,
                }
            )
            failures.append({"sample_id": sample_id, "error": f"{type(exc).__name__}: {exc}"})
    return sample_rows, gene_changes, drug_changes, failures


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--zip", type=Path, default=DEFAULT_ZIP)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--base-overlay", type=Path, default=DEFAULT_BASE_OVERLAY)
    parser.add_argument("--draft-overlay", type=Path, default=DEFAULT_DRAFT_OVERLAY)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    inputs = extract_zip_inputs(args.zip, args.input_dir)
    sample_rows, gene_changes, drug_changes, failures = run_retest(
        inputs,
        args.base_overlay,
        args.draft_overlay,
        limit=args.limit or None,
    )

    summary_rows = [
        {"指标": "输入Excel数", "结果": len(sample_rows)},
        {"指标": "PASS样本数", "结果": sum(1 for row in sample_rows if row["status"] == "PASS")},
        {"指标": "FAIL样本数", "结果": sum(1 for row in sample_rows if row["status"] == "FAIL")},
        {"指标": "gene section变化条目", "结果": len(gene_changes)},
        {"指标": "drug section变化条目", "结果": len(drug_changes)},
        {"指标": "base overlay", "结果": str(args.base_overlay)},
        {"指标": "draft overlay", "结果": str(args.draft_overlay)},
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    write_sheet(ws, summary_rows)
    ws = wb.create_sheet("样本对比")
    write_sheet(ws, sample_rows)
    ws = wb.create_sheet("gene_section变化")
    write_sheet(ws, gene_changes)
    ws = wb.create_sheet("drug_section变化")
    write_sheet(ws, drug_changes)
    ws = wb.create_sheet("失败样本")
    write_sheet(ws, failures)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    print(f"inputs={len(inputs)} tested={len(sample_rows)} pass={summary_rows[1]['结果']} fail={summary_rows[2]['结果']}")
    print(f"gene_changes={len(gene_changes)} drug_changes={len(drug_changes)}")
    print(f"output={args.output}")
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
