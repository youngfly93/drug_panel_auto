#!/usr/bin/env python3
"""Apply approved CRC358 pending-review decisions into a reviewed overlay draft.

The default output is a draft YAML under ``tmp/``. Production overlay updates
must be explicit via ``--output panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml``.
Only rows marked ``通过`` or ``修改后通过`` in the batch9 review workbook are promoted.
"""

from __future__ import annotations

import argparse
import copy
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


DEFAULT_BASE_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_REVIEW_WORKBOOK = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/CRC358_batch9_待医学审核合入包_20260614.xlsx"
)
DEFAULT_OUTPUT = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/reviewed_part3_knowledge.approved_from_review_batch9.yaml"
)
DEFAULT_SUMMARY = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/approved_review_apply_summary_batch9.json"
)

APPROVED = {"通过", "修改后通过"}
PENDING_STATUSES = {"", "待医学审核", "待审核"}
FULL_REVIEW_SHEETS = ("新增gene完整审核", "新增drug完整审核")
PII_PATTERNS = {
    "sample_id": re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    "report_no": re.compile(r"报告编号"),
    "name_label": re.compile(r"姓名[:：]"),
    "sender": re.compile(r"送检者"),
    "date": re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
}


def clean(value: Any) -> str:
    return str(value or "").strip()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def gene_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (clean(row.get("gene")).upper(), clean(row.get("c_hgvs")), clean(row.get("p_hgvs")))


def drug_key(row: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        clean(row.get("gene")).upper(),
        clean(row.get("c_hgvs")),
        clean(row.get("p_hgvs")),
        clean(row.get("type") or "benefit"),
        clean(row.get("applicability")),
        clean(row.get("drug_name")),
        clean(row.get("header")),
    )


def pii_hits(text: str) -> list[str]:
    return [name for name, pattern in PII_PATTERNS.items() if pattern.search(text)]


def scan_row_text(row: dict[str, Any]) -> str:
    fields = (
        "gene",
        "c_hgvs",
        "p_hgvs",
        "header",
        "drug_name",
        "intro",
        "mutation_analysis",
        "relation",
        "clinical",
        "reviewed_intro",
        "reviewed_mutation_analysis",
        "reviewed_relation",
        "reviewed_clinical",
    )
    return "\n".join(clean(row.get(field)) for field in fields)


def read_review_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [name for name in FULL_REVIEW_SHEETS if name in wb.sheetnames]
    if not sheet_names:
        if "新增候选" not in wb.sheetnames:
            raise ValueError("workbook missing full review sheets and fallback sheet: 新增候选")
        sheet_names = ["新增候选"]

    rows: list[dict[str, str]] = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows or len(raw_rows[0]) == 1 and clean(raw_rows[0][0]) == "无数据":
            continue
        headers = [clean(value) for value in raw_rows[0]]
        for raw in raw_rows[1:]:
            item = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
            item["_sheet"] = sheet_name
            rows.append(item)
    return rows


def final_text(row: dict[str, str], original_field: str, reviewed_field: str) -> tuple[str, str]:
    status = clean(row.get("review_status"))
    original = clean(row.get(original_field))
    reviewed = clean(row.get(reviewed_field))
    if status == "修改后通过":
        if not original and not reviewed:
            return "", ""
        if not reviewed:
            return "", f"{reviewed_field}缺失：修改后通过必须填写最终定稿文本"
        return reviewed, ""
    if status == "通过":
        return reviewed or original, ""
    return "", ""


def build_gene_row(row: dict[str, str]) -> tuple[dict[str, str] | None, list[str]]:
    issues: list[str] = []
    gene = clean(row.get("gene")).upper()
    if not gene:
        return None, ["gene缺失"]
    intro, issue = final_text(row, "intro", "reviewed_intro")
    if issue:
        issues.append(issue)
    mutation_analysis, issue = final_text(row, "mutation_analysis", "reviewed_mutation_analysis")
    if issue:
        issues.append(issue)
    if mutation_analysis and not clean(row.get("c_hgvs")):
        issues.append("mutation_analysis缺少c_hgvs")
    if not intro and not mutation_analysis:
        return None, issues or ["无可合入gene正文"]
    out = {
        "gene": gene,
        "c_hgvs": clean(row.get("c_hgvs")),
        "p_hgvs": clean(row.get("p_hgvs")),
    }
    if intro:
        out["intro"] = intro
    if mutation_analysis:
        out["mutation_analysis"] = mutation_analysis
    return out, issues


def build_drug_row(row: dict[str, str]) -> tuple[dict[str, str] | None, list[str]]:
    issues: list[str] = []
    gene = clean(row.get("gene")).upper()
    if not gene:
        return None, ["gene缺失"]
    relation, issue = final_text(row, "relation", "reviewed_relation")
    if issue:
        issues.append(issue)
    clinical, issue = final_text(row, "clinical", "reviewed_clinical")
    if issue:
        issues.append(issue)
    drug_name = clean(row.get("drug_name"))
    if not drug_name:
        issues.append("drug_name缺失")
    if not relation or not clinical:
        issues.append("relation/clinical未成对")
    if issues:
        return None, issues
    out = {
        "gene": gene,
        "c_hgvs": clean(row.get("c_hgvs")),
        "p_hgvs": clean(row.get("p_hgvs")),
        "type": clean(row.get("type")) or "benefit",
        "applicability": clean(row.get("applicability")),
        "header": clean(row.get("header")),
        "drug_name": drug_name,
        "relation": relation,
        "clinical": clinical,
    }
    return out, []


def strip_empty(row: dict[str, str]) -> dict[str, str]:
    return {key: value for key, value in row.items() if value}


def apply_review_decisions(
    base_overlay: dict[str, Any],
    review_rows: list[dict[str, str]],
) -> tuple[dict[str, Any], list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    merged = copy.deepcopy(base_overlay)
    gene_rows = copy.deepcopy(merged.get("gene_sections") or [])
    drug_rows = copy.deepcopy(merged.get("drug_sections") or [])
    seen_gene = {gene_key(row) for row in gene_rows}
    seen_drug = {drug_key(row) for row in drug_rows}
    added: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []
    issues: list[dict[str, str]] = []

    for row in review_rows:
        status = clean(row.get("review_status"))
        section = clean(row.get("section"))
        if status in PENDING_STATUSES:
            skipped.append(
                {
                    "reason": "未审核通过",
                    "review_status": status or "空",
                    "section": section,
                    "gene": clean(row.get("gene")).upper(),
                    "c_hgvs": clean(row.get("c_hgvs")),
                    "p_hgvs": clean(row.get("p_hgvs")),
                    "drug_name": clean(row.get("drug_name")),
                }
            )
            continue
        if status not in APPROVED:
            skipped.append(
                {
                    "reason": "审核结论不合入",
                    "review_status": status,
                    "section": section,
                    "gene": clean(row.get("gene")).upper(),
                    "c_hgvs": clean(row.get("c_hgvs")),
                    "p_hgvs": clean(row.get("p_hgvs")),
                    "drug_name": clean(row.get("drug_name")),
                }
            )
            continue
        hits = pii_hits(scan_row_text(row))
        if hits:
            issues.append(
                {
                    "issue": "PII风险",
                    "detail": ",".join(hits),
                    "section": section,
                    "gene": clean(row.get("gene")).upper(),
                    "c_hgvs": clean(row.get("c_hgvs")),
                    "p_hgvs": clean(row.get("p_hgvs")),
                    "drug_name": clean(row.get("drug_name")),
                }
            )
            continue

        if section == "gene_sections":
            new_row, row_issues = build_gene_row(row)
            if row_issues:
                issues.extend(
                    {
                        "issue": issue,
                        "section": section,
                        "gene": clean(row.get("gene")).upper(),
                        "c_hgvs": clean(row.get("c_hgvs")),
                        "p_hgvs": clean(row.get("p_hgvs")),
                        "drug_name": clean(row.get("drug_name")),
                    }
                    for issue in row_issues
                )
                continue
            assert new_row is not None
            key = gene_key(new_row)
            if key in seen_gene:
                skipped.append(
                    {
                        "reason": "生产或前序审核行已存在同key",
                        "review_status": status,
                        "section": section,
                        "gene": new_row["gene"],
                        "c_hgvs": new_row.get("c_hgvs", ""),
                        "p_hgvs": new_row.get("p_hgvs", ""),
                        "drug_name": "",
                    }
                )
                continue
            gene_rows.append(strip_empty(new_row))
            seen_gene.add(key)
            added.append(
                {
                    "section": section,
                    "gene": new_row["gene"],
                    "c_hgvs": new_row.get("c_hgvs", ""),
                    "p_hgvs": new_row.get("p_hgvs", ""),
                    "drug_name": "",
                    "review_status": status,
                }
            )
            continue

        if section == "drug_sections":
            new_row, row_issues = build_drug_row(row)
            if row_issues:
                issues.extend(
                    {
                        "issue": issue,
                        "section": section,
                        "gene": clean(row.get("gene")).upper(),
                        "c_hgvs": clean(row.get("c_hgvs")),
                        "p_hgvs": clean(row.get("p_hgvs")),
                        "drug_name": clean(row.get("drug_name")),
                    }
                    for issue in row_issues
                )
                continue
            assert new_row is not None
            key = drug_key(new_row)
            if key in seen_drug:
                skipped.append(
                    {
                        "reason": "生产或前序审核行已存在同key",
                        "review_status": status,
                        "section": section,
                        "gene": new_row["gene"],
                        "c_hgvs": new_row.get("c_hgvs", ""),
                        "p_hgvs": new_row.get("p_hgvs", ""),
                        "drug_name": new_row.get("drug_name", ""),
                    }
                )
                continue
            drug_rows.append(strip_empty(new_row))
            seen_drug.add(key)
            added.append(
                {
                    "section": section,
                    "gene": new_row["gene"],
                    "c_hgvs": new_row.get("c_hgvs", ""),
                    "p_hgvs": new_row.get("p_hgvs", ""),
                    "drug_name": new_row.get("drug_name", ""),
                    "review_status": status,
                }
            )
            continue

        issues.append(
            {
                "issue": "未知section",
                "section": section,
                "gene": clean(row.get("gene")).upper(),
                "c_hgvs": clean(row.get("c_hgvs")),
                "p_hgvs": clean(row.get("p_hgvs")),
                "drug_name": clean(row.get("drug_name")),
            }
        )

    source = dict(merged.get("source") or {})
    source["approved_review_applied_at"] = datetime.now(timezone.utc).isoformat()
    source["approved_review_policy"] = (
        "Only rows marked 通过/修改后通过 are promoted; production keys win; "
        "modified rows must provide reviewed_* final text."
    )
    merged["source"] = source
    merged["gene_sections"] = gene_rows
    merged["drug_sections"] = drug_rows
    return merged, added, skipped, issues


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-overlay", type=Path, default=DEFAULT_BASE_OVERLAY)
    parser.add_argument("--review-workbook", type=Path, default=DEFAULT_REVIEW_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--summary-json", type=Path, default=DEFAULT_SUMMARY)
    args = parser.parse_args()

    base = load_yaml(args.base_overlay)
    review_rows = read_review_rows(args.review_workbook)
    merged, added, skipped, issues = apply_review_decisions(base, review_rows)
    dump_yaml(args.output, merged)
    summary = {
        "status": "approved_review_applied_candidate",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "base_overlay": str(args.base_overlay),
        "review_workbook": str(args.review_workbook),
        "output": str(args.output),
        "review_rows": len(review_rows),
        "added": len(added),
        "skipped": len(skipped),
        "approved_skipped": sum(1 for row in skipped if clean(row.get("review_status")) in APPROVED),
        "non_approved_skipped": sum(1 for row in skipped if clean(row.get("review_status")) not in APPROVED),
        "skipped_by_status": dict(Counter(clean(row.get("review_status")) or "空" for row in skipped)),
        "issues": len(issues),
        "added_by_section": {
            "gene_sections": sum(1 for row in added if row.get("section") == "gene_sections"),
            "drug_sections": sum(1 for row in added if row.get("section") == "drug_sections"),
        },
        "issues_detail": issues,
    }
    args.summary_json.parent.mkdir(parents=True, exist_ok=True)
    args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"review_rows={len(review_rows)} added={len(added)} skipped={len(skipped)} issues={len(issues)}")
    print(f"output={args.output}")
    print(f"summary_json={args.summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
