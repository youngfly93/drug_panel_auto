#!/usr/bin/env python3
"""按"各 panel 共用"的真实 Excel 结构，生成**完全虚构**的子宫内膜29试跑 Excel。

做法：克隆模板 Excel 的 sheet 名 + 表头结构 → 把 Variations/Msisensor/TMB/QC 写成
虚构子宫内膜数据 → 其余携带真实样本数据的 sheet 只保留表头、清空数据行 → 加一个
Meta sheet 填虚构病人信息（方便表单自动带出；网页也可手填）。产物全去标识，可直接
上传试跑。文件名含"子宫内膜癌分子分型29基因检测"以触发项目类型自动识别。

用法（--template 传任意一份"共用格式"的真实/示例 Excel，仅取其结构，数据全部覆盖）：
  python scripts/gen_endometrial_test_excels.py \
    --template "<某份共用格式Excel.xlsx>" --out-dir "output/endometrial_试跑Excel"
"""
from __future__ import annotations

import argparse
import warnings
from copy import copy
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

# 只写入这些 sheet 的虚构数据；其余 sheet 清空数据行只留表头
SYNTH_SHEETS = {"Variations", "Msisensor", "TMB", "QC"}

# 3 类试跑病例（虚构）。variants 列：(类别, 基因, 转录本, cHGVS, pHGVS_S, pHGVS_A, Function, Freq, ExIn, CLNSIG, Chr, MType)
CASES = {
    "A_MSS": {
        "patient": {"患者姓名": "测试甲", "性别": "女", "年龄": 55, "样本类型": "组织",
                    "临床诊断": "子宫内膜癌", "样本编号": "LZ900001", "报告编号": "MLJY-LZ900001",
                    "项目名称": "子宫内膜癌分子分型29基因检测", "检测项目": "子宫内膜癌分子分型29基因检测",
                    "取材手段": "手术", "取材部位": "子宫", "送检日期": "2026-03-02", "报告日期": "2026-03-10",
                    "检测方法": "NGS高通量测序"},
        "msi": ("MSS", "1.51"), "tmb": ("6", "0.52", "5.77"),
        "variants": [
            ("Ⅱ类", "PTEN", "NM_000314.8", "c.697C>T", "p.R233*", "p.Arg233*", "Nonsense", "38.2", "EX7", "Pathogenic", "chr10", "SNV"),
            ("Ⅱ类", "PIK3CA", "NM_006218.4", "c.1633G>A", "p.E545K", "p.Glu545Lys", "Missense", "27.4", "EX10", "Pathogenic", "chr3", "SNV"),
            ("Ⅱ类", "ARID1A", "NM_006015.6", "c.5965C>T", "p.R1989*", "p.Arg1989*", "Nonsense", "31.0", "EX20", "Pathogenic", "chr1", "SNV"),
            ("Ⅱ类", "CTNNB1", "NM_001904.4", "c.110C>T", "p.S37F", "p.Ser37Phe", "Missense", "22.6", "EX3", "Pathogenic", "chr3", "SNV"),
        ],
    },
    "B_MSI-H": {
        "patient": {"患者姓名": "测试乙", "性别": "女", "年龄": 62, "样本类型": "组织、血液",
                    "临床诊断": "子宫内膜癌", "样本编号": "LZ900002", "报告编号": "MLJY-LZ900002",
                    "项目名称": "子宫内膜癌分子分型29基因检测", "检测项目": "子宫内膜癌分子分型29基因检测",
                    "取材手段": "手术", "取材部位": "子宫", "送检日期": "2026-03-04", "报告日期": "2026-03-12",
                    "检测方法": "NGS高通量测序"},
        "msi": ("MSI-H", "28.30"), "tmb": ("21", "0.52", "20.18"),
        "variants": [
            ("Ⅱ类", "PTEN", "NM_000314.8", "c.800del", "p.K267fs", "p.Lys267fs", "Frameshift", "41.5", "EX8", "Pathogenic", "chr10", "Deletion"),
            ("Ⅱ类", "PIK3CA", "NM_006218.4", "c.3140A>G", "p.H1047R", "p.His1047Arg", "Missense", "33.1", "EX21", "Pathogenic", "chr3", "SNV"),
            ("Ⅱ类", "KRAS", "NM_004985.5", "c.35G>A", "p.G12D", "p.Gly12Asp", "Missense", "29.8", "EX2", "Pathogenic", "chr12", "SNV"),
        ],
    },
    "C_复杂": {
        "patient": {"患者姓名": "测试丙", "性别": "女", "年龄": 68, "样本类型": "组织",
                    "临床诊断": "子宫内膜浆液性癌", "样本编号": "LZ900003", "报告编号": "MLJY-LZ900003",
                    "项目名称": "子宫内膜癌分子分型29基因检测", "检测项目": "子宫内膜癌分子分型29基因检测",
                    "取材手段": "手术", "取材部位": "子宫", "送检日期": "2026-03-06", "报告日期": "2026-03-14",
                    "检测方法": "NGS高通量测序"},
        "msi": ("MSI-H", "31.10"), "tmb": ("35", "0.52", "33.64"),
        "variants": [
            ("Ⅱ类", "TP53", "NM_000546.6", "c.524G>A", "p.R175H", "p.Arg175His", "Missense", "62.1", "EX5", "Pathogenic", "chr17", "SNV"),
            ("Ⅱ类", "PPP2R1A", "NM_014225.6", "c.547C>T", "p.R183W", "p.Arg183Trp", "Missense", "44.0", "EX5", "Pathogenic", "chr19", "SNV"),
            ("Ⅱ类", "FBXW7", "NM_033632.3", "c.1393C>T", "p.R465C", "p.Arg465Cys", "Missense", "28.3", "EX9", "Pathogenic", "chr4", "SNV"),
            ("Ⅱ类", "PIK3CA", "NM_006218.4", "c.1633G>A", "p.E545K", "p.Glu545Lys", "Missense", "19.7", "EX10", "Pathogenic", "chr3", "SNV"),
            ("Ⅱ类", "ERBB2", "NM_004448.4", "c.929C>T", "p.S310F", "p.Ser310Phe", "Missense", "17.2", "EX8", "Pathogenic", "chr17", "SNV"),
        ],
    },
}

VAR_FIELDS = ["ExistIn552", "ExistInsmall358", "Gene_Symbol", "Transcript", "cHGVS",
              "pHGVS_S", "pHGVS_A", "Function", "Freq(%)", "ExIn_ID", "CLNSIG", "Chr", "MType"]


def header_index(ws):
    return {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}


def fill_variations(ws, variants):
    hidx = header_index(ws)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, v in enumerate(variants):
        r = 2 + i
        row = dict(zip(VAR_FIELDS, (v[0], 1, v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], v[9], v[10], v[11])))
        for name, val in row.items():
            if name in hidx:
                ws.cell(row=r, column=hidx[name], value=val)


def fill_msisensor(ws, status, percent):
    # 真实布局：第3行(1-based) = msisensor2 | total | num | % | STATUS
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in r:
            if cell.value in ("MSS", "MSI-H", "MSI-L"):
                cell.value = status
            # % 列在 status 左一格
        # 设置 msisensor2 行的 % 列
    # 直接定位 msisensor2 行
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").startswith("msisensor2"):
            ws.cell(row=r, column=4, value=percent)
            ws.cell(row=r, column=5, value=status)


def fill_tmb(ws, var_num, bed, tmb):
    # 真实布局：第2行=表头(Var_num/Bed_size/TMB)，第3行=数据
    ws.cell(row=3, column=1, value=var_num)
    ws.cell(row=3, column=3, value=tmb)


def fill_qc(ws):
    # 写几条通用合格 QC 值（覆盖真实样本数值）
    repl = {"Initial bases on target": "1250000", "Average sequencing depth": "1600",
            "lib diversity": "0.95", "insert": "185"}
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
        label = str(r[0].value or "").strip()
        for k, val in repl.items():
            if label.startswith(k) and len(r) > 1 and r[1].value not in (None, ""):
                r[1].value = val


def clear_other_sheet(ws):
    """保留第1行表头，清空其余数据（去掉真实样本携带数据）。"""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def add_meta_sheet(wb, patient):
    if "Meta" in wb.sheetnames:
        del wb["Meta"]
    ws = wb.create_sheet("Meta", 0)
    keys = list(patient.keys())
    for c, k in enumerate(keys, start=1):
        ws.cell(row=1, column=c, value=k)
        ws.cell(row=2, column=c, value=patient[k])
    return ws


def build_one(template: Path, out_dir: Path, case_id: str, spec: dict) -> Path:
    wb = openpyxl.load_workbook(template)
    add_meta_sheet(wb, spec["patient"])
    for sn in list(wb.sheetnames):
        ws = wb[sn]
        if sn == "Variations":
            fill_variations(ws, spec["variants"])
        elif sn == "Msisensor":
            fill_msisensor(ws, spec["msi"][0], spec["msi"][1])
        elif sn == "TMB":
            fill_tmb(ws, spec["tmb"][0], spec["tmb"][1], spec["tmb"][2])
        elif sn == "QC":
            fill_qc(ws)
        elif sn == "Meta":
            continue
        else:
            clear_other_sheet(ws)
    out_dir.mkdir(parents=True, exist_ok=True)
    sid = spec["patient"]["样本编号"].lower()
    fname = f"测试病例{case_id}-子宫内膜癌-子宫内膜癌分子分型29基因检测-mljy-{sid}.xlsx"
    out = out_dir / fname
    wb.save(out)
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template", required=True, type=Path, help="共用格式的模板 Excel")
    ap.add_argument("--out-dir", required=True, type=Path)
    args = ap.parse_args(argv)
    for cid, spec in CASES.items():
        out = build_one(args.template, args.out_dir, cid, spec)
        print(f"  ✅ {cid}: {out}  (变异{len(spec['variants'])} / MSI={spec['msi'][0]})")


if __name__ == "__main__":
    main()
