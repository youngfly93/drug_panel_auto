#!/usr/bin/env python3
"""Build a reviewed Part3 overlay draft from report-team approved XLSX rows.

Input is the workbook produced by ``harvest_crc_part3_candidates.py``. Only rows
whose review conclusion contains "通过" are converted. The script writes a draft
YAML by default; use ``--apply`` only after clinical/report-team sign-off.

It also supports the legacy candidate workbook produced by
``harvest_crc_legacy_knowledge_candidates.py``. Legacy rows are converted to
gene-level overrides only; variant-specific "function hint" columns are not
promoted as generic text.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def _clean(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _is_approved(value: Any) -> bool:
    text = _clean(value)
    return "通过" in text and "不通过" not in text


_PATIENT_ID_PATTERNS = [
    re.compile(r"\b(?:LZ|MLJY|MLO|KY|YB|XB|B)\d{4,}\b", re.IGNORECASE),
    re.compile(r"\b\d{4}[-/.年]\d{1,2}[-/.月]\d{1,2}日?\b"),
]

_PATIENT_SPECIFIC_MARKERS = (
    "突变丰度",
    "拷贝数为",
    "扩增，拷贝数",
    "此突变在样本",
    "本次样本",
    "本病例",
)

_GENE_LEVEL_SPECIFIC_MARKERS = (
    "该样本检出",
    "本次检测到",
    "本次检测检出",
)

_HGVS_OR_FREQUENCY_RE = re.compile(
    r"(?:\bc\.\d|p\.[A-Z][A-Za-z0-9_*?=]+|\bchr(?:omosome)?\b|\d+(?:\.\d+)?%)",
    re.IGNORECASE,
)

def _rows_by_header(ws) -> list[dict[str, Any]]:
    headers = [_clean(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    return rows


def _has_patient_id_or_date(text: str) -> bool:
    return any(pattern.search(text) for pattern in _PATIENT_ID_PATTERNS)


def _is_safe_gene_level_text(text: str) -> bool:
    if not text:
        return False
    if _has_patient_id_or_date(text):
        return False
    if any(marker in text for marker in _PATIENT_SPECIFIC_MARKERS):
        return False
    if any(marker in text for marker in _GENE_LEVEL_SPECIFIC_MARKERS):
        return False
    return not _HGVS_OR_FREQUENCY_RE.search(text)


def _is_safe_variant_level_text(text: str) -> bool:
    if not text:
        return False
    if _has_patient_id_or_date(text):
        return False
    if any(marker in text for marker in _PATIENT_SPECIFIC_MARKERS):
        return False
    return True


def _load_existing_keys(overlay_path: Path) -> set[tuple[str, str, str]]:
    if not overlay_path.exists():
        return set()
    data = yaml.safe_load(overlay_path.read_text(encoding="utf-8")) or {}
    keys = set()
    for row in data.get("gene_sections") or []:
        gene = _clean(row.get("gene")).upper()
        c_hgvs = _clean(row.get("c_hgvs"))
        p_hgvs = _clean(row.get("p_hgvs"))
        if gene:
            keys.add((gene, c_hgvs, p_hgvs))
    return keys


def _approved_variant_sections(workbook_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path)
    if "需补库位点" not in wb.sheetnames:
        return []
    sections: list[dict[str, str]] = []
    for row in _rows_by_header(wb["需补库位点"]):
        if not _is_approved(row.get("审核结论")):
            continue
        gene = _clean(row.get("基因")).upper()
        c_hgvs = _clean(row.get("cHGVS"))
        p_hgvs = _clean(row.get("pHGVS"))
        intro = _clean(row.get("基础候选简介"))
        analysis = _clean(row.get("基础候选解析"))
        if intro and not _is_safe_variant_level_text(intro):
            intro = ""
        if analysis and not _is_safe_variant_level_text(analysis):
            analysis = ""
        if not gene or not c_hgvs or not (intro or analysis):
            continue
        item: dict[str, str] = {"gene": gene, "c_hgvs": c_hgvs}
        if p_hgvs:
            item["p_hgvs"] = p_hgvs
        if intro:
            item["intro"] = intro
        if analysis:
            item["mutation_analysis"] = analysis
        sections.append(item)
    return sections


def _section_field(candidate_type: str) -> str:
    text = _clean(candidate_type)
    if "简介" in text:
        return "intro"
    if "解析" in text or "分析" in text or "说明" in text:
        return "mutation_analysis"
    return ""


def _approved_exact_candidate_sections(workbook_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path)
    if "历史精确位点候选" not in wb.sheetnames:
        return []

    grouped: dict[tuple[str, str, str], dict[str, list[str]]] = {}
    for row in _rows_by_header(wb["历史精确位点候选"]):
        if not _is_approved(row.get("审核结论")):
            continue
        gene = _clean(row.get("基因")).upper()
        c_hgvs = _clean(row.get("cHGVS"))
        p_hgvs = _clean(row.get("pHGVS"))
        field = _section_field(_clean(row.get("候选类型")))
        text = _clean(row.get("候选上下文")) or _clean(row.get("命中段落"))
        if not gene or not c_hgvs or not field or not _is_safe_variant_level_text(text):
            continue
        grouped.setdefault((gene, c_hgvs, p_hgvs), {"intro": [], "mutation_analysis": []})
        grouped[(gene, c_hgvs, p_hgvs)][field].append(text)

    sections: list[dict[str, str]] = []
    for (gene, c_hgvs, p_hgvs), fields in sorted(grouped.items()):
        intro = _join_paragraphs(*fields["intro"])
        analysis = _join_paragraphs(*fields["mutation_analysis"])
        if not intro and not analysis:
            continue
        item: dict[str, str] = {"gene": gene, "c_hgvs": c_hgvs}
        if p_hgvs:
            item["p_hgvs"] = p_hgvs
        if intro:
            item["intro"] = intro
        if analysis:
            item["mutation_analysis"] = analysis
        sections.append(item)
    return sections


def _approved_historical_gene_sections(workbook_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path)
    if "历史基因级候选" not in wb.sheetnames:
        return []

    grouped: dict[str, dict[str, list[str]]] = {}
    for row in _rows_by_header(wb["历史基因级候选"]):
        if not _is_approved(row.get("审核结论")):
            continue
        gene = _clean(row.get("基因")).upper()
        field = _section_field(_clean(row.get("候选类型")))
        text = _clean(row.get("候选上下文")) or _clean(row.get("命中段落"))
        if not gene or not field or not _is_safe_gene_level_text(text):
            continue
        grouped.setdefault(gene, {"intro": [], "mutation_analysis": []})
        grouped[gene][field].append(text)

    sections: list[dict[str, str]] = []
    for gene, fields in sorted(grouped.items()):
        intro = _join_paragraphs(*fields["intro"])
        analysis = _join_paragraphs(*fields["mutation_analysis"])
        if not intro and not analysis:
            continue
        item: dict[str, str] = {"gene": gene}
        if intro:
            item["intro"] = intro
        if analysis:
            item["mutation_analysis"] = analysis
        sections.append(item)
    return sections


def _join_paragraphs(*parts: str) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for part in parts:
        text = _clean(part)
        if not text:
            continue
        if "{XX癌" in text or "运营系统调取" in text:
            continue
        if text in seen:
            continue
        seen.add(text)
        out.append(text)
    return "\n".join(out)


def _approved_legacy_gene_sections(
    workbook_path: Path,
    *,
    approve_all: bool,
) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path)
    if "旧库补库候选" not in wb.sheetnames:
        return []

    sections: list[dict[str, str]] = []
    for row in _rows_by_header(wb["旧库补库候选"]):
        if not approve_all and not _is_approved(row.get("审核结论")):
            continue
        gene = _clean(row.get("基因")).upper()
        if not gene:
            continue

        intro = _clean(row.get("旧库基因简介"))
        # The legacy workbook's "旧库功能提示" column can be variant-specific
        # ("该样本检出的突变可能导致..."). Do not promote it into a gene-level
        # override where it would apply to every variant of the gene.
        analysis = _join_paragraphs(
            _clean(row.get("旧库结构域/变异说明")),
            _clean(row.get("旧库肠癌解析")),
        )
        if intro and not _is_safe_gene_level_text(intro):
            intro = ""
        if analysis and not _is_safe_gene_level_text(analysis):
            analysis = ""
        if not intro and not analysis:
            continue

        item: dict[str, str] = {"gene": gene}
        if intro:
            item["intro"] = intro
        if analysis:
            item["mutation_analysis"] = analysis
        sections.append(item)
    return sections


def build_draft(
    *,
    review_xlsx: Path,
    existing_overlay: Path,
    output: Path,
    apply: bool,
    approve_all: bool,
) -> dict[str, Any]:
    approved_variants = _approved_variant_sections(review_xlsx)
    approved_exact = _approved_exact_candidate_sections(review_xlsx)
    approved_historical_genes = _approved_historical_gene_sections(review_xlsx)
    approved_legacy_genes = _approved_legacy_gene_sections(review_xlsx, approve_all=approve_all)
    approved_genes = approved_historical_genes + approved_legacy_genes
    approved = approved_variants + approved_exact + approved_genes
    existing_keys = _load_existing_keys(existing_overlay)
    new_sections = []
    skipped_existing = 0
    for row in approved:
        key = (_clean(row.get("gene")).upper(), _clean(row.get("c_hgvs")), _clean(row.get("p_hgvs")))
        if key in existing_keys:
            skipped_existing += 1
            continue
        new_sections.append(row)

    draft = {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "purpose": (
                "Report-team approved Part3 knowledge additions generated from "
                f"{review_xlsx.name}."
            ),
            "scope": "gene_or_variant_level_reviewed_knowledge",
            "source_type": "report_team_approved_candidate_table",
        },
        "gene_sections": new_sections,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(yaml.safe_dump(draft, allow_unicode=True, sort_keys=False), encoding="utf-8")

    if apply and new_sections:
        existing = yaml.safe_load(existing_overlay.read_text(encoding="utf-8")) or {}
        existing.setdefault("gene_sections", [])
        existing["gene_sections"].extend(new_sections)
        existing_overlay.write_text(
            yaml.safe_dump(existing, allow_unicode=True, sort_keys=False),
            encoding="utf-8",
        )

    return {
        "approved_rows": len(approved),
        "approved_variant_rows": len(approved_variants),
        "approved_exact_rows": len(approved_exact),
        "approved_historical_gene_rows": len(approved_historical_genes),
        "approved_legacy_gene_rows": len(approved_legacy_genes),
        "approved_gene_rows": len(approved_genes),
        "new_sections": len(new_sections),
        "skipped_existing": skipped_existing,
        "draft_output": str(output),
        "applied": bool(apply and new_sections),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--review-xlsx", type=Path, required=True)
    parser.add_argument(
        "--existing-overlay",
        type=Path,
        default=Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml"),
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--approve-all",
        action="store_true",
        help="Treat all legacy candidate rows as approved without editing the workbook.",
    )
    args = parser.parse_args()
    result = build_draft(
        review_xlsx=args.review_xlsx,
        existing_overlay=args.existing_overlay,
        output=args.output,
        apply=args.apply,
        approve_all=args.approve_all,
    )
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
