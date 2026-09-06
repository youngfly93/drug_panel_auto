#!/usr/bin/env python3
# 步骤: 01 从肺癌超集工作簿派生产品成员旗标
# 上游: 原始超集 Excel、panel.yaml 的 derived_input 或显式基因列表
# 输出: .work/ 下的 *-derived-<panel>.xlsx 及无病例正文的哈希回执
# 种子: 无（确定性集合运算；不推断医学结果）
"""Change product membership only, preserving every other XLSX ZIP member.

Example: python scripts/derive_panel_input.py input.xlsx --panel lung_13
The flag is NOT a variant classification. Both the membership flag and the
original classification/event contract must be checked by the report producer.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import posixpath
import re
import zipfile
from pathlib import Path
from typing import Iterable

import yaml
from lxml import etree
from openpyxl.utils.cell import column_index_from_string, get_column_letter

ROOT = Path(__file__).resolve().parents[1]
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
TARGETS = ("Variations", "Hereditary_tumor")
SMALL_FLAG = re.compile(r"^ExistInsmall\d+[A-Za-z0-9_]*$", re.I)
CELL_REF = re.compile(r"^(\$?)([A-Z]+)(\$?)(\d+)$")


def _xml(data: bytes):
    return etree.fromstring(
        data, etree.XMLParser(resolve_entities=False, no_network=True)
    )


def _text(cell, strings: list[str]) -> str:
    if cell is None:
        return ""
    if cell.get("t") == "inlineStr":
        return "".join(cell.itertext())
    value = cell.find(f"{{{NS}}}v")
    text = value.text or "" if value is not None else ""
    return strings[int(text)] if cell.get("t") == "s" and text else text


def _sheet_paths(archive: zipfile.ZipFile) -> dict[str, str]:
    relationships = _xml(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {
        r.get("Id"): posixpath.normpath("xl/" + r.get("Target", ""))
        if not r.get("Target", "").startswith("/")
        else r.get("Target").lstrip("/")
        for r in relationships
        if r.get("TargetMode") != "External"
    }
    return {
        s.get("name"): targets[s.get(f"{{{REL}}}id")]
        for s in _xml(archive.read("xl/workbook.xml")).iter(f"{{{NS}}}sheet")
    }


def _change_sheet(data: bytes, strings: list[str], genes: set[str], flag: str):
    root = _xml(data)
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        raise ValueError("Missing sheetData")
    rows = list(sheet_data)
    header = next(
        (r for r in rows if any(_text(c, strings).strip() == "Gene_Symbol" for c in r)),
        None,
    )
    if header is None:
        raise ValueError("Missing Gene_Symbol header")
    headers = {
        column_index_from_string(CELL_REF.fullmatch(c.get("r")).group(2)): _text(
            c, strings
        ).strip()
        for c in header
    }
    if list(headers.values()).count("Gene_Symbol") != 1:
        raise ValueError("Gene_Symbol must occur exactly once")
    old = sorted(c for c, name in headers.items() if SMALL_FLAG.fullmatch(name))
    if flag in headers.values() and flag not in [headers[c] for c in old]:
        raise ValueError("Flag name collides with an existing non-membership column")
    gene_col = next(c for c, name in headers.items() if name == "Gene_Symbol")
    max_col = max(
        column_index_from_string(CELL_REF.fullmatch(c.get("r")).group(2))
        for r in rows
        for c in r
    )
    new_col = old[0] if old else max_col + 1
    removed = set(old[1:])  # Replace the first old column in place; remove any others.
    if removed and (
        root.findall(f".//{{{NS}}}f") or root.find(f"{{{NS}}}tableParts") is not None
    ):
        raise ValueError(
            "Multiple flag columns with formulas/Excel tables require explicit review"
        )

    def shift_col(index: int) -> int:
        return max(1, index - sum(c < index for c in removed))

    def shift_ref(ref: str) -> str:
        def cell(match):
            return (
                match.group(1)
                + get_column_letter(shift_col(column_index_from_string(match.group(2))))
                + match.group(3)
                + match.group(4)
            )

        return re.sub(r"(\$?)([A-Z]+)(\$?)(\d+)", cell, ref)

    flagged = 0
    header_number = int(header.get("r"))
    for row in rows:
        row_number = int(row.get("r"))
        indexed = {
            column_index_from_string(CELL_REF.fullmatch(c.get("r")).group(2)): c
            for c in row
        }
        gene = _text(indexed.get(gene_col), strings).strip().upper()
        style_source = indexed.get(new_col)
        style = style_source.get("s") if style_source is not None else None
        for column, cell in indexed.items():
            if column in old:
                row.remove(cell)
            elif removed:
                cell.set("r", shift_ref(cell.get("r")))
        if row_number >= header_number:
            cell = etree.Element(
                f"{{{NS}}}c", r=f"{get_column_letter(new_col)}{row_number}"
            )
            if style is not None:
                cell.set("s", style)
            if row_number == header_number:
                cell.set("t", "inlineStr")
                etree.SubElement(
                    etree.SubElement(cell, f"{{{NS}}}is"), f"{{{NS}}}t"
                ).text = flag
            elif gene in genes:
                etree.SubElement(cell, f"{{{NS}}}v").text = "1"
                flagged += 1
            # Empty membership cells intentionally have neither type nor value.
            before = next(
                (
                    c
                    for c in row
                    if column_index_from_string(CELL_REF.fullmatch(c.get("r")).group(2))
                    > new_col
                ),
                None,
            )
            if before is None:
                row.append(cell)
            else:
                before.addprevious(cell)
        row.attrib.pop("spans", None)
    for element in root.iter():
        if element.tag in {f"{{{NS}}}c", f"{{{NS}}}row"}:
            continue
        for attribute in ("ref", "sqref", "topLeftCell", "activeCell"):
            if removed and element.get(attribute):
                element.set(attribute, shift_ref(element.get(attribute)))
    columns = root.find(f"{{{NS}}}cols")
    if columns is not None and removed:
        for column in list(columns):
            kept = [
                c
                for c in range(int(column.get("min")), int(column.get("max")) + 1)
                if c not in removed
            ]
            if not kept:
                columns.remove(column)
            else:
                column.set("min", str(shift_col(min(kept))))
                column.set("max", str(shift_col(max(kept))))
    dimension = root.find(f"{{{NS}}}dimension")
    if dimension is not None:
        last_column = get_column_letter(max(shift_col(max_col), new_col))
        last_row = max(int(row.get("r")) for row in rows)
        dimension.set(
            "ref",
            f"A1:{last_column}{last_row}",
        )
    return etree.tostring(root, encoding="UTF-8", xml_declaration=True), {
        "removed_flags": [headers[c] for c in old],
        "membership_column": flag,
        "flagged_rows": flagged,
        "header_row": header_number,
    }


def derive_panel_input(
    source: Path,
    *,
    panel_id: str,
    genes: Iterable[str],
    flag_column: str,
    output_dir: Path,
) -> dict:
    source, output_dir = Path(source).resolve(), Path(output_dir).resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ValueError("Source must be an existing .xlsx")
    if not re.fullmatch(r"[a-z][a-z0-9_]*", panel_id):
        raise ValueError("Invalid panel id")
    if not SMALL_FLAG.fullmatch(flag_column):
        raise ValueError("Membership column must be an ExistInsmall<N> product flag")
    normalized = {str(g).strip().upper() for g in (genes or []) if str(g).strip()}
    if not normalized or any(
        not re.fullmatch(r"[A-Z][A-Z0-9-]*", g) for g in normalized
    ):
        raise ValueError("A nonempty explicit gene-symbol list is required")
    if ".work" not in output_dir.parts:
        raise ValueError("Derived patient inputs must be written under .work/")
    output = output_dir / f"{source.stem}-derived-{panel_id}.xlsx"
    if (
        output.exists()
        or output.with_suffix(".derivation.json").exists()
        or output == source
    ):
        raise FileExistsError("Refusing to overwrite an existing workbook")
    changes, summary = {}, {}
    with zipfile.ZipFile(source) as archive:
        names = archive.namelist()
        if len(names) != len(set(names)):
            raise ValueError("Duplicate XLSX ZIP members")
        paths = _sheet_paths(archive)
        strings = []
        if "xl/sharedStrings.xml" in names:
            strings = [
                "".join(si.itertext())
                for si in _xml(archive.read("xl/sharedStrings.xml"))
            ]
        for name in TARGETS:
            if name not in paths:
                raise ValueError(f"Missing required worksheet: {name}")
            changes[paths[name]], summary[name] = _change_sheet(
                archive.read(paths[name]), strings, normalized, flag_column
            )
        output_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(output, "x", zipfile.ZIP_DEFLATED) as destination:
            for item in archive.infolist():
                # ZipFile.writestr mutates ZipInfo.header_offset; never hand it
                # the source archive's live ZipInfo used by verification below.
                destination.writestr(
                    copy.copy(item), changes.get(item.filename, archive.read(item))
                )
        with zipfile.ZipFile(output) as destination:
            assert all(
                archive.read(name) == destination.read(name)
                for name in names
                if name not in changes
            )
    receipt = {
        "panel_id": panel_id,
        "membership_column": flag_column,
        "gene_count": len(normalized),
        "worksheets": summary,
        "source_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "derived_sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
        "unchanged_zip_members": len(names) - len(changes),
        "other_worksheet_payloads_byte_identical": True,
        "clinical_values_inferred": False,
        "paired_small_panel_validation": False,
    }
    output.with_suffix(".derivation.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2) + "\n"
    )
    return {**receipt, "output": str(output)}


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path)
    parser.add_argument("--panel", required=True)
    parser.add_argument("--genes", help="Comma/space-separated gene symbols")
    parser.add_argument(
        "--genes-file", type=Path, help="YAML list or mapping with a genes list"
    )
    parser.add_argument(
        "--flag-column", help="Overrides panel.yaml's derived_input.membership_column"
    )
    parser.add_argument(
        "--output-dir", type=Path, default=ROOT / ".work/derived_panel_inputs"
    )
    args = parser.parse_args(argv)
    if not re.fullmatch(r"[a-z][a-z0-9_]*", args.panel):
        parser.error("Invalid panel id")
    if args.genes and args.genes_file:
        parser.error("Use --genes or --genes-file, not both")
    package_path = ROOT / "panels" / args.panel / "panel.yaml"
    profile = (
        (yaml.safe_load(package_path.read_text()).get("derived_input") or {})
        if package_path.is_file()
        else {}
    )
    genes = profile.get("genes") or []
    if args.genes:
        genes = re.split(r"[,\s]+", args.genes)
    if args.genes_file:
        payload = yaml.safe_load(args.genes_file.read_text())
        genes = payload.get("genes") if isinstance(payload, dict) else payload
    flag = args.flag_column or profile.get("membership_column")
    if not flag:
        match = re.fullmatch(r"lung_(\d+)(?:_pdl1)?", args.panel)
        if match:
            flag = "ExistInsmall" + match.group(1)
    if not flag:
        parser.error("An explicit product flag is required")
    print(
        json.dumps(
            derive_panel_input(
                args.source,
                panel_id=args.panel,
                genes=genes,
                flag_column=flag,
                output_dir=args.output_dir,
            ),
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
