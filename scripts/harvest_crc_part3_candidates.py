#!/usr/bin/env python3
"""Harvest review candidates for CRC Part3 knowledge gaps from final DOCX reports.

This script prepares report-team review material only. It deliberately does not
write to the production knowledge overlay.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


MECHANISM_KEYWORDS = (
    "编码",
    "蛋白",
    "受体",
    "激酶",
    "信号通路",
    "抑癌",
    "原癌",
    "转录",
    "修复",
    "甲基化",
    "染色质",
    "细胞周期",
)
CRC_KEYWORDS = ("结直肠癌", "结肠癌", "直肠癌", "大肠癌", "肠癌")
EXCLUDE_SNIPPET_KEYWORDS = (
    "检测基因列表",
    "GeneList",
    "Gene List",
    "未检出",
    "附录",
    "参考文献",
    "检测结果说明",
    "报告导读",
    "致患者信",
)
PATIENT_SPECIFIC_MARKERS = (
    "该样本检出",
    "此突变在样本中",
    "突变丰度",
    "拷贝数为",
    "扩增，拷贝数",
)
HGVS_IN_TEXT_RE = re.compile(r"\bc\.\d|p\.[A-Z][A-Za-z0-9*]+|\bchr\d", re.I)
FREQUENCY_RE = re.compile(r"\d+(?:\.\d+)?\s*%")
PRODUCT_TOKEN_RE = re.compile(
    r"(基因|分子分型|HRD|hrd|精准治疗|靶向|PD[-_\s]?L1|pd[-_\s]?l1|MSI|msi|TMB|tmb)"
)
SAMPLE_TOKEN_RE = re.compile(r"^(?:MLJY|mljy|[Ll][ZzWw]\d{4,}|[A-Z]{2,}\d{5,}.*)$")
PANEL_SIZE_RE = re.compile(r"(\d{1,4})\s*基因")
RESOURCE_PREFIX = "._"


@dataclass(frozen=True)
class GapCandidate:
    priority: str
    gene: str
    c_hgvs: str
    p_hgvs: str
    class_label: str
    current_status: str
    base_intro: str
    base_analysis: str


def _clean(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _stable_id(relative_path: Path) -> str:
    digest = hashlib.sha256(str(relative_path).encode("utf-8")).hexdigest()
    return f"rpt_{digest[:16]}"


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()[:16]


def _normalize_product_family(value: str) -> str:
    text = _clean(value).replace("＋", "+")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"PD[-_]?L1|PDL1", "pd-l1", text, flags=re.IGNORECASE)
    text = re.sub(r"MSI", "msi", text, flags=re.IGNORECASE)
    text = re.sub(r"TMB", "tmb", text, flags=re.IGNORECASE)
    return text.strip("-_ ") or "未识别产品族"


def _infer_product_family(filename: str) -> str:
    stem = Path(filename).stem
    if stem.startswith(RESOURCE_PREFIX):
        stem = stem[len(RESOURCE_PREFIX) :]
    stem = re.sub(r"终版\d*$|终版$|补充报告|修改版|已审核", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"PD[-_\s]?L1", "PDL1", stem, flags=re.IGNORECASE)
    parts = [
        part.strip()
        for part in re.split(r"[-－—–]", stem)
        if part and part.strip()
    ]
    candidates: list[str] = []
    for part in parts:
        if SAMPLE_TOKEN_RE.match(part):
            continue
        if PRODUCT_TOKEN_RE.search(part):
            candidates.append(part)
    if candidates:
        return _normalize_product_family(candidates[-1])
    return "未识别产品族"


def _docx_paragraphs(path: Path) -> list[str]:
    """Extract Word paragraph text with a lightweight XML reader."""
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception:
        return []
    paragraphs: list[str] = []
    for match in re.finditer(r"<w:p[ >].*?</w:p>", xml, re.S):
        para_xml = match.group(0)
        text = "".join(re.findall(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", para_xml, re.S))
        text = html.unescape(re.sub(r"\s+", " ", text)).strip()
        if text:
            paragraphs.append(text)
    return paragraphs


def _is_next_variant_header(text: str) -> bool:
    return bool(
        re.match(r"^[◆◇]?\s*[A-Z0-9][A-Z0-9/-]{1,12}\s*[：:]\s*c\.", text)
        or re.match(r"^[◆◇]?\s*[A-Z0-9][A-Z0-9/-]{1,12}\s*[：:]\s*chr", text)
    )


def _candidate_section(paragraphs: list[str], start: int, max_paragraphs: int = 14) -> str:
    parts = []
    for idx in range(start, min(len(paragraphs), start + max_paragraphs)):
        text = paragraphs[idx]
        if idx > start and _is_next_variant_header(text):
            break
        if idx > start and re.search(r"^(2\.|3\.|4\.|5\.)", text):
            break
        parts.append(text)
    return "\n".join(parts)


def _looks_like_intro(gene: str, text: str) -> bool:
    if len(text) < 45 or len(text) > 1400:
        return False
    if not (
        text.startswith(f"{gene}基因")
        or text.startswith(f"{gene}又名")
        or text.startswith(gene)
    ):
        return False
    if any(key in text for key in EXCLUDE_SNIPPET_KEYWORDS):
        return False
    if _is_patient_specific_text(text):
        return False
    if "常见突变基因有" in text or "检测基因" in text:
        return False
    return any(key in text for key in MECHANISM_KEYWORDS)


def _is_patient_specific_text(text: str) -> bool:
    if any(marker in text for marker in PATIENT_SPECIFIC_MARKERS):
        return True
    if HGVS_IN_TEXT_RE.search(text):
        return True
    # A standalone percentage in a Part3 paragraph is often abundance. Keep the
    # gene-level pool conservative; exact-site candidates are handled separately.
    return bool(FREQUENCY_RE.search(text))


def _looks_like_analysis(gene: str, text: str) -> bool:
    if len(text) < 50 or len(text) > 1600:
        return False
    if gene not in text:
        return False
    if any(key in text for key in EXCLUDE_SNIPPET_KEYWORDS):
        return False
    if _is_patient_specific_text(text):
        return False
    if "常见突变基因有" in text or "检测基因" in text:
        return False
    if not (
        text.startswith(gene)
        or text.startswith(f"{gene}基因")
        or text.startswith(f"该样本检出{gene}")
        or f"{gene}基因" in text[:80]
    ):
        return False
    return (
        any(key in text for key in CRC_KEYWORDS)
        or "该样本检出" in text
        or "临床意义" in text
        or "疾病的发生发展" in text
    )


def _load_gap_candidates(path: Path, priorities: set[str]) -> list[GapCandidate]:
    wb = load_workbook(path)
    if "待审核候选内容" not in wb.sheetnames:
        raise SystemExit(f"{path} 缺少 sheet: 待审核候选内容")
    ws = wb["待审核候选内容"]
    headers = [_clean(cell.value) for cell in ws[1]]
    idx = {name: i for i, name in enumerate(headers)}
    result: list[GapCandidate] = []
    seen: set[tuple[str, str, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: row[i] if i < len(row) else "" for h, i in idx.items()}
        priority = _clean(d.get("优先级"))
        if priority not in priorities:
            continue
        gene = _clean(d.get("基因")).upper()
        c_hgvs = _clean(d.get("cHGVS"))
        p_hgvs = _clean(d.get("pHGVS"))
        key = (gene, c_hgvs, p_hgvs)
        if key in seen:
            continue
        seen.add(key)
        result.append(
            GapCandidate(
                priority=priority,
                gene=gene,
                c_hgvs=c_hgvs,
                p_hgvs=p_hgvs,
                class_label=_clean(d.get("等级")),
                current_status=_clean(d.get("当前状态")),
                base_intro=_clean(d.get("候选基因简介")),
                base_analysis=_clean(d.get("候选变异解析")),
            )
        )
    return result


def _iter_docx(corpus_dir: Path) -> list[Path]:
    return sorted(
        p
        for p in corpus_dir.rglob("*.docx")
        if not p.name.startswith("._") and not p.name.startswith("~$")
    )


def _source_info(path: Path, corpus_dir: Path) -> dict[str, Any]:
    try:
        relative_path = path.relative_to(corpus_dir)
    except ValueError:
        relative_path = Path(path.name)
    product_family = _infer_product_family(path.name)
    size_match = PANEL_SIZE_RE.search(product_family)
    return {
        "source_id": _stable_id(relative_path),
        "content_hash": _file_hash(path),
        "product_family": product_family,
        "panel_size": int(size_match.group(1)) if size_match else "",
    }


def harvest(
    *,
    gap_xlsx: Path,
    corpus_dir: Path,
    output: Path,
    priorities: set[str],
    max_gene_hits: int,
    max_exact_hits: int,
) -> dict[str, Any]:
    candidates = _load_gap_candidates(gap_xlsx, priorities)
    if not candidates:
        raise SystemExit("缺口表中没有符合优先级的候选。")

    by_gene: dict[str, list[GapCandidate]] = defaultdict(list)
    for candidate in candidates:
        by_gene[candidate.gene].append(candidate)

    exact_rows: list[list[Any]] = []
    gene_rows: list[list[Any]] = []
    exact_seen: Counter[tuple[str, str, str]] = Counter()
    gene_seen: Counter[str] = Counter()
    docx_files = _iter_docx(corpus_dir)

    for docx in docx_files:
        paragraphs = _docx_paragraphs(docx)
        if not paragraphs:
            continue
        body = "\n".join(paragraphs)
        present_genes = [gene for gene in by_gene if gene in body]
        if not present_genes:
            continue

        source = _source_info(docx, corpus_dir)
        for gene in present_genes:
            for cand in by_gene[gene]:
                key = (cand.gene, cand.c_hgvs, cand.p_hgvs)
                if exact_seen[key] >= max_exact_hits:
                    continue
                search_terms = [cand.c_hgvs]
                if cand.p_hgvs:
                    search_terms.append(cand.p_hgvs)
                for idx, text in enumerate(paragraphs):
                    if cand.gene not in text:
                        continue
                    if not any(term and term in text for term in search_terms):
                        continue
                    exact_rows.append(
                        [
                            cand.priority,
                            cand.gene,
                            cand.c_hgvs,
                            cand.p_hgvs,
                            "精确位点候选",
                            source["source_id"],
                            source["content_hash"],
                            source["product_family"],
                            source["panel_size"],
                            text,
                            _candidate_section(paragraphs, idx),
                            "待审核",
                            "",
                        ]
                    )
                    exact_seen[key] += 1
                    break

            if gene_seen[gene] >= max_gene_hits:
                continue
            for idx, text in enumerate(paragraphs):
                if gene_seen[gene] >= max_gene_hits:
                    break
                if not (_looks_like_intro(gene, text) or _looks_like_analysis(gene, text)):
                    continue
                candidate_type = "基因简介候选" if _looks_like_intro(gene, text) else "基因解析候选"
                gene_rows.append(
                    [
                        gene,
                        candidate_type,
                        source["source_id"],
                        source["content_hash"],
                        source["product_family"],
                        source["panel_size"],
                        text,
                        text,
                        "待审核",
                        "",
                    ]
                )
                gene_seen[gene] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "候选汇总"
    summary_rows = [
        ["缺口表", gap_xlsx.name],
        ["历史报告目录", corpus_dir.name],
        ["纳入优先级", ", ".join(sorted(priorities))],
        ["待处理位点数", len(candidates)],
        ["扫描终版报告数", len(docx_files)],
        ["精确位点候选数", len(exact_rows)],
        ["基因级候选数", len(gene_rows)],
        ["说明", "本表只供报告组审核，不会自动进入生产知识库。"],
    ]
    _append_table(ws, ["项目", "结果"], summary_rows)

    ws2 = wb.create_sheet("需补库位点")
    _append_table(
        ws2,
        [
            "优先级",
            "基因",
            "cHGVS",
            "pHGVS",
            "等级",
            "当前状态",
            "基础候选简介",
            "基础候选解析",
            "审核结论",
            "备注",
        ],
        [
            [
                c.priority,
                c.gene,
                c.c_hgvs,
                c.p_hgvs,
                c.class_label,
                c.current_status,
                c.base_intro,
                c.base_analysis,
                "待审核",
                "",
            ]
            for c in candidates
        ],
    )

    ws3 = wb.create_sheet("历史精确位点候选")
    _append_table(
        ws3,
        [
            "优先级",
            "基因",
            "cHGVS",
            "pHGVS",
            "候选类型",
            "source_id",
            "content_hash",
            "产品族",
            "基因数",
            "命中段落",
            "候选上下文",
            "审核结论",
            "备注",
        ],
        exact_rows,
    )

    ws4 = wb.create_sheet("历史基因级候选")
    _append_table(
        ws4,
        [
            "基因",
            "候选类型",
            "source_id",
            "content_hash",
            "产品族",
            "基因数",
            "命中段落",
            "候选上下文",
            "审核结论",
            "备注",
        ],
        gene_rows,
    )

    ws5 = wb.create_sheet("审核说明")
    _append_table(
        ws5,
        ["项目", "说明"],
        [
            ["怎么用", "报告组优先看“需补库位点”，再看是否有历史精确位点候选。没有精确候选时，看基因级候选。"],
            ["来源说明", "source_id/content_hash 为去标识来源，不显示患者姓名、样本号、完整文件名或报告原文之外的身份信息。"],
            ["审核通过", "确认内容适用于结直肠癌358报告后，在“审核结论”填“通过”。"],
            ["审核不通过", "填“不通过”并写原因，例如证据过旧、癌种不适用、措辞不适合复用。"],
            ["入库原则", "通过后再写入 reviewed_part3_knowledge.yaml；未审核内容不得进生产报告。"],
            ["复测", "入库后重新跑同批Excel，检查固定套话风险是否下降、报告段落是否可用。"],
        ],
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "gap_candidates": len(candidates),
        "docx_files": len(docx_files),
        "exact_candidates": len(exact_rows),
        "gene_candidates": len(gene_rows),
        "output": str(output),
    }


def _append_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for col in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for cell in col:
                max_len = max(max_len, min(len(str(cell.value or "")), 80))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 46)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--gap-xlsx", type=Path, required=True)
    parser.add_argument("--corpus-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--priorities",
        default="P0,P1,P2",
        help="Comma-separated priority labels from the gap workbook.",
    )
    parser.add_argument("--max-gene-hits", type=int, default=8)
    parser.add_argument("--max-exact-hits", type=int, default=5)
    args = parser.parse_args()
    priorities = {p.strip() for p in args.priorities.split(",") if p.strip()}
    result = harvest(
        gap_xlsx=args.gap_xlsx,
        corpus_dir=args.corpus_dir,
        output=args.output,
        priorities=priorities,
        max_gene_hits=args.max_gene_hits,
        max_exact_hits=args.max_exact_hits,
    )
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
