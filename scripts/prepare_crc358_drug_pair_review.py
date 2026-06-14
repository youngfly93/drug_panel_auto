#!/usr/bin/env python3
"""Prepare CRC358 drug-interpretation candidate pairs for medical review.

This writes review artifacts only. It does not merge unreviewed drug text into
the production CRC358 knowledge overlay.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path(
    "tmp/knowledge_buildout_after_batch5_drugclass_20260614/"
    "CRC358_医学知识库候选审核表_v0.1.xlsx"
)
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch5_drugclass_20260614")
REQUIRED_CONTENT_TYPES = {"drug_relation", "drug_clinical"}
PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"送检者"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]
SAMPLE_CONTEXT_PATTERNS = [
    re.compile(r"该样本同时检出"),
    re.compile(r"同时检出[A-Z0-9/+-]+基因"),
]


def clean(value: Any) -> str:
    return str(value or "").strip()


def has_pii(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def has_sample_specific_context(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in SAMPLE_CONTEXT_PATTERNS)


def pair_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    return (
        clean(row.get("gene")).upper(),
        clean(row.get("c_hgvs")),
        clean(row.get("p_hgvs")),
        clean(row.get("drug_type") or "benefit"),
        clean(row.get("drug_name")),
    )


def read_candidate_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "候选审核表" not in wb.sheetnames:
        raise ValueError("workbook missing sheet: 候选审核表")
    ws = wb["候选审核表"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        row = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        if row.get("content_type") in REQUIRED_CONTENT_TYPES:
            out.append(row)
    return out


def is_uncovered_drug_candidate(row: dict[str, str]) -> bool:
    if row.get("current_reviewed_status") != "暂无reviewed覆盖":
        return False
    if row.get("content_type") not in REQUIRED_CONTENT_TYPES:
        return False
    if row.get("source_type") and row.get("source_type") != "historical_final_report":
        return False
    return True


def cleanup_reason(row: dict[str, str]) -> str:
    reasons = []
    if not row.get("drug_name") or row.get("drug_name") == "--":
        reasons.append("缺少明确药名")
    if not row.get("gene"):
        reasons.append("缺少基因")
    if not (row.get("c_hgvs") and row.get("p_hgvs")):
        reasons.append("缺少完整 c/p 位点")
    if has_pii(row.get("candidate_text", "")):
        reasons.append("疑似包含PII")
    if has_sample_specific_context(row.get("candidate_text", "")):
        reasons.append("包含同样本其他变异上下文")
    return "；".join(reasons) or "未形成 relation/clinical 成对规则"


def split_drug_pairs(rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Split uncovered drug rows into complete reviewable pairs and cleanup rows."""

    candidates = [row for row in rows if is_uncovered_drug_candidate(row)]
    grouped: dict[tuple[str, str, str, str, str], dict[str, list[dict[str, str]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    cleanup: list[dict[str, Any]] = []

    for row in candidates:
        reason = cleanup_reason(row)
        if reason != "未形成 relation/clinical 成对规则":
            cleanup.append(
                {
                    "candidate_id": row.get("candidate_id", ""),
                    "gene": row.get("gene", ""),
                    "c_hgvs": row.get("c_hgvs", ""),
                    "p_hgvs": row.get("p_hgvs", ""),
                    "content_type": row.get("content_type", ""),
                    "drug_type": row.get("drug_type", ""),
                    "drug_name": row.get("drug_name", ""),
                    "整理原因": reason,
                    "candidate_text": row.get("candidate_text", ""),
                }
            )
            continue
        grouped[pair_key(row)][row["content_type"]].append(row)

    reviewable: list[dict[str, Any]] = []
    for key, by_type in sorted(grouped.items()):
        gene, c_hgvs, p_hgvs, drug_type, drug_name = key
        relation_rows = by_type.get("drug_relation", [])
        clinical_rows = by_type.get("drug_clinical", [])
        if not relation_rows or not clinical_rows:
            for row in relation_rows + clinical_rows:
                cleanup.append(
                    {
                        "candidate_id": row.get("candidate_id", ""),
                        "gene": row.get("gene", ""),
                        "c_hgvs": row.get("c_hgvs", ""),
                        "p_hgvs": row.get("p_hgvs", ""),
                        "content_type": row.get("content_type", ""),
                        "drug_type": row.get("drug_type", ""),
                        "drug_name": row.get("drug_name", ""),
                        "整理原因": "缺少配套的药物关联分析或临床解析",
                        "candidate_text": row.get("candidate_text", ""),
                    }
                )
            continue
        relation = relation_rows[0]
        clinical = clinical_rows[0]
        reviewable.append(
            {
                "gene": gene,
                "c_hgvs": c_hgvs,
                "p_hgvs": p_hgvs,
                "drug_type": drug_type,
                "drug_name": drug_name,
                "relation_candidate_id": relation.get("candidate_id", ""),
                "clinical_candidate_id": clinical.get("candidate_id", ""),
                "source_count_relation": relation.get("source_count", ""),
                "source_count_clinical": clinical.get("source_count", ""),
                "confidence_relation": relation.get("confidence", ""),
                "confidence_clinical": clinical.get("confidence", ""),
                "relation": relation.get("candidate_text", ""),
                "clinical": clinical.get("candidate_text", ""),
                "review_status": "待医学审核",
                "review_notes": "",
            }
        )
    return reviewable, cleanup


def build_pending_overlay(reviewable: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "source_type": "historical_final_report_drug_pair_review",
            "status": "pending_medical_review",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "merge_policy": "review-only artifact; do not merge into production before medical approval",
        },
        "drug_sections": [
            {
                "gene": row["gene"],
                "c_hgvs": row["c_hgvs"],
                "p_hgvs": row["p_hgvs"],
                "type": row["drug_type"] or "benefit",
                "header": f'{row["gene"]}：{row["c_hgvs"]}，{row["p_hgvs"]}突变相应靶向药物',
                "drug_name": row["drug_name"],
                "relation": row["relation"],
                "clinical": row["clinical"],
            }
            for row in reviewable
        ],
    }


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_review_workbook(
    path: Path,
    reviewable: list[dict[str, Any]],
    cleanup: list[dict[str, Any]],
    overlay: dict[str, Any],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(
        ws,
        [
            {"项目": "定位", "说明": "CRC358 药物解析成对审核包；不是生产入库文件。"},
            {"项目": "可成对审核", "说明": len(reviewable)},
            {"项目": "需人工整理", "说明": len(cleanup)},
            {"项目": "成对条件", "说明": "同一 gene/c_hgvs/p_hgvs/drug_type/drug_name 同时有 relation 和 clinical。"},
            {"项目": "安全边界", "说明": "缺药名、缺位点、疑似PII、只有半段解析的条目均不进入待审 overlay。"},
        ],
    )
    ws = wb.create_sheet("可成对审核")
    write_sheet(ws, reviewable)
    ws = wb.create_sheet("缺药名或缺位点")
    write_sheet(ws, cleanup)
    ws = wb.create_sheet("待审overlay预览")
    write_sheet(ws, overlay.get("drug_sections") or [])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    rows = read_candidate_rows(args.input)
    reviewable, cleanup = split_drug_pairs(rows)
    overlay = build_pending_overlay(reviewable)
    review_xlsx = args.out_dir / "CRC358_batch5_药物解析成对审核包_20260614.xlsx"
    overlay_yaml = args.out_dir / "reviewed_part3_drug_pairs_pending_review_batch5.yaml"
    write_review_workbook(review_xlsx, reviewable, cleanup, overlay)
    dump_yaml(overlay_yaml, overlay)
    print(f"reviewable_pairs={len(reviewable)} cleanup_rows={len(cleanup)}")
    print(f"review_xlsx={review_xlsx}")
    print(f"pending_overlay={overlay_yaml}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
