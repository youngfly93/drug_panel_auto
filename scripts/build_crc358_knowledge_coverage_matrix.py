#!/usr/bin/env python3
"""Build a CRC358 production knowledge coverage matrix.

The script is read-only for production knowledge bases. It extracts the 358
gene list from the approved CRC358 golden template, compares it with the
current reviewed Part3 overlay, targeted drug table, immune gene table, and
CRC rule overrides, then writes review artifacts under ``tmp/``.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_TEMPLATE = Path("panels/crc_358_msi/templates/crc_358_msi_golden_template_v0.docx")
DEFAULT_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_CRC_RULES = Path("panels/crc_358_msi/rules/crc.yaml")
DEFAULT_TARGETED_DB = Path("data/knowledge_bases/processed/targeted_drug_db_public.xlsx")
DEFAULT_IMMUNE_DB = Path("data/knowledge_bases/processed/immune_gene_list_public.xlsx")
DEFAULT_GENE_DB = Path("data/knowledge_bases/processed/gene_knowledge_db.xlsx")
DEFAULT_OUT_DIR = Path("tmp/knowledge_coverage")

LOW_INFO_PATTERNS = (
    "具体临床价值需结合",
    "潜在意义",
    "可能影响",
    "综合判断",
    "肿瘤发生发展",
    "证据等级",
    "相关治疗策略",
    "有待明确",
    "尚未见",
    "较少",
)
SPECIAL_RULE_GENES = {"KRAS", "NRAS", "BRAF", "NTRK1", "NTRK2", "NTRK3", "ERBB2", "EGFR", "TSC1", "TP53"}


def clean(value: Any) -> str:
    return str(value or "").strip()


def norm_gene(value: Any) -> str:
    return clean(value).upper()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def docx_table_cells(path: Path) -> list[list[str]]:
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8", "ignore")
    tables: list[list[str]] = []
    for tbl in re.findall(r"<w:tbl[ >].*?</w:tbl>", xml, re.S):
        cells: list[str] = []
        for tc in re.findall(r"<w:tc[ >].*?</w:tc>", tbl, re.S):
            texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", tc, re.S)
            text = "".join(re.sub(r"<[^>]+>", "", item) for item in texts)
            cells.append(html.unescape(text).strip())
        tables.append(cells)
    return tables


def extract_crc358_panel_genes(template_path: Path) -> list[str]:
    for cells in docx_table_cells(template_path):
        if cells and "Gene List for MLseq" in cells[0]:
            genes = [norm_gene(cell) for cell in cells[1:] if norm_gene(cell)]
            duplicate_genes = [gene for gene, count in Counter(genes).items() if count > 1]
            if duplicate_genes:
                raise ValueError(f"Duplicated genes in template gene list: {duplicate_genes[:10]}")
            return genes
    raise ValueError(f"Gene List for MLseq table not found in {template_path}")


def low_info_hits(text: str) -> list[str]:
    return [pattern for pattern in LOW_INFO_PATTERNS if pattern in clean(text)]


def overlay_indexes(path: Path) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    data = load_yaml(path)
    gene_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    drug_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in data.get("gene_sections") or []:
        gene = norm_gene(row.get("gene"))
        if gene:
            gene_rows[gene].append(row)
    for row in data.get("drug_sections") or []:
        gene = norm_gene(row.get("gene"))
        if gene:
            drug_rows[gene].append(row)
    return gene_rows, drug_rows


def crc_rule_indexes(path: Path) -> dict[str, Any]:
    data = load_yaml(path)
    class_i = {norm_gene(gene) for gene in data.get("class_i_genes") or []}
    class_ii = {norm_gene(gene) for gene in data.get("class_ii_genes") or []}
    important = {norm_gene(gene) for gene in data.get("crc_important_genes") or []}
    display = {norm_gene(row.get("name")) for row in data.get("panel_display_genes") or [] if row.get("name")}
    overrides: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in data.get("reviewed_variant_overrides") or []:
        gene = norm_gene(row.get("gene"))
        if gene:
            overrides[gene].append(row)
    return {
        "class_i": class_i,
        "class_ii": class_ii,
        "important": important,
        "display": display,
        "overrides": overrides,
    }


def targeted_drug_indexes(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, Any] = {
        "exact": defaultdict(list),
        "internal": defaultdict(list),
        "public": defaultdict(list),
        "composite_mentions": defaultdict(list),
        "sheet_rows": {},
    }
    sheet_aliases = {
        "targeted_drug_tips": "exact",
        "internal_targeted_drug_tips": "internal",
        "public_targeted_drug_tips": "public",
    }
    for sheet_name, bucket in sheet_aliases.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [clean(value) for value in rows[0]]
        result["sheet_rows"][sheet_name] = max(0, len(rows) - 1)
        gene_idx = headers.index("基因名称") if "基因名称" in headers else 0
        for excel_row, raw in enumerate(rows[1:], start=2):
            gene_value = norm_gene(raw[gene_idx] if gene_idx < len(raw) else "")
            if not gene_value:
                continue
            row = {
                "excel_row": excel_row,
                "sheet": sheet_name,
                "gene_value": gene_value,
                "raw": raw,
                "headers": headers,
            }
            if "/" in gene_value:
                for part in gene_value.split("/"):
                    part = norm_gene(part)
                    if part:
                        result["composite_mentions"][part].append(row)
            else:
                result[bucket][gene_value].append(row)
    wb.close()
    return result


def immune_indexes(path: Path) -> dict[str, set[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"positive": set(), "negative": set(), "hyper": set()}
    if "immune_gene_list" not in wb.sheetnames:
        return result
    ws = wb["immune_gene_list"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) > 0 and norm_gene(row[0]):
            result["positive"].add(norm_gene(row[0]))
        if len(row) > 1 and norm_gene(row[1]):
            result["negative"].add(norm_gene(row[1]))
        if len(row) > 2 and norm_gene(row[2]):
            result["hyper"].add(norm_gene(row[2]))
    wb.close()
    return result


def base_gene_db_index(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    genes: set[str] = set()
    if "基因变异解析" in wb.sheetnames:
        ws = wb["基因变异解析"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            gene = norm_gene(row[0] if row else "")
            if gene and re.match(r"^[A-Z0-9-]+$", gene):
                genes.add(gene)
    wb.close()
    return genes


def yes(value: bool) -> str:
    return "是" if value else "否"


def recommend_action(
    gene: str,
    has_overlay: bool,
    has_intro: bool,
    has_mutation: bool,
    low_info_count: int,
    has_base: bool,
    is_critical: bool,
    is_special: bool,
    targeted_exact_count: int,
) -> tuple[str, str, str]:
    if not has_overlay and is_critical:
        return "P0", "关键基因缺少 reviewed Part3 覆盖", "优先补基因简介和 CRC 语境变异解析"
    if is_special:
        return "P0", "药物/免疫规则需结构化确认", "核对触发条件，必要时拆成位点/变异类型/组合规则"
    if not has_overlay and not has_base:
        return "P1", "panel 基因无生产解析覆盖", "补 gene-level reviewed 简介与基础解析"
    if not has_overlay:
        return "P2", "仅基础库覆盖，未进入 reviewed overlay", "评估是否需要升级为 CRC reviewed 文案"
    if not has_intro or not has_mutation:
        return "P1", "reviewed 条目字段不完整", "补齐 intro/mutation_analysis"
    if low_info_count:
        return "P1", "存在低信息量/套话风险", "用更具体的通路、CRC 证据或位点意义替换"
    if is_critical and targeted_exact_count == 0:
        return "P2", "关键基因无精确用药主表行", "确认是否应有用药提示或仅作为 Part3 解释"
    return "", "", "保持现状，后续随真实样本补库"


def build_matrix(args: argparse.Namespace) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    panel_genes = extract_crc358_panel_genes(args.template)
    gene_overlay, drug_overlay = overlay_indexes(args.overlay)
    rules = crc_rule_indexes(args.crc_rules)
    targeted = targeted_drug_indexes(args.targeted_db)
    immune = immune_indexes(args.immune_db)
    base_genes = base_gene_db_index(args.gene_db)
    panel_set = set(panel_genes)

    matrix: list[dict[str, Any]] = []
    low_info_rows: list[dict[str, Any]] = []
    rule_todos: list[dict[str, Any]] = []

    for index, gene in enumerate(panel_genes, start=1):
        rows = gene_overlay.get(gene, [])
        has_overlay = bool(rows)
        has_intro = any(clean(row.get("intro")) for row in rows)
        has_mutation = any(clean(row.get("mutation_analysis")) for row in rows)
        variant_sections = [row for row in rows if clean(row.get("c_hgvs")) or clean(row.get("p_hgvs"))]
        low_rows = [(row, low_info_hits(row.get("mutation_analysis") or "")) for row in rows]
        low_rows = [(row, hits) for row, hits in low_rows if len(hits) >= 3]
        phgvs_blank = [row for row in rows if clean(row.get("c_hgvs")) and not clean(row.get("p_hgvs"))]
        target_exact = targeted["exact"].get(gene, [])
        internal_rows = targeted["internal"].get(gene, [])
        public_rows = targeted["public"].get(gene, [])
        composite_rows = targeted["composite_mentions"].get(gene, [])
        override_rows = rules["overrides"].get(gene, [])
        is_critical = gene in rules["class_i"] or gene in rules["class_ii"] or gene in rules["important"]
        is_special = gene in SPECIAL_RULE_GENES
        priority, gap_reason, action = recommend_action(
            gene=gene,
            has_overlay=has_overlay,
            has_intro=has_intro,
            has_mutation=has_mutation,
            low_info_count=len(low_rows),
            has_base=gene in base_genes,
            is_critical=is_critical,
            is_special=is_special,
            targeted_exact_count=len(target_exact),
        )
        if has_overlay and has_intro and has_mutation and not low_rows:
            coverage_level = "reviewed覆盖"
        elif has_overlay:
            coverage_level = "reviewed需完善"
        elif gene in base_genes:
            coverage_level = "仅基础库覆盖"
        else:
            coverage_level = "缺口"
        matrix.append(
            {
                "序号": index,
                "gene": gene,
                "Part3覆盖等级": coverage_level,
                "优先级": priority,
                "问题原因": gap_reason,
                "建议动作": action,
                "class_i": yes(gene in rules["class_i"]),
                "class_ii": yes(gene in rules["class_ii"]),
                "CRC重点基因": yes(gene in rules["important"]),
                "报告展示基因": yes(gene in rules["display"]),
                "reviewed_gene_sections": len(rows),
                "reviewed_variant_sections": len(variant_sections),
                "有基因简介": yes(has_intro),
                "有变异解析": yes(has_mutation),
                "低信息量条目数": len(low_rows),
                "c有_p空条目数": len(phgvs_blank),
                "基础库覆盖": yes(gene in base_genes),
                "Part3药物解析条数": len(drug_overlay.get(gene, [])),
                "用药主表精确行数": len(target_exact),
                "internal行数": len(internal_rows),
                "public行数": len(public_rows),
                "合并语义行提及": len(composite_rows),
                "override条数": len(override_rows),
                "免疫正相关": yes(gene in immune["positive"]),
                "免疫负相关": yes(gene in immune["negative"]),
                "免疫超进展": yes(gene in immune["hyper"]),
                "审核状态": "待审核" if priority else "",
                "审核意见": "",
            }
        )
        for row, hits in low_rows:
            low_info_rows.append(
                {
                    "gene": gene,
                    "c_hgvs": clean(row.get("c_hgvs")),
                    "p_hgvs": clean(row.get("p_hgvs")),
                    "命中套话数": len(hits),
                    "命中套话": "；".join(hits),
                    "当前基因简介": clean(row.get("intro")),
                    "当前变异解析": clean(row.get("mutation_analysis")),
                    "建议": "报告组判断是否替换为更具体的 CRC/通路/位点解释",
                    "审核状态": "待审核",
                    "审核意见": "",
                }
            )
        if is_special:
            rule_todos.append(
                {
                    "优先级": "P0",
                    "gene": gene,
                    "规则化主题": special_rule_topic(gene),
                    "当前证据": f"主表精确行 {len(target_exact)}；override {len(override_rows)}；合并语义提及 {len(composite_rows)}",
                    "建议动作": special_rule_action(gene),
                    "审核状态": "待审核",
                    "审核意见": "",
                }
            )

    overlay_genes = set(gene_overlay)
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template": str(args.template),
        "panel_gene_count": len(panel_genes),
        "panel_gene_unique_count": len(panel_set),
        "reviewed_overlay_gene_count_total": len(overlay_genes),
        "reviewed_overlay_gene_count_in_panel": len(overlay_genes & panel_set),
        "reviewed_overlay_gene_count_outside_panel": len(overlay_genes - panel_set),
        "panel_genes_without_reviewed_overlay": sum(1 for row in matrix if row["reviewed_gene_sections"] == 0),
        "panel_genes_without_any_gene_text": sum(1 for row in matrix if row["Part3覆盖等级"] == "缺口"),
        "panel_genes_low_info": sum(1 for row in matrix if row["低信息量条目数"] > 0),
        "panel_genes_with_targeted_drug_rows": sum(1 for row in matrix if row["用药主表精确行数"] > 0),
        "panel_genes_with_immune_tags": sum(
            1
            for row in matrix
            if row["免疫正相关"] == "是" or row["免疫负相关"] == "是" or row["免疫超进展"] == "是"
        ),
        "targeted_sheet_rows": targeted["sheet_rows"],
        "priority_counts": dict(Counter(row["优先级"] or "无" for row in matrix)),
        "coverage_level_counts": dict(Counter(row["Part3覆盖等级"] for row in matrix)),
    }
    return matrix, summary, low_info_rows, rule_todos


def special_rule_topic(gene: str) -> str:
    if gene in {"KRAS", "NRAS", "BRAF"}:
        return "RAS/RAF 组合野生型、任一突变、BRAF V600E/non-V600E"
    if gene in {"NTRK1", "NTRK2", "NTRK3"}:
        return "NTRK 融合限定"
    if gene == "ERBB2":
        return "ERBB2 扩增 vs 点突变"
    if gene == "EGFR":
        return "EGFR 位点级靶药/免疫/超进展分流"
    if gene == "TSC1":
        return "TSC1 loss-of-function 用药触发"
    if gene == "TP53":
        return "TP53 药物提示触发边界"
    return "特殊规则"


def special_rule_action(gene: str) -> str:
    if gene in {"KRAS", "NRAS", "BRAF"}:
        return "拆成组合条件规则，避免把单基因行误当作抗EGFR获益/耐药完整逻辑"
    if gene in {"NTRK1", "NTRK2", "NTRK3"}:
        return "仅融合阳性触发拉罗替尼/恩曲替尼等相关提示，普通 SNV 不自动触发"
    if gene == "ERBB2":
        return "扩增、点突变、融合分别建条件，且标记 CRC 适用范围"
    if gene == "EGFR":
        return "免疫负相关仅保留 L858R/EX19del；超进展仅保留扩增；其他位点走靶药/Part3"
    if gene == "TSC1":
        return "确认无义/移码/剪接等 LOF 才触发 mTOR 相关药物"
    if gene == "TP53":
        return "确认是否按 II 类/特定位点/功能缺失触发药物，不泛化所有 TP53"
    return "补结构化适用条件"


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict[str, Any]], widths: dict[str, int] | None = None) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="00A9B7")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for body_row in ws.iter_rows(min_row=2):
        for cell in body_row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(header) * 2, 10), 28)
    ws.sheet_view.showGridLines = False


def write_outputs(
    matrix: list[dict[str, Any]],
    summary: dict[str, Any],
    low_info_rows: list[dict[str, Any]],
    rule_todos: list[dict[str, Any]],
    out_dir: Path,
) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    xlsx_path = out_dir / f"CRC358_知识库覆盖矩阵_{stamp}.xlsx"
    json_path = out_dir / f"crc358_knowledge_coverage_summary_{stamp}.json"
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = [{"指标": key, "结果": json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value} for key, value in summary.items()]
    add_sheet(wb, "覆盖汇总", ["指标", "结果"], summary_rows, {"A": 36, "B": 90})
    matrix_headers = [
        "序号",
        "gene",
        "Part3覆盖等级",
        "优先级",
        "问题原因",
        "建议动作",
        "class_i",
        "class_ii",
        "CRC重点基因",
        "报告展示基因",
        "reviewed_gene_sections",
        "reviewed_variant_sections",
        "有基因简介",
        "有变异解析",
        "低信息量条目数",
        "c有_p空条目数",
        "基础库覆盖",
        "Part3药物解析条数",
        "用药主表精确行数",
        "internal行数",
        "public行数",
        "合并语义行提及",
        "override条数",
        "免疫正相关",
        "免疫负相关",
        "免疫超进展",
        "审核状态",
        "审核意见",
    ]
    add_sheet(
        wb,
        "358基因覆盖矩阵",
        matrix_headers,
        matrix,
        {"A": 8, "B": 13, "C": 16, "D": 10, "E": 26, "F": 42, "G": 10, "H": 10, "I": 12, "J": 12},
    )
    priority_rows = [row for row in matrix if row["优先级"]]
    priority_rows.sort(key=lambda row: ({"P0": 0, "P1": 1, "P2": 2}.get(row["优先级"], 9), row["gene"]))
    add_sheet(wb, "补库优先级", matrix_headers, priority_rows)
    rule_headers = ["优先级", "gene", "规则化主题", "当前证据", "建议动作", "审核状态", "审核意见"]
    add_sheet(wb, "药物规则化待办", rule_headers, rule_todos, {"A": 10, "B": 12, "C": 42, "D": 36, "E": 62, "F": 12, "G": 34})
    low_headers = ["gene", "c_hgvs", "p_hgvs", "命中套话数", "命中套话", "当前基因简介", "当前变异解析", "建议", "审核状态", "审核意见"]
    add_sheet(wb, "低信息量条目", low_headers, low_info_rows, {"A": 12, "B": 18, "C": 18, "D": 12, "E": 34, "F": 60, "G": 78, "H": 42, "I": 12, "J": 34})
    source_rows = [
        {"来源": "panel gene list", "路径": summary["template"], "说明": "从 Gene List for MLseq (n=358) 表提取"},
        {"来源": "reviewed Part3 overlay", "路径": str(DEFAULT_OVERLAY), "说明": "生产 reviewed Part3 文案"},
        {"来源": "targeted drug DB", "路径": str(DEFAULT_TARGETED_DB), "说明": "当前用药提示主表与 public/internal 源表"},
        {"来源": "immune gene list", "路径": str(DEFAULT_IMMUNE_DB), "说明": "免疫正/负/超进展基因标签"},
        {"来源": "CRC rules", "路径": str(DEFAULT_CRC_RULES), "说明": "重点基因、展示基因、override"},
    ]
    add_sheet(wb, "来源与说明", ["来源", "路径", "说明"], source_rows, {"A": 24, "B": 88, "C": 68})
    wb.save(xlsx_path)
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return xlsx_path, json_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--overlay", type=Path, default=DEFAULT_OVERLAY)
    parser.add_argument("--crc-rules", type=Path, default=DEFAULT_CRC_RULES)
    parser.add_argument("--targeted-db", type=Path, default=DEFAULT_TARGETED_DB)
    parser.add_argument("--immune-db", type=Path, default=DEFAULT_IMMUNE_DB)
    parser.add_argument("--gene-db", type=Path, default=DEFAULT_GENE_DB)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    matrix, summary, low_info_rows, rule_todos = build_matrix(args)
    xlsx_path, json_path = write_outputs(matrix, summary, low_info_rows, rule_todos, args.out_dir)
    print(json.dumps({"xlsx": str(xlsx_path), "json": str(json_path), **summary}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
