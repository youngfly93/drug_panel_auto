#!/usr/bin/env python3
"""Prepare a CRC358 subset release workbook from machine triage results.

Only rows with ``machine_suggestion == 建议通过`` are marked as ``通过``.
All other rows are marked ``暂缓`` so release gates can validate a clean,
auditable subset without silently promoting rows that still need focused review.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TRIAGE_WORKBOOK = Path(
    "tmp/knowledge_buildout_after_batch12_medical_triage_20260614/CRC358_batch12_机器医学质控分流_20260614.xlsx"
)
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch13_subset_release_20260614")
FULL_REVIEW_SHEETS = ("新增gene完整审核", "新增drug完整审核")


def clean(value: Any) -> str:
    return str(value or "").strip()


def read_triage_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "质控总表" not in wb.sheetnames:
        raise ValueError(f"triage workbook missing sheet: 质控总表 ({path})")
    ws = wb["质控总表"]
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [clean(value) for value in raw_rows[0]]
    rows: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        row = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        rows.append(row)
    return rows


def should_approve(
    row: dict[str, str],
    *,
    approve_cross_cancer_gene_intros: bool = False,
    approve_historical_drug_pairs: bool = False,
) -> bool:
    if clean(row.get("machine_suggestion")) == "建议通过":
        return True
    if clean(row.get("risk_flags")):
        return False
    section = clean(row.get("section"))
    source = clean(row.get("review_source"))
    if approve_cross_cancer_gene_intros:
        if source == "all_cancer_final_report_gene_intro_support" and section == "gene_sections" and clean(row.get("intro")):
            return True
    if approve_historical_drug_pairs:
        if (
            source == "historical_final_report_drug_pair_review"
            and section == "drug_sections"
            and clean(row.get("drug_name"))
            and clean(row.get("relation"))
            and clean(row.get("clinical"))
        ):
            return True
    return False


def release_row(
    row: dict[str, str],
    *,
    approve_cross_cancer_gene_intros: bool = False,
    approve_historical_drug_pairs: bool = False,
    reject_historical_drug_pairs: bool = False,
) -> dict[str, str]:
    out = dict(row)
    suggestion = clean(row.get("machine_suggestion"))
    if should_approve(
        row,
        approve_cross_cancer_gene_intros=approve_cross_cancer_gene_intros,
        approve_historical_drug_pairs=approve_historical_drug_pairs,
    ):
        out["review_status"] = "通过"
        if suggestion == "建议通过":
            default_note = "机器质控子集放行：基础知识库支撑且无自动质控风险"
        elif clean(row.get("review_source")) == "all_cancer_final_report_gene_intro_support":
            default_note = "机器质控子集放行：跨癌种历史终版基因简介，无自动质控风险"
        else:
            default_note = "机器质控子集放行：历史终版药物成对内容完整且无自动质控风险"
        out["review_notes"] = clean(row.get("review_notes")) or default_note
    elif reject_historical_drug_pairs and clean(row.get("review_source")) == "historical_final_report_drug_pair_review":
        out["review_status"] = "不入库"
        gene = clean(row.get("gene")).upper()
        if gene == "MAP2K4":
            default_note = (
                "最终处置：不入库。该候选主要为临床前细胞系证据，药物获批适应证不直接面向CRC/MAP2K4位点；"
                "通用入库会导致用药提示过宽。"
            )
        elif gene == "SDHB":
            default_note = (
                "最终处置：不入库。该候选药物解析依赖GIST/SDH缺陷语境，CRC358通用位点知识库缺少肿瘤类型/applicability限定；"
                "通用入库会导致用药提示过宽。"
            )
        elif gene == "XRCC2":
            default_note = (
                "最终处置：不入库。该候选PARP抑制剂解析依赖DNA修复缺陷/篮子试验语境，单个XRCC2位点不足以作为通用用药提示；"
                "需待HRD/适用条件规则建立后再评估。"
            )
        else:
            default_note = "最终处置：不入库。药物解析需特定适用条件，当前通用知识库无法安全限定。"
        out["review_notes"] = clean(row.get("review_notes")) or default_note
    else:
        out["review_status"] = "暂缓"
        out["review_notes"] = clean(row.get("review_notes")) or f"暂缓：{suggestion or '未给出建议'}，需人工精审后再放行"
    out.setdefault("reviewed_intro", "")
    out.setdefault("reviewed_mutation_analysis", "")
    out.setdefault("reviewed_relation", "")
    out.setdefault("reviewed_clinical", "")
    return out


def workbook_row(row: dict[str, str]) -> dict[str, str]:
    fields = [
        "action",
        "reason",
        "review_source",
        "section",
        "gene",
        "c_hgvs",
        "p_hgvs",
        "type",
        "applicability",
        "header",
        "drug_name",
        "has_intro",
        "has_mutation_analysis",
        "has_relation",
        "has_clinical",
        "intro",
        "mutation_analysis",
        "relation",
        "clinical",
        "review_status",
        "reviewed_intro",
        "reviewed_mutation_analysis",
        "reviewed_relation",
        "reviewed_clinical",
        "review_notes",
        "machine_suggestion",
        "machine_reason",
        "risk_flags",
    ]
    out = {field: clean(row.get(field)) for field in fields}
    out["action"] = "added"
    out["reason"] = clean(row.get("machine_reason"))
    out["has_intro"] = "是" if clean(row.get("intro")) else "否"
    out["has_mutation_analysis"] = "是" if clean(row.get("mutation_analysis")) else "否"
    out["has_relation"] = "是" if clean(row.get("relation")) else "否"
    out["has_clinical"] = "是" if clean(row.get("clinical")) else "否"
    return out


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(path: Path, rows: list[dict[str, str]], summary_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(ws, summary_rows)
    all_rows = [workbook_row(row) for row in rows]
    ws = wb.create_sheet("新增候选")
    write_sheet(ws, all_rows)
    ws = wb.create_sheet("新增gene完整审核")
    write_sheet(ws, [row for row in all_rows if row.get("section") == "gene_sections"])
    ws = wb.create_sheet("新增drug完整审核")
    write_sheet(ws, [row for row in all_rows if row.get("section") == "drug_sections"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_summary(rows: list[dict[str, str]], output: Path, policy: str) -> dict[str, Any]:
    return {
        "status": "machine_triage_subset_release_workbook",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "output": str(output),
        "total_rows": len(rows),
        "review_status_counts": dict(Counter(row["review_status"] for row in rows)),
        "machine_suggestion_counts": dict(Counter(row.get("machine_suggestion", "") for row in rows)),
        "section_counts": dict(Counter(row.get("section", "") for row in rows)),
        "policy": policy,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--triage-workbook", type=Path, default=DEFAULT_TRIAGE_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--batch-label", default="batch13")
    parser.add_argument("--approve-cross-cancer-gene-intros", action="store_true")
    parser.add_argument("--approve-historical-drug-pairs", action="store_true")
    parser.add_argument("--reject-historical-drug-pairs", action="store_true")
    args = parser.parse_args()
    if args.approve_historical_drug_pairs and args.reject_historical_drug_pairs:
        raise ValueError("--approve-historical-drug-pairs and --reject-historical-drug-pairs are mutually exclusive")

    rows = [
        release_row(
            row,
            approve_cross_cancer_gene_intros=args.approve_cross_cancer_gene_intros,
            approve_historical_drug_pairs=args.approve_historical_drug_pairs,
            reject_historical_drug_pairs=args.reject_historical_drug_pairs,
        )
        for row in read_triage_rows(args.triage_workbook)
    ]
    args.out_dir.mkdir(parents=True, exist_ok=True)
    output = args.out_dir / f"CRC358_{args.batch_label}_机器建议通过子集放行_20260614.xlsx"
    summary_json = args.out_dir / f"crc358_{args.batch_label}_subset_release_summary_20260614.json"
    policy_parts = ["建议通过 -> 通过"]
    if args.approve_cross_cancer_gene_intros:
        policy_parts.append("跨癌种历史终版基因简介 -> 通过")
    if args.approve_historical_drug_pairs:
        policy_parts.append("历史终版药物成对内容 -> 通过")
    if args.reject_historical_drug_pairs:
        policy_parts.append("历史终版药物成对内容 -> 不入库")
    policy_parts.append("其他 -> 暂缓，不进入生产合入")
    summary = build_summary(rows, output, "；".join(policy_parts))
    summary_rows = [
        {"项目": "定位", "结果": "机器建议通过子集放行工作簿；用于生成 release-ready overlay 草稿"},
        {"项目": "来源", "结果": str(args.triage_workbook)},
        {"项目": "策略", "结果": summary["policy"]},
        {"项目": "总行数", "结果": summary["total_rows"]},
        {"项目": "审核状态计数", "结果": json.dumps(summary["review_status_counts"], ensure_ascii=False)},
        {"项目": "机器建议计数", "结果": json.dumps(summary["machine_suggestion_counts"], ensure_ascii=False)},
        {"项目": "输出", "结果": str(output)},
    ]
    write_workbook(output, rows, summary_rows)
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"output_xlsx={output}")
    print(f"output_json={summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
