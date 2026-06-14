#!/usr/bin/env python3
"""Prepare CRC358 gene-intro/mutation-analysis gaps for medical review.

The output is a review package and a pending overlay preview. It does not write
to the production CRC358 reviewed knowledge file.
"""

from __future__ import annotations

import argparse
import difflib
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CANDIDATES = Path(
    "tmp/knowledge_buildout_after_batch5_drugclass_20260614/"
    "CRC358_医学知识库候选审核表_v0.1.xlsx"
)
DEFAULT_BASE_KB = Path("data/knowledge_bases/processed/gene_knowledge_db.xlsx")
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch6_gene_gap_20260614")
TARGET_CONTENT_TYPES = {"gene_intro", "mutation_analysis"}
PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"送检者"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]


def clean(value: Any) -> str:
    return str(value or "").strip()


def compact(text: str) -> str:
    return re.sub(r"\s+", "", clean(text))


def has_pii(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def strip_section_leak(text: str) -> str:
    """Trim text accidentally captured past the Part 3 mutation-analysis block."""
    text = clean(text)
    for marker in ("靶向药物/免疫用药提示解析", "潜在获益靶向/免疫药物解析", "潜在负相关靶向/免疫药物解析"):
        if marker in text:
            text = text.split(marker, 1)[0].strip()
    return text


def similarity(a: str, b: str) -> float:
    a_key = compact(a)
    b_key = compact(b)
    if not a_key or not b_key:
        return 0.0
    return round(difflib.SequenceMatcher(None, a_key, b_key).ratio(), 4)


def as_int(value: Any) -> int:
    try:
        return int(float(clean(value) or "0"))
    except ValueError:
        return 0


def read_candidate_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "候选审核表" not in wb.sheetnames:
        raise ValueError("workbook missing sheet: 候选审核表")
    ws = wb["候选审核表"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        row = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        if row.get("content_type") in TARGET_CONTENT_TYPES:
            out.append(row)
    return out


def load_base_gene_knowledge(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "基因变异解析" not in wb.sheetnames:
        raise ValueError("base gene KB missing sheet: 基因变异解析")
    ws = wb["基因变异解析"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [clean(value) for value in rows[0]]
    idx = {header: i for i, header in enumerate(headers)}
    required = {"基因名称", "基因简介", "基因变异解析"}
    missing = required - set(idx)
    if missing:
        raise ValueError(f"base gene KB missing columns: {sorted(missing)}")
    out: dict[str, dict[str, str]] = {}
    for raw in rows[1:]:
        gene = clean(raw[idx["基因名称"]] if idx["基因名称"] < len(raw) else "").upper()
        if not gene:
            continue
        out.setdefault(
            gene,
            {
                "base_intro": strip_section_leak(raw[idx["基因简介"]] if idx["基因简介"] < len(raw) else ""),
                "base_mutation_analysis": strip_section_leak(
                    raw[idx["基因变异解析"]] if idx["基因变异解析"] < len(raw) else ""
                ),
            },
        )
    return out


def is_relevant_gap(row: dict[str, str]) -> bool:
    return (
        row.get("kb_gap_class") == "低置信待补证据"
        and row.get("current_reviewed_status") == "暂无reviewed覆盖"
        and row.get("content_type") in TARGET_CONTENT_TYPES
        and row.get("source_type") == "historical_final_report"
        and bool(row.get("gene"))
    )


def pick_best(rows: list[dict[str, str]]) -> dict[str, str]:
    if not rows:
        return {}
    return sorted(
        rows,
        key=lambda row: (
            as_int(row.get("source_count")),
            len(row.get("candidate_text", "")),
            row.get("confidence", ""),
        ),
        reverse=True,
    )[0]


def build_gene_gap_review(
    candidate_rows: list[dict[str, str]],
    base_kb: dict[str, dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in candidate_rows:
        if is_relevant_gap(row):
            grouped[row["gene"].upper()][row["content_type"]].append(row)

    base_supported: list[dict[str, Any]] = []
    base_missing: list[dict[str, Any]] = []
    for gene in sorted(grouped):
        intro_row = pick_best(grouped[gene].get("gene_intro", []))
        analysis_row = pick_best(grouped[gene].get("mutation_analysis", []))
        base = base_kb.get(gene, {})
        candidate_intro = strip_section_leak(intro_row.get("candidate_text", ""))
        candidate_analysis = strip_section_leak(analysis_row.get("candidate_text", ""))
        base_intro = base.get("base_intro", "")
        base_analysis = base.get("base_mutation_analysis", "")
        rec_intro = candidate_intro if candidate_intro and not has_pii(candidate_intro) else base_intro
        rec_analysis = candidate_analysis if candidate_analysis and not has_pii(candidate_analysis) else base_analysis
        row = {
            "gene": gene,
            "intro_candidate_id": intro_row.get("candidate_id", ""),
            "analysis_candidate_id": analysis_row.get("candidate_id", ""),
            "analysis_c_hgvs": analysis_row.get("c_hgvs", ""),
            "analysis_p_hgvs": analysis_row.get("p_hgvs", ""),
            "intro_source_count": intro_row.get("source_count", ""),
            "analysis_source_count": analysis_row.get("source_count", ""),
            "base_has_intro": "是" if base_intro else "否",
            "base_has_mutation_analysis": "是" if base_analysis else "否",
            "intro_similarity_to_base": similarity(candidate_intro, base_intro),
            "analysis_similarity_to_base": similarity(candidate_analysis, base_analysis),
            "recommended_intro": rec_intro,
            "recommended_mutation_analysis": rec_analysis,
            "historical_intro": candidate_intro,
            "base_intro": base_intro,
            "historical_mutation_analysis": candidate_analysis,
            "base_mutation_analysis": base_analysis,
            "review_status": "待医学审核",
            "review_notes": "",
        }
        if base_intro or base_analysis:
            row["machine_suggestion"] = "基础知识库已有同基因内容；建议报告组集中审核后进入 reviewed overlay"
            base_supported.append(row)
        else:
            row["machine_suggestion"] = "基础知识库缺失；需从更多历史终版/公共库/人工整理补证据"
            base_missing.append(row)
    return base_supported, base_missing


def build_pending_overlay(base_supported: list[dict[str, Any]]) -> dict[str, Any]:
    gene_sections = []
    for row in base_supported:
        if row.get("recommended_intro"):
            gene_sections.append(
                {
                    "gene": row["gene"],
                    "c_hgvs": "",
                    "p_hgvs": "",
                    "intro": row["recommended_intro"],
                    "mutation_analysis": "",
                }
            )
        if row.get("recommended_mutation_analysis") and row.get("analysis_c_hgvs"):
            gene_sections.append(
                {
                    "gene": row["gene"],
                    "c_hgvs": row.get("analysis_c_hgvs", ""),
                    "p_hgvs": row.get("analysis_p_hgvs", ""),
                    "intro": "",
                    "mutation_analysis": row["recommended_mutation_analysis"],
                }
            )
    return {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "source_type": "base_gene_kb_supported_gap_review",
            "status": "pending_medical_review",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "merge_policy": "review-only artifact; do not merge into production before medical approval",
        },
        "gene_sections": gene_sections,
    }


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_review_workbook(
    path: Path,
    base_supported: list[dict[str, Any]],
    base_missing: list[dict[str, Any]],
    overlay: dict[str, Any],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(
        ws,
        [
            {"项目": "定位", "说明": "CRC358 基因简介/变异解析缺口审核包；不是生产入库文件。"},
            {"项目": "基础库可支撑基因", "说明": len(base_supported)},
            {"项目": "基础库仍缺基因", "说明": len(base_missing)},
            {"项目": "处理范围", "说明": "仅 gene_intro / mutation_analysis；不处理药物解析或动态变异说明。"},
            {"项目": "安全边界", "说明": "生成 pending_medical_review overlay 预览，医学审核通过后才可合入生产。"},
        ],
    )
    ws = wb.create_sheet("基础库可支撑")
    write_sheet(ws, base_supported)
    ws = wb.create_sheet("基础库仍缺")
    write_sheet(ws, base_missing)
    ws = wb.create_sheet("待审overlay预览")
    write_sheet(ws, overlay.get("gene_sections") or [])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidates", type=Path, default=DEFAULT_CANDIDATES)
    parser.add_argument("--base-kb", type=Path, default=DEFAULT_BASE_KB)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    candidate_rows = read_candidate_rows(args.candidates)
    base_kb = load_base_gene_knowledge(args.base_kb)
    base_supported, base_missing = build_gene_gap_review(candidate_rows, base_kb)
    overlay = build_pending_overlay(base_supported)

    review_xlsx = args.out_dir / "CRC358_batch6_基因简介与变异解析补库审核包_20260614.xlsx"
    overlay_yaml = args.out_dir / "reviewed_part3_gene_gap_pending_review_batch6.yaml"
    write_review_workbook(review_xlsx, base_supported, base_missing, overlay)
    dump_yaml(overlay_yaml, overlay)

    print(f"base_supported_genes={len(base_supported)} base_missing_genes={len(base_missing)}")
    print(f"pending_gene_sections={len(overlay.get('gene_sections') or [])}")
    print(f"review_xlsx={review_xlsx}")
    print(f"pending_overlay={overlay_yaml}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
