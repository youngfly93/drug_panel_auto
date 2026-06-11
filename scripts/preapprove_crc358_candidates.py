#!/usr/bin/env python3
"""Create a machine pre-review workbook for CRC358 knowledge candidates.

This is a draft accelerator, not a production approval mechanism. The output
copy marks only conservative historical-final-report rows as approved so the
promotion script can build a test overlay draft.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


DEFAULT_INPUT = Path("tmp/knowledge_buildout/CRC358_医学知识库候选审核表_v0.1.xlsx")
DEFAULT_OUTPUT = Path("tmp/knowledge_buildout/CRC358_医学知识库机器预审表_v0.1.xlsx")
SAFE_CONTENT_TYPES = {"gene_intro", "mutation_analysis", "drug_relation", "drug_clinical"}
SAFE_CONFIDENCE = {"高", "中"}
PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]


def clean(value: object) -> str:
    return str(value or "").strip()


def has_pii(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def is_safe_gene(gene: str) -> bool:
    return bool(re.fullmatch(r"[A-Z][A-Z0-9-]{1,15}", gene or ""))


def should_preapprove(row: dict[str, str]) -> tuple[bool, str]:
    if row.get("source_type") != "historical_final_report":
        return False, "非历史终版来源"
    if row.get("confidence") not in SAFE_CONFIDENCE:
        return False, "置信度不足"
    if row.get("content_type") not in SAFE_CONTENT_TYPES:
        return False, "样本特异或暂不适合自动入草稿"
    if row.get("current_reviewed_status") != "暂无reviewed覆盖":
        return False, "正式库已有覆盖"
    if not is_safe_gene(row.get("gene", "")):
        return False, "复合基因/异常基因名需人工处理"
    if has_pii(row.get("candidate_text", "")):
        return False, "疑似PII风险"
    if row.get("content_type") in {"mutation_analysis", "drug_relation", "drug_clinical"} and not (
        row.get("c_hgvs") or row.get("p_hgvs")
    ):
        return False, "非基因简介但缺位点信息"
    return True, "机器预审通过：历史终版重复出现，未检出PII，适合进入测试overlay草稿"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    wb = load_workbook(args.input)
    if "候选审核表" not in wb.sheetnames:
        raise ValueError("workbook missing sheet: 候选审核表")
    ws = wb["候选审核表"]
    headers = [clean(cell.value) for cell in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    required = {"review_status", "reviewed_text", "review_notes"}
    missing = required - set(idx)
    if missing:
        raise ValueError(f"workbook missing columns: {sorted(missing)}")

    approved = 0
    rejected = 0
    for row_num in range(2, ws.max_row + 1):
        row = {h: clean(ws.cell(row=row_num, column=col).value) for h, col in idx.items()}
        ok, note = should_preapprove(row)
        if ok:
            ws.cell(row=row_num, column=idx["review_status"]).value = "通过"
            ws.cell(row=row_num, column=idx["reviewed_text"]).value = row.get("candidate_text", "")
            ws.cell(row=row_num, column=idx["review_notes"]).value = note
            approved += 1
        else:
            if not row.get("review_notes"):
                ws.cell(row=row_num, column=idx["review_notes"]).value = "机器预审未入草稿：" + note
            rejected += 1

    if "机器预审说明" in wb.sheetnames:
        del wb["机器预审说明"]
    guide = wb.create_sheet("机器预审说明", 0)
    guide.append(["项目", "说明"])
    guide.append(["定位", "用于生成测试overlay草稿，不等同生产医学终审"])
    guide.append(["通过规则", "历史终版来源；高/中置信；非PII；未被正式reviewed覆盖；非复合基因；非样本特异说明"])
    guide.append(["通过条数", approved])
    guide.append(["未入草稿条数", rejected])
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F6B78")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 90

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"approved={approved} rejected={rejected} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
