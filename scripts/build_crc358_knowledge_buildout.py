#!/usr/bin/env python3
"""Build CRC358 candidate medical knowledge workbooks from final DOCX reports.

This script intentionally writes only review artifacts under ``tmp/``. It does
not promote unreviewed medical text into the production panel overlay.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CORPUS = Path("各癌种基因报告近年汇总/肠癌")
DEFAULT_OUT = Path("tmp/knowledge_buildout")
DEFAULT_CIVIC = Path(
    "tmp/civic_crc358_public_candidate_pilot_20260611/"
    "CRC358_CIViC公共候选知识库pilot_20260611.xlsx"
)
DEFAULT_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_GAP_STATS = Path("知识库缺口统计_20260605.xlsx")
TODAY = "2026-06-11"

CONTENT_TYPES = {
    "gene_intro": "基因简介",
    "variant_description": "基因变异说明",
    "mutation_analysis": "基因变异解析",
    "drug_relation": "基因变异与药物关联分析",
    "drug_clinical": "药物疗效临床解析",
    "public_gene_intro": "公共库候选基因简介",
    "public_mutation_analysis": "公共库候选变异解析",
}

PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"送检者"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]

HGVS_C_RE = re.compile(r"c\.[A-Za-z0-9_+\-*?>=.]+(?:delins|del|dup|ins)?[A-Za-z0-9_+\-*?>=.]*")
HGVS_P_RE = re.compile(r"p\.[A-Za-z0-9_+\-*?>=.]+")
PMID_RE = re.compile(r"PMID[:：]?\s*(\d{5,9})|\[(\d{5,9})\]")
NCT_RE = re.compile(r"NCT\d{8}")
GENE_RE = re.compile(r"^[A-Z][A-Z0-9-]{1,15}(?:/[A-Z][A-Z0-9-]{1,15})*")
GENE_ALIASES = {
    "DMNT3A": "DNMT3A",
    "DNM3TA": "DNMT3A",
}


@dataclass
class VariantContext:
    gene: str = ""
    c_hgvs: str = ""
    p_hgvs: str = ""
    frequency: str = ""
    header: str = ""


@dataclass
class Candidate:
    candidate_id: str
    source_hash: str
    source_type: str
    source_family: str
    product_family: str
    gene: str
    c_hgvs: str = ""
    p_hgvs: str = ""
    content_type: str = ""
    candidate_text: str = ""
    header: str = ""
    drug_name: str = ""
    drug_type: str = ""
    pmids: str = ""
    nct_ids: str = ""
    source_count: int = 1
    source_hashes: set[str] = field(default_factory=set)
    current_reviewed_status: str = "未覆盖"
    confidence: str = "低"
    system_suggestion: str = ""
    review_status: str = "待审核"
    reviewed_text: str = ""
    review_notes: str = ""

    def gap_classification(self) -> tuple[str, str]:
        if self.current_reviewed_status != "暂无reviewed覆盖":
            return ("已覆盖", "无需处理")
        if self.content_type == "variant_description":
            return (
                "动态生成项_不入reviewed知识库",
                "变异说明含样本特异描述，应由程序动态生成，不作为 reviewed 文案入库",
            )
        if self.content_type in {"drug_relation", "drug_clinical"}:
            if not self.drug_name or self.drug_name == "--":
                return (
                    "药物解析缺药名_需人工整理",
                    "药物解析候选缺少明确药名，不能直接进入 reviewed 药物知识库",
                )
            if not (self.c_hgvs or self.p_hgvs):
                return (
                    "药物解析缺位点_需人工整理",
                    "药物解析候选缺少明确位点或适用条件，需人工整理为 gene/applicability 规则后再入库",
                )
        if self.confidence == "禁入":
            return ("禁入", "疑似包含PII或格式风险，禁止直接入库")
        if self.source_type == "public_civic_candidate":
            return ("公共库候选_待审核", "需结合历史终版或内部口径审核后再定稿")
        if self.confidence in {"高", "中"}:
            return ("可审核入库候选", "建议优先进入报告组/医学审核")
        return ("低置信待补证据", "单来源或证据不足，需补充来源或人工整理")

    def as_row(self) -> dict[str, object]:
        gap_class, gap_action = self.gap_classification()
        return {
            "candidate_id": self.candidate_id,
            "source_family": self.source_family,
            "product_family": self.product_family,
            "source_type": self.source_type,
            "source_count": self.source_count,
            "source_hashes": ";".join(sorted(self.source_hashes)) or self.source_hash,
            "gene": self.gene,
            "c_hgvs": self.c_hgvs,
            "p_hgvs": self.p_hgvs,
            "content_type": self.content_type,
            "content_type_cn": CONTENT_TYPES.get(self.content_type, self.content_type),
            "header": self.header,
            "drug_name": self.drug_name,
            "drug_type": self.drug_type,
            "candidate_text": self.candidate_text,
            "pmids": self.pmids,
            "nct_ids": self.nct_ids,
            "current_reviewed_status": self.current_reviewed_status,
            "confidence": self.confidence,
            "system_suggestion": self.system_suggestion,
            "kb_gap_class": gap_class,
            "kb_gap_action": gap_action,
            "review_status": self.review_status,
            "reviewed_text": self.reviewed_text,
            "review_notes": self.review_notes,
        }


def norm_space(text: str) -> str:
    text = html.unescape(text or "")
    text = re.sub(r'HYPERLINK\s+\\l\s+"[^"]+"\s*', "", text)
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def docx_paragraphs(path: Path) -> list[str]:
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8", "ignore")
    paragraphs: list[str] = []
    for match in re.finditer(r"<w:p[ >].*?</w:p>", xml, re.S):
        raw = match.group(0)
        texts = re.findall(r"<w:t[^>]*>(.*?)</w:t>", raw, re.S)
        text = norm_space("".join(re.sub(r"<[^>]+>", "", item) for item in texts))
        if text:
            paragraphs.append(text)
    return paragraphs


def stable_file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def short_hash(text: str, length: int = 16) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:length]


def classify_product_family(name: str) -> tuple[str, str, bool, bool]:
    lower = name.lower()
    has_msi = "msi" in lower
    is_final = "终版" in name
    match = re.search(r"结直肠癌\s*(\d+)\s*基因", name)
    if not match:
        return "crc_other", "", has_msi, is_final
    panel_size = match.group(1)
    product_name = f"结直肠癌{panel_size}基因" + ("+MSI" if has_msi else "")
    if panel_size == "358":
        return "crc_358_msi" if has_msi else "crc_358", product_name, has_msi, is_final
    if panel_size == "301":
        return "crc_301_msi" if has_msi else "crc_301", product_name, has_msi, is_final
    return f"crc_{panel_size}_msi" if has_msi else f"crc_{panel_size}", product_name, has_msi, is_final


def is_docx_candidate(path: Path) -> bool:
    return path.suffix.lower() == ".docx" and not path.name.startswith("._")


def find_part3_start(paragraphs: list[str]) -> int:
    for i, text in enumerate(paragraphs):
        if text.startswith("HYPERLINK"):
            continue
        if i < 80:
            continue
        if text.startswith("第三部分") and ("基因变异" in text or "药物" in text):
            return i
    for i, text in enumerate(paragraphs):
        if text == "基因变异解析" and i > 80:
            return max(0, i - 2)
    return 0


def find_part3_end(paragraphs: list[str], start: int) -> int:
    for i in range(start + 1, len(paragraphs)):
        text = paragraphs[i]
        if "第四部分" in text or text.startswith("附录"):
            return i
        if text.startswith("PMID:"):
            return i
    return len(paragraphs)


def strip_bullet(text: str) -> str:
    text = norm_space(text)
    text = re.sub(r"^[u◆●•·\-\s]+", "", text)
    return text.strip()


def normalize_gene_symbol(gene: str) -> str:
    gene = norm_space(gene).upper()
    return GENE_ALIASES.get(gene, gene)


def parse_variant_heading(text: str) -> VariantContext | None:
    text = strip_bullet(text)
    if len(text) > 180:
        return None
    if any(label in text for label in ("基因简介", "基因变异说明", "基因变异解析", "药物疗效")):
        return None
    if text.startswith(("PMID", "HYPERLINK", "检测", "报告", "备注", "潜在", "靶向")):
        return None
    if "：" not in text and ":" not in text:
        return None
    left, right = re.split(r"[：:]", text, maxsplit=1)
    gene_match = GENE_RE.match(left.strip())
    if not gene_match:
        return None
    gene = normalize_gene_symbol(gene_match.group(0).strip())
    c_match = HGVS_C_RE.search(right)
    p_match = HGVS_P_RE.search(right)
    freq = ""
    f_match = re.search(r"(\d+(?:\.\d+)?)\s*%", right)
    if f_match:
        freq = f_match.group(1) + "%"
    return VariantContext(
        gene=gene,
        c_hgvs=c_match.group(0).rstrip("，,；;") if c_match else "",
        p_hgvs=p_match.group(0).rstrip("，,；;") if p_match else "",
        frequency=freq,
        header=text,
    )


def is_composite_context(context: VariantContext) -> bool:
    """Return True for composite biomarker headings, not single-gene variants."""
    return "/" in (context.gene or "") and not (context.c_hgvs or context.p_hgvs)


def pmids_of(text: str) -> str:
    ids: set[str] = set()
    for a, b in PMID_RE.findall(text or ""):
        ids.add(a or b)
    return ";".join(sorted(ids))


def nct_ids_of(text: str) -> str:
    return ";".join(sorted(set(NCT_RE.findall(text or ""))))


def pii_risk(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def candidate_text_matches_context(gene: str, content_type: str, text: str) -> bool:
    """Reject text blocks that were captured under an unrelated gene heading."""
    expected = normalize_gene_symbol(gene)
    if not expected:
        return False
    text = norm_space(text)
    first_gene = ""
    if content_type == "gene_intro":
        match = re.search(r"^([A-Z][A-Z0-9-]{1,15})\s*(?:（[^）]*）)?基因", text)
        if not match:
            match = re.search(r"^([A-Z][A-Z0-9-]{1,15})\s*(?:编码|是|属于|参与)", text)
        if match:
            first_gene = match.group(1)
    elif content_type in {"variant_description", "mutation_analysis"}:
        match = re.search(r"该样本检出([A-Z][A-Z0-9-]{1,15})基因", text[:220])
        if match:
            first_gene = match.group(1)
        else:
            match = re.search(r"([A-Z][A-Z0-9-]{1,15})基因编码", text[:220])
            if match:
                first_gene = match.group(1)
    if not first_gene:
        return True
    return normalize_gene_symbol(first_gene) == expected


def make_candidate(
    source_hash: str,
    source_type: str,
    product_family: str,
    context: VariantContext,
    content_type: str,
    text: str,
    header: str = "",
    drug_name: str = "",
    drug_type: str = "",
) -> Candidate | None:
    text = norm_space(text)
    if len(text) < 12:
        return None
    if not context.gene:
        return None
    gene = normalize_gene_symbol(context.gene)
    if not candidate_text_matches_context(gene, content_type, text):
        return None
    seed = "|".join(
        [
            source_type,
            product_family,
            gene,
            context.c_hgvs,
            context.p_hgvs,
            content_type,
            text,
            header,
            drug_name,
            drug_type,
        ]
    )
    return Candidate(
        candidate_id="cand_" + short_hash(seed, 14),
        source_hash=source_hash,
        source_hashes={source_hash},
        source_type=source_type,
        source_family="肠癌",
        product_family=product_family,
        gene=gene,
        c_hgvs=context.c_hgvs,
        p_hgvs=context.p_hgvs,
        content_type=content_type,
        candidate_text=text,
        header=header or context.header,
        drug_name=drug_name,
        drug_type=drug_type,
        pmids=pmids_of(text),
        nct_ids=nct_ids_of(text),
    )


def extract_candidates_from_paragraphs(
    paragraphs: list[str],
    source_hash: str,
    product_family: str,
) -> list[Candidate]:
    start = find_part3_start(paragraphs)
    end = find_part3_end(paragraphs, start)
    part3 = paragraphs[start:end]
    candidates: list[Candidate] = []
    context = VariantContext()
    mode = ""
    buffer: list[str] = []
    drug_type = ""
    drug_header = ""
    drug_name = ""
    drug_context = VariantContext()

    def flush() -> None:
        nonlocal buffer, mode
        if not mode or not buffer:
            buffer = []
            return
        text = "\n".join(buffer).strip()
        if mode in {"gene_intro", "variant_description", "mutation_analysis"}:
            cand = make_candidate(
                source_hash,
                "historical_final_report",
                product_family,
                context,
                mode,
                text,
            )
        else:
            target_context = drug_context if drug_context.gene else context
            cand = make_candidate(
                source_hash,
                "historical_final_report",
                product_family,
                target_context,
                mode,
                text,
                header=drug_header,
                drug_name=drug_name,
                drug_type=drug_type,
            )
        if cand:
            candidates.append(cand)
        buffer = []

    for raw in part3:
        text = norm_space(raw)
        if not text:
            continue

        if "潜在获益" in text and "药物" in text:
            flush()
            drug_type = "benefit"
            mode = ""
            continue
        if any(marker in text for marker in ("潜在负相关", "可能耐药", "慎用药物")) and "药物" in text:
            flush()
            drug_type = "caution"
            mode = ""
            continue

        if text in {"基因简介：", "基因简介:"}:
            flush()
            mode = "gene_intro"
            continue
        if text in {"基因变异说明：", "基因变异说明:"}:
            flush()
            mode = "variant_description"
            continue
        if text in {
            "基因变异解析：",
            "基因变异解析:",
            "基因结构域：",
            "基因结构域:",
        }:
            flush()
            mode = "mutation_analysis"
            continue
        if text in {"基因变异与药物关联分析：", "基因变异与药物关联分析:"}:
            flush()
            mode = "drug_relation"
            continue
        if text in {"药物疗效临床解析：", "药物疗效临床解析:"}:
            flush()
            mode = "drug_clinical"
            continue

        parsed = parse_variant_heading(text)
        if parsed:
            flush()
            if is_composite_context(parsed):
                context = VariantContext()
                drug_context = VariantContext()
                drug_header = ""
                drug_name = ""
                mode = ""
                continue
            if "突变相应" in text or "未突变相应" in text:
                drug_context = parsed
                drug_header = text
                drug_name = ""
                mode = ""
            else:
                context = parsed
                mode = ""
            continue

        if drug_type and drug_header and not mode and "药物" not in text and len(text) < 120:
            drug_name = text
            continue

        if mode:
            buffer.append(text)

    flush()
    return candidates


def load_reviewed_status(path: Path) -> tuple[set[tuple[str, str, str]], set[str]]:
    variant_keys: set[tuple[str, str, str]] = set()
    gene_keys: set[str] = set()
    if not path.exists():
        return variant_keys, gene_keys
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    for section in (data.get("gene_sections") or []) + (data.get("drug_sections") or []):
        gene = norm_space(section.get("gene")).upper()
        if not gene:
            continue
        c_hgvs = norm_space(section.get("c_hgvs"))
        p_hgvs = norm_space(section.get("p_hgvs"))
        if c_hgvs or p_hgvs:
            variant_keys.add((gene, c_hgvs, p_hgvs))
        else:
            gene_keys.add(gene)
    return variant_keys, gene_keys


def reviewed_status(candidate: Candidate, variant_keys: set[tuple[str, str, str]], gene_keys: set[str]) -> str:
    key = (candidate.gene.upper(), candidate.c_hgvs, candidate.p_hgvs)
    key_c_only = (candidate.gene.upper(), candidate.c_hgvs, "")
    if key in variant_keys or key_c_only in variant_keys:
        return "位点级reviewed已覆盖"
    if candidate.gene.upper() in gene_keys:
        return "基因级reviewed已覆盖"
    return "暂无reviewed覆盖"


def dedupe_candidates(candidates: Iterable[Candidate]) -> list[Candidate]:
    groups: dict[tuple[str, str, str, str, str, str, str], Candidate] = {}

    def product_key(cand: Candidate) -> str:
        # Gene introductions are cancer-family knowledge. Treat identical CRC
        # final-report wording across panel sizes as supporting evidence rather
        # than unrelated single-source rows. Variant/drug rows stay panel-scoped.
        if cand.source_type == "historical_final_report" and cand.content_type == "gene_intro":
            return "crc_all_panels"
        return cand.product_family

    def merge_product_family(existing: Candidate, incoming: Candidate) -> None:
        values = {
            item
            for text in (existing.product_family, incoming.product_family)
            for item in str(text or "").split(";")
            if item
        }
        if values:
            existing.product_family = ";".join(sorted(values))

    for cand in candidates:
        key_gene = cand.gene.upper()
        c_hgvs = cand.c_hgvs if cand.content_type not in {"gene_intro", "public_gene_intro"} else ""
        p_hgvs = cand.p_hgvs if cand.content_type not in {"gene_intro", "public_gene_intro"} else ""
        text_key = re.sub(r"\s+", "", cand.candidate_text)
        key = (product_key(cand), key_gene, c_hgvs, p_hgvs, cand.content_type, cand.drug_name, text_key)
        if key not in groups:
            groups[key] = cand
            continue
        existing = groups[key]
        merge_product_family(existing, cand)
        existing.source_hashes.update(cand.source_hashes or {cand.source_hash})
        existing.source_count = len(existing.source_hashes)
        if cand.pmids:
            existing.pmids = ";".join(sorted(set(filter(None, (existing.pmids + ";" + cand.pmids).split(";")))))
        if cand.nct_ids:
            existing.nct_ids = ";".join(sorted(set(filter(None, (existing.nct_ids + ";" + cand.nct_ids).split(";")))))

    rows = list(groups.values())
    for cand in rows:
        cand.source_count = len(cand.source_hashes) or cand.source_count
        if cand.source_type == "public_civic_candidate":
            cand.confidence = "低"
            cand.system_suggestion = "公共数据库候选；需结合历史终版报告或内部口径后再定稿"
        elif pii_risk(cand.candidate_text):
            cand.confidence = "禁入"
            cand.system_suggestion = "疑似包含患者标识或日期；禁止直接入库"
        elif cand.content_type == "variant_description":
            cand.confidence = "低"
            cand.system_suggestion = "该类说明含样本特异描述，优先由程序动态生成，不建议直接入库"
        elif cand.source_count >= 3:
            cand.confidence = "高"
            cand.system_suggestion = "多份终版报告重复出现；建议优先集中审核"
        elif cand.source_count == 2:
            cand.confidence = "中"
            cand.system_suggestion = "有两份终版报告支持；建议审核后入候选库"
        else:
            cand.confidence = "低"
            cand.system_suggestion = "单份终版报告来源；建议结合更多报告或公共库后审核"
    return sorted(
        rows,
        key=lambda c: (
            {"高": 0, "中": 1, "低": 2, "禁入": 3}.get(c.confidence, 9),
            c.gene,
            c.c_hgvs,
            c.content_type,
        ),
    )


def load_civic_candidates(path: Path) -> list[Candidate]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "候选中文解读" not in wb.sheetnames:
        return []
    ws = wb["候选中文解读"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip() for x in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    out: list[Candidate] = []
    for row in rows[1:]:
        gene = norm_space(row[idx.get("基因", -1)] if idx.get("基因", -1) >= 0 else "")
        if not gene:
            continue
        trigger = norm_space(row[idx.get("触发位点", -1)] if idx.get("触发位点", -1) >= 0 else "")
        c_match = HGVS_C_RE.search(trigger)
        p_match = HGVS_P_RE.search(trigger)
        context = VariantContext(
            gene=gene.upper(),
            c_hgvs=c_match.group(0) if c_match else "",
            p_hgvs=p_match.group(0) if p_match else "",
            header=trigger,
        )
        for col, content_type in (("候选中文基因简介", "public_gene_intro"), ("候选中文变异解析", "public_mutation_analysis")):
            pos = idx.get(col, -1)
            text = norm_space(row[pos] if pos >= 0 else "")
            if not text:
                continue
            cand = make_candidate(
                "civic_" + short_hash(gene + trigger + content_type, 8),
                "public_civic_candidate",
                "crc_358_msi",
                context,
                content_type,
                text,
                header=trigger,
            )
            if cand:
                cand.source_family = "公共数据库"
                cand.review_status = norm_space(row[idx.get("报告组审核结论", -1)] if idx.get("报告组审核结论", -1) >= 0 else "") or "待审核"
                cand.review_notes = norm_space(row[idx.get("报告组修改意见", -1)] if idx.get("报告组修改意见", -1) >= 0 else "")
                out.append(cand)
    return out


def split_trigger_site(trigger: str) -> tuple[str, str]:
    c_match = HGVS_C_RE.search(trigger or "")
    p_match = HGVS_P_RE.search(trigger or "")
    return (c_match.group(0) if c_match else "", p_match.group(0) if p_match else "")


def load_gap_rows(civic_path: Path, gap_stats_path: Path) -> list[dict[str, object]]:
    gaps: dict[tuple[str, str, str, str], dict[str, object]] = {}

    if civic_path.exists():
        wb = load_workbook(civic_path, read_only=True, data_only=True)
        if "候选中文解读" in wb.sheetnames:
            ws = wb["候选中文解读"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(x or "").strip() for x in rows[0]]
                idx = {h: i for i, h in enumerate(headers)}
                for row in rows[1:]:
                    gene = norm_space(row[idx.get("基因", -1)] if idx.get("基因", -1) >= 0 else "")
                    trigger = norm_space(row[idx.get("触发位点", -1)] if idx.get("触发位点", -1) >= 0 else "")
                    intro = norm_space(row[idx.get("候选中文基因简介", -1)] if idx.get("候选中文基因简介", -1) >= 0 else "")
                    analysis = norm_space(row[idx.get("候选中文变异解析", -1)] if idx.get("候选中文变异解析", -1) >= 0 else "")
                    if not gene or intro or analysis:
                        continue
                    c_hgvs, p_hgvs = split_trigger_site(trigger)
                    key = (gene, c_hgvs, p_hgvs, "public_civic_pilot")
                    gaps[key] = {
                        "gene": gene,
                        "c_hgvs": c_hgvs,
                        "p_hgvs": p_hgvs,
                        "source": "public_civic_pilot",
                        "occurrences": 1,
                        "current_status": norm_space(row[idx.get("当前reviewed状态", -1)] if idx.get("当前reviewed状态", -1) >= 0 else ""),
                        "priority": norm_space(row[idx.get("优先级", -1)] if idx.get("优先级", -1) >= 0 else ""),
                        "suggestion": norm_space(row[idx.get("建议", -1)] if idx.get("建议", -1) >= 0 else ""),
                    }

    if gap_stats_path.exists():
        wb = load_workbook(gap_stats_path, read_only=True, data_only=True)
        if "缺口明细" in wb.sheetnames:
            ws = wb["缺口明细"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(x or "").strip() for x in rows[0]]
                idx = {h: i for i, h in enumerate(headers)}
                for row in rows[1:]:
                    gene = norm_space(row[idx.get("基因", -1)] if idx.get("基因", -1) >= 0 else "")
                    c_hgvs = norm_space(row[idx.get("cHGVS", -1)] if idx.get("cHGVS", -1) >= 0 else "")
                    p_hgvs = norm_space(row[idx.get("pHGVS", -1)] if idx.get("pHGVS", -1) >= 0 else "")
                    if not gene:
                        continue
                    key = (gene, c_hgvs, p_hgvs, "pressure_gap_aggregate")
                    if key not in gaps:
                        gaps[key] = {
                            "gene": gene,
                            "c_hgvs": c_hgvs,
                            "p_hgvs": p_hgvs,
                            "source": "pressure_gap_aggregate",
                            "occurrences": 0,
                            "current_status": norm_space(row[idx.get("覆盖状态", -1)] if idx.get("覆盖状态", -1) >= 0 else ""),
                            "priority": norm_space(row[idx.get("优先级", -1)] if idx.get("优先级", -1) >= 0 else ""),
                            "suggestion": norm_space(row[idx.get("建议动作", -1)] if idx.get("建议动作", -1) >= 0 else ""),
                        }
                    gaps[key]["occurrences"] = int(gaps[key].get("occurrences") or 0) + 1

    return sorted(
        gaps.values(),
        key=lambda r: (
            str(r.get("priority") or ""),
            str(r.get("gene") or ""),
            str(r.get("c_hgvs") or ""),
            str(r.get("p_hgvs") or ""),
            str(r.get("source") or ""),
        ),
    )


def should_extract_candidates(product_family: str, is_final: bool, include_all_crc_panels: bool) -> bool:
    if not is_final:
        return False
    if product_family == "crc_358_msi":
        return True
    if not include_all_crc_panels:
        return False
    return product_family.startswith("crc_") and product_family != "crc_other"


def scan_reports(
    corpus: Path,
    overlay: Path,
    civic: Path,
    *,
    include_all_crc_panels: bool = False,
) -> tuple[list[dict[str, object]], list[Candidate], list[Candidate]]:
    variant_keys, gene_keys = load_reviewed_status(overlay)
    inventory: list[dict[str, object]] = []
    raw: list[Candidate] = []
    for path in sorted(corpus.iterdir(), key=lambda p: p.name):
        if path.name.startswith("._"):
            inventory.append(
                {
                    "source_hash": "",
                    "source_family": "肠癌",
                    "product_family": "",
                    "product_name": "",
                    "has_msi": "",
                    "is_final": "",
                    "file_size_kb": round(path.stat().st_size / 1024, 1),
                    "docx_valid": False,
                    "paragraph_count": 0,
                    "part3_found": False,
                    "candidate_count": 0,
                    "skip_reason": "macOS资源叉文件",
                    "parse_error": "",
                }
            )
            continue
        if path.suffix.lower() != ".docx":
            continue
        product_family, product_name, has_msi, is_final = classify_product_family(path.name)
        source_hash = stable_file_hash(path)
        row: dict[str, object] = {
            "source_hash": source_hash,
            "source_family": "肠癌",
            "product_family": product_family,
            "product_name": product_name,
            "has_msi": has_msi,
            "is_final": is_final,
            "file_size_kb": round(path.stat().st_size / 1024, 1),
            "docx_valid": False,
            "paragraph_count": 0,
            "part3_found": False,
            "candidate_count": 0,
            "skip_reason": "" if is_final else "文件名未标记终版",
            "parse_error": "",
        }
        try:
            paragraphs = docx_paragraphs(path)
            row["docx_valid"] = True
            row["paragraph_count"] = len(paragraphs)
            start = find_part3_start(paragraphs)
            row["part3_found"] = start > 0
            if should_extract_candidates(product_family, is_final, include_all_crc_panels):
                candidates = extract_candidates_from_paragraphs(paragraphs, source_hash, product_family)
                for cand in candidates:
                    cand.current_reviewed_status = reviewed_status(cand, variant_keys, gene_keys)
                raw.extend(candidates)
                row["candidate_count"] = len(candidates)
        except Exception as exc:  # pragma: no cover - defensive per-document logging
            row["parse_error"] = f"{type(exc).__name__}: {exc}"
        inventory.append(row)

    public_candidates = load_civic_candidates(civic)
    for cand in public_candidates:
        cand.current_reviewed_status = reviewed_status(cand, variant_keys, gene_keys)
    deduped = dedupe_candidates(raw + public_candidates)
    return inventory, raw, deduped


def write_sheet(ws, rows: list[dict[str, object]], freeze: str = "A2") -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                max_len = max(max_len, min(len(str(item.value or "")), 60))
                item.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 55)


def write_workbooks(
    out_dir: Path,
    inventory: list[dict[str, object]],
    raw: list[Candidate],
    deduped: list[Candidate],
    gaps: list[dict[str, object]],
) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(timezone.utc).isoformat()
    paths = {
        "inventory": out_dir / "crc_report_inventory.xlsx",
        "raw": out_dir / "crc358_part3_candidates_raw.xlsx",
        "review": out_dir / "CRC358_医学知识库候选审核表_v0.1.xlsx",
        "coverage": out_dir / "CRC358_知识库覆盖率基线_v0.1.xlsx",
    }

    inv_wb = Workbook()
    inv_ws = inv_wb.active
    inv_ws.title = "语料inventory"
    write_sheet(inv_ws, inventory)
    summary = inv_wb.create_sheet("汇总", 0)
    product_counts = Counter(str(r.get("product_family") or "") for r in inventory if r.get("product_family"))
    summary_rows = [
        {"项目": "生成时间", "结果": generated_at},
        {"项目": "语料目录", "结果": str(DEFAULT_CORPUS)},
        {"项目": "总DOCX/资源叉记录", "结果": len(inventory)},
        {"项目": "crc_358_msi终版数", "结果": sum(1 for r in inventory if r.get("product_family") == "crc_358_msi" and r.get("is_final"))},
        {"项目": "可解析DOCX数", "结果": sum(1 for r in inventory if r.get("docx_valid"))},
        {"项目": "抽取候选raw条数", "结果": len(raw)},
    ] + [{"项目": f"产品族:{k}", "结果": v} for k, v in product_counts.most_common()]
    write_sheet(summary, summary_rows)
    inv_wb.save(paths["inventory"])

    raw_wb = Workbook()
    raw_ws = raw_wb.active
    raw_ws.title = "raw候选"
    write_sheet(raw_ws, [c.as_row() for c in raw])
    raw_wb.save(paths["raw"])

    review_wb = Workbook()
    review_ws = review_wb.active
    review_ws.title = "候选审核表"
    write_sheet(review_ws, [c.as_row() for c in deduped])
    gap_ws = review_wb.create_sheet("缺口清单")
    write_sheet(gap_ws, gaps)
    overview = review_wb.create_sheet("汇总", 0)
    confidence_counts = Counter(c.confidence for c in deduped)
    status_counts = Counter(c.current_reviewed_status for c in deduped)
    type_counts = Counter(c.content_type for c in deduped)
    gap_class_counts = Counter(c.gap_classification()[0] for c in deduped)
    overview_rows = [
        {"项目": "生成时间", "结果": generated_at},
        {"项目": "定位", "结果": "CRC358医学知识库v0.1候选版；不直接进入生产报告"},
        {"项目": "去重后候选条数", "结果": len(deduped)},
        {"项目": "高置信候选", "结果": confidence_counts.get("高", 0)},
        {"项目": "中置信候选", "结果": confidence_counts.get("中", 0)},
        {"项目": "低置信候选", "结果": confidence_counts.get("低", 0)},
        {"项目": "禁入候选", "结果": confidence_counts.get("禁入", 0)},
        {"项目": "显式缺口条目", "结果": len(gaps)},
        {"项目": "可审核入库候选", "结果": gap_class_counts.get("可审核入库候选", 0)},
        {"项目": "低置信待补证据", "结果": gap_class_counts.get("低置信待补证据", 0)},
        {"项目": "动态生成项_不入reviewed知识库", "结果": gap_class_counts.get("动态生成项_不入reviewed知识库", 0)},
    ] + [{"项目": f"覆盖:{k}", "结果": v} for k, v in status_counts.items()] + [
        {"项目": f"缺口分类:{k}", "结果": v} for k, v in gap_class_counts.items()
    ] + [
        {"项目": f"类型:{CONTENT_TYPES.get(k, k)}", "结果": v} for k, v in type_counts.items()
    ]
    write_sheet(overview, overview_rows)
    guide = review_wb.create_sheet("审核说明")
    guide_rows = [
        {"字段": "review_status", "说明": "报告组集中定稿字段；允许值建议为 待审核/通过/修改后通过/不通过/暂缓"},
        {"字段": "reviewed_text", "说明": "修改后通过时填写最终正文；为空时默认使用 candidate_text"},
        {"字段": "confidence", "说明": "系统按历史终版重复次数和风险自动分层，只做审核优先级参考"},
        {"字段": "current_reviewed_status", "说明": "当前正式 overlay 是否已覆盖，不代表候选内容已可用"},
        {"字段": "source_hashes", "说明": "去标识化来源 hash；不输出真实报告文件名、姓名、样本号"},
    ]
    write_sheet(guide, guide_rows)
    review_wb.save(paths["review"])

    cov_wb = Workbook()
    cov_ws = cov_wb.active
    cov_ws.title = "覆盖基线"
    coverage_rows = [
        {"指标": "生成时间", "结果": generated_at},
        {"指标": "crc_358_msi终版报告数", "结果": sum(1 for r in inventory if r.get("product_family") == "crc_358_msi" and r.get("is_final"))},
        {"指标": "历史raw候选条数", "结果": len(raw)},
        {"指标": "去重候选条数", "结果": len(deduped)},
        {"指标": "高置信候选", "结果": confidence_counts.get("高", 0)},
        {"指标": "中置信候选", "结果": confidence_counts.get("中", 0)},
        {"指标": "低置信候选", "结果": confidence_counts.get("低", 0)},
        {"指标": "显式缺口条目", "结果": len(gaps)},
        {"指标": "可审核入库候选", "结果": gap_class_counts.get("可审核入库候选", 0)},
        {"指标": "低置信待补证据", "结果": gap_class_counts.get("低置信待补证据", 0)},
        {"指标": "动态生成项_不入reviewed知识库", "结果": gap_class_counts.get("动态生成项_不入reviewed知识库", 0)},
    ]
    write_sheet(cov_ws, coverage_rows)
    cov_gap_ws = cov_wb.create_sheet("缺口聚合")
    write_sheet(cov_gap_ws, gaps)
    cov_type_ws = cov_wb.create_sheet("候选类型统计")
    type_rows = [{"content_type": k, "content_type_cn": CONTENT_TYPES.get(k, k), "count": v} for k, v in type_counts.items()]
    write_sheet(cov_type_ws, type_rows)
    cov_wb.save(paths["coverage"])
    return paths


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--corpus", type=Path, default=DEFAULT_CORPUS)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--civic-pilot", type=Path, default=DEFAULT_CIVIC)
    parser.add_argument("--overlay", type=Path, default=DEFAULT_OVERLAY)
    parser.add_argument("--gap-stats", type=Path, default=DEFAULT_GAP_STATS)
    parser.add_argument(
        "--include-all-crc-panels",
        action="store_true",
        help=(
            "Extract candidates from all CRC final-report product families, not only "
            "crc_358_msi. Use with conservative promotion rules."
        ),
    )
    args = parser.parse_args()

    inventory, raw, deduped = scan_reports(
        args.corpus,
        args.overlay,
        args.civic_pilot,
        include_all_crc_panels=args.include_all_crc_panels,
    )
    gaps = load_gap_rows(args.civic_pilot, args.gap_stats)
    paths = write_workbooks(args.out_dir, inventory, raw, deduped, gaps)
    print(f"inventory: {paths['inventory']}")
    print(f"raw: {paths['raw']}")
    print(f"review: {paths['review']}")
    print(f"coverage: {paths['coverage']}")
    print(f"raw_candidates={len(raw)} deduped_candidates={len(deduped)} explicit_gaps={len(gaps)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
