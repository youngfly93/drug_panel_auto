#!/usr/bin/env python3
"""Triage CRC358 pending medical-review rows into actionable review buckets.

This script does not approve or merge medical content. It reads the batch9
review workbook, runs structural and text-risk checks, and writes a reviewer
workbook with machine suggestions to make one-pass medical review practical.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_REVIEW_WORKBOOK = Path(
    "tmp/knowledge_buildout_after_batch9_pending_merge_20260614/CRC358_batch9_待医学审核合入包_20260614.xlsx"
)
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch12_medical_triage_20260614")
FULL_REVIEW_SHEETS = ("新增gene完整审核", "新增drug完整审核")

PII_PATTERNS = {
    "sample_id": re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    "report_no": re.compile(r"报告编号"),
    "name_label": re.compile(r"姓名[:：]"),
    "sender": re.compile(r"送检者"),
    "date": re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
}
SECTION_LEAK_PATTERNS = {
    "reading_notes": re.compile(r"3\.\s*阅读说明"),
    "reference_instruction": re.compile(r"文中参考文献及临床试验编号说明"),
    "glossary_instruction": re.compile(r"医学及生物学常见名词说明"),
}
SAMPLE_CONTEXT_PATTERNS = {
    "same_sample_other_variant": re.compile(r"(同时检出|伴随检出|还检出|另检出).{0,80}(突变|变异|扩增|融合)"),
}
GENE_SYMBOL_RE = re.compile(r"(?<![A-Za-z0-9-])([A-Z][A-Z0-9-]{1,12})(?![A-Za-z0-9-])")


def clean(value: Any) -> str:
    return str(value or "").strip()


def scan_hits(text: str, patterns: dict[str, re.Pattern[str]]) -> list[str]:
    return [name for name, pattern in patterns.items() if pattern.search(text)]


def row_text(row: dict[str, str]) -> str:
    fields = (
        "gene",
        "c_hgvs",
        "p_hgvs",
        "header",
        "drug_name",
        "intro",
        "mutation_analysis",
        "relation",
        "clinical",
    )
    return "\n".join(clean(row.get(field)) for field in fields)


def read_review_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for sheet_name in FULL_REVIEW_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows or len(raw_rows[0]) == 1 and clean(raw_rows[0][0]) == "无数据":
            continue
        headers = [clean(value) for value in raw_rows[0]]
        for idx, raw in enumerate(raw_rows[1:], 2):
            item = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
            item["_sheet"] = sheet_name
            item["_row_number"] = str(idx)
            rows.append(item)
    return rows


def text_first_gene_matches_context(row: dict[str, str]) -> bool:
    gene = clean(row.get("gene")).upper()
    candidates = []
    for field in ("intro", "mutation_analysis", "relation", "clinical"):
        text = clean(row.get(field))
        if text:
            candidates.append(text)
    if not gene or not candidates:
        return False
    text = "\n".join(candidates)
    for match in GENE_SYMBOL_RE.finditer(text):
        symbol = match.group(1).upper()
        if symbol in {"DNA", "RNA", "FDA", "NMPA", "COSMIC", "HR", "HER2", "PI3K", "AKT", "TGF", "VEGF"}:
            continue
        return symbol == gene
    return True


def risk_flags(row: dict[str, str]) -> list[str]:
    text = row_text(row)
    flags: list[str] = []
    flags.extend(f"PII:{name}" for name in scan_hits(text, PII_PATTERNS))
    flags.extend(f"章节串漏:{name}" for name in scan_hits(text, SECTION_LEAK_PATTERNS))
    flags.extend(f"样本上下文:{name}" for name in scan_hits(text, SAMPLE_CONTEXT_PATTERNS))
    if not text_first_gene_matches_context(row):
        flags.append("正文首个基因与上下文不一致或正文为空")

    section = clean(row.get("section"))
    if section == "gene_sections":
        has_intro = bool(clean(row.get("intro")))
        has_analysis = bool(clean(row.get("mutation_analysis")))
        if not has_intro and not has_analysis:
            flags.append("gene正文缺失")
        if has_analysis and not clean(row.get("c_hgvs")):
            flags.append("mutation_analysis缺少c_hgvs")
    elif section == "drug_sections":
        for field in ("drug_name", "relation", "clinical"):
            if not clean(row.get(field)):
                flags.append(f"{field}缺失")
    else:
        flags.append(f"未知section:{section}")
    return flags


def suggestion_for(row: dict[str, str]) -> tuple[str, str, str]:
    flags = risk_flags(row)
    source = clean(row.get("review_source"))
    section = clean(row.get("section"))
    has_intro = clean(row.get("has_intro")) == "是"
    has_analysis = clean(row.get("has_mutation_analysis")) == "是"
    has_relation = clean(row.get("has_relation")) == "是"
    has_clinical = clean(row.get("has_clinical")) == "是"

    if any(flag.startswith("PII:") or flag.startswith("章节串漏:") for flag in flags):
        return "建议修改后通过", "存在隐私/章节串漏风险，需在 reviewed_* 字段填写清理后的最终文本", "；".join(flags)
    if any(flag.startswith("样本上下文:") for flag in flags):
        return "建议暂缓", "疑似依赖同一样本其他变异上下文，不宜作为通用知识库直接入库", "；".join(flags)
    if flags:
        return "建议修改后通过", "存在结构或上下文风险，需人工修订后再入库", "；".join(flags)

    if source == "base_gene_kb_supported_gap_review" and section == "gene_sections":
        if has_intro and not has_analysis:
            return "建议通过", "基础知识库支撑的基因简介，结构完整且未见自动质控风险", ""
        if has_analysis and clean(row.get("c_hgvs")):
            return "建议通过", "基础知识库支撑的位点级变异解析，带c/p位点且未见自动质控风险", ""
    if source == "all_cancer_final_report_gene_intro_support" and section == "gene_sections":
        return "建议人工精审", "全癌种历史终版多来源基因简介，可复用但需确认适用于CRC358语境", ""
    if source == "historical_final_report_drug_pair_review" and section == "drug_sections" and has_relation and has_clinical:
        return "建议人工精审", "历史终版药物成对内容完整，药物证据需逐条医学确认", ""
    return "建议人工精审", "自动质控未发现硬风险，但来源/用途需要人工确认", ""


def triage_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in rows:
        suggestion, reason, flags = suggestion_for(row)
        reviewed_intro = ""
        reviewed_mutation_analysis = ""
        reviewed_relation = ""
        reviewed_clinical = ""
        if suggestion == "建议修改后通过":
            reviewed_intro = clean(row.get("intro"))
            reviewed_mutation_analysis = clean(row.get("mutation_analysis"))
            reviewed_relation = clean(row.get("relation"))
            reviewed_clinical = clean(row.get("clinical"))
        out.append(
            {
                "machine_suggestion": suggestion,
                "machine_reason": reason,
                "risk_flags": flags,
                "suggested_review_status": "通过" if suggestion == "建议通过" else "待医学审核",
                "review_source": clean(row.get("review_source")),
                "section": clean(row.get("section")),
                "gene": clean(row.get("gene")).upper(),
                "c_hgvs": clean(row.get("c_hgvs")),
                "p_hgvs": clean(row.get("p_hgvs")),
                "type": clean(row.get("type")),
                "applicability": clean(row.get("applicability")),
                "header": clean(row.get("header")),
                "drug_name": clean(row.get("drug_name")),
                "intro": clean(row.get("intro")),
                "mutation_analysis": clean(row.get("mutation_analysis")),
                "relation": clean(row.get("relation")),
                "clinical": clean(row.get("clinical")),
                "review_status": clean(row.get("review_status")),
                "reviewed_intro": reviewed_intro,
                "reviewed_mutation_analysis": reviewed_mutation_analysis,
                "reviewed_relation": reviewed_relation,
                "reviewed_clinical": reviewed_clinical,
                "review_notes": clean(row.get("review_notes")),
                "source_sheet": clean(row.get("_sheet")),
                "source_row": clean(row.get("_row_number")),
            }
        )
    return out


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(path: Path, rows: list[dict[str, str]], summary_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(ws, summary_rows)
    buckets = [
        ("质控总表", rows),
        ("建议通过", [row for row in rows if row["machine_suggestion"] == "建议通过"]),
        ("建议人工精审", [row for row in rows if row["machine_suggestion"] == "建议人工精审"]),
        ("建议修改后通过", [row for row in rows if row["machine_suggestion"] == "建议修改后通过"]),
        ("建议暂缓", [row for row in rows if row["machine_suggestion"] == "建议暂缓"]),
    ]
    for sheet_name, sheet_rows in buckets:
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, sheet_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_summary(rows: list[dict[str, str]]) -> dict[str, Any]:
    return {
        "status": "triage_only_not_medical_approval",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_rows": len(rows),
        "by_suggestion": dict(Counter(row["machine_suggestion"] for row in rows)),
        "by_source": dict(Counter(row["review_source"] for row in rows)),
        "by_section": dict(Counter(row["section"] for row in rows)),
        "risk_rows": sum(1 for row in rows if row["risk_flags"]),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--review-workbook", type=Path, default=DEFAULT_REVIEW_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    raw_rows = read_review_rows(args.review_workbook)
    triaged = triage_rows(raw_rows)
    summary = build_summary(triaged)
    out_xlsx = args.out_dir / "CRC358_batch12_机器医学质控分流_20260614.xlsx"
    out_json = args.out_dir / "crc358_batch12_medical_triage_summary_20260614.json"
    summary_rows = [
        {"项目": "定位", "结果": "机器质控分流；不是医学审批结论；不合入生产"},
        {"项目": "输入审核包", "结果": str(args.review_workbook)},
        {"项目": "总候选数", "结果": summary["total_rows"]},
        {"项目": "建议通过", "结果": summary["by_suggestion"].get("建议通过", 0)},
        {"项目": "建议人工精审", "结果": summary["by_suggestion"].get("建议人工精审", 0)},
        {"项目": "建议修改后通过", "结果": summary["by_suggestion"].get("建议修改后通过", 0)},
        {"项目": "建议暂缓", "结果": summary["by_suggestion"].get("建议暂缓", 0)},
        {"项目": "用法", "结果": "报告组可从“建议通过”开始批量确认；其余sheet逐条处理后再回填batch9审核包"},
    ]
    write_workbook(out_xlsx, triaged, summary_rows)
    args.out_dir.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"output_xlsx={out_xlsx}")
    print(f"output_json={out_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
