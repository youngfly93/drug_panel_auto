#!/usr/bin/env python3
"""Prepare CRC358 targeted-priority production promotion candidate files.

This is intentionally a dry-run/candidate builder. It writes merged production
candidate YAML files under ``tmp/`` and never overwrites production files.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_PROD_PART3 = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_PROD_CRC_RULES = Path("panels/crc_358_msi/rules/crc.yaml")
DEFAULT_TARGET_PART3 = Path("tmp/knowledge_buildout/reviewed_part3_knowledge_targeted_priority_v0.1.yaml")
DEFAULT_TARGET_DRUGS = Path("tmp/knowledge_buildout/crc358_targeted_drug_overrides_priority_v0.1.yaml")
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout/production_candidate")


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def gene_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (
        str(row.get("gene") or "").upper(),
        str(row.get("c_hgvs") or ""),
        str(row.get("p_hgvs") or ""),
    )


def drug_section_key(row: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        str(row.get("gene") or "").upper(),
        str(row.get("c_hgvs") or ""),
        str(row.get("p_hgvs") or ""),
        str(row.get("type") or "benefit"),
        str(row.get("applicability") or ""),
        str(row.get("drug_name") or ""),
        str(row.get("header") or ""),
    )


def rule_override_key(row: dict[str, Any]) -> tuple[str, str, str, tuple[str, ...], tuple[str, ...]]:
    benefit = tuple(str(x) for x in row.get("benefit_drugs") or [])
    caution_raw = row.get("caution_drugs")
    if isinstance(caution_raw, list):
        caution = tuple(str(x) for x in caution_raw)
    else:
        caution = (str(caution_raw or ""),)
    return (
        str(row.get("gene") or "").upper(),
        str(row.get("c_hgvs") or ""),
        str(row.get("applicability") or ""),
        benefit,
        caution,
    )


def merge_part3(prod: dict[str, Any], additive: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    merged = dict(prod)
    source = dict(prod.get("source") or {})
    source["candidate_merged_at"] = datetime.now(timezone.utc).isoformat()
    source["candidate_merge_policy"] = "production reviewed rows win; targeted-priority rows append only when key is absent"
    source["candidate_status"] = "pending_medical_review"
    merged["source"] = source

    rows_added: list[dict[str, Any]] = []
    gene_rows = list(prod.get("gene_sections") or [])
    seen_gene = {gene_key(row) for row in gene_rows}
    for row in additive.get("gene_sections") or []:
        key = gene_key(row)
        if key in seen_gene:
            continue
        gene_rows.append(row)
        seen_gene.add(key)
        rows_added.append({"file": "reviewed_part3_knowledge.yaml", "section": "gene_sections", "gene": row.get("gene", ""), "detail": "gene-level intro"})

    drug_rows = list(prod.get("drug_sections") or [])
    seen_drug = {drug_section_key(row) for row in drug_rows}
    for row in additive.get("drug_sections") or []:
        key = drug_section_key(row)
        if key in seen_drug:
            continue
        drug_rows.append(row)
        seen_drug.add(key)
        rows_added.append(
            {
                "file": "reviewed_part3_knowledge.yaml",
                "section": "drug_sections",
                "gene": row.get("gene", ""),
                "detail": f"{row.get('type', 'benefit')} / {row.get('applicability', '')}",
            }
        )

    merged["gene_sections"] = gene_rows
    merged["drug_sections"] = drug_rows
    return merged, rows_added


def merge_crc_rules(prod: dict[str, Any], additive: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    merged = dict(prod)
    existing = list(prod.get("reviewed_variant_overrides") or [])
    seen = {rule_override_key(row) for row in existing}
    rows_added: list[dict[str, Any]] = []
    for row in additive.get("reviewed_variant_overrides") or []:
        key = rule_override_key(row)
        if key in seen:
            continue
        existing.append(row)
        seen.add(key)
        rows_added.append(
            {
                "file": "crc.yaml",
                "section": "reviewed_variant_overrides",
                "gene": row.get("gene", ""),
                "detail": f"{row.get('applicability', '')} / {', '.join(row.get('benefit_drugs') or [])}",
            }
        )
    merged["reviewed_variant_overrides"] = existing
    return merged, rows_added


def write_summary_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "合入候选摘要"
    if not rows:
        rows = [{"file": "", "section": "", "gene": "", "detail": "无新增条目"}]
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 70)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 70))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--prod-part3", type=Path, default=DEFAULT_PROD_PART3)
    parser.add_argument("--prod-crc-rules", type=Path, default=DEFAULT_PROD_CRC_RULES)
    parser.add_argument("--target-part3", type=Path, default=DEFAULT_TARGET_PART3)
    parser.add_argument("--target-drugs", type=Path, default=DEFAULT_TARGET_DRUGS)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    prod_part3 = load_yaml(args.prod_part3)
    target_part3 = load_yaml(args.target_part3)
    prod_crc = load_yaml(args.prod_crc_rules)
    target_drugs = load_yaml(args.target_drugs)

    candidate_part3, part3_added = merge_part3(prod_part3, target_part3)
    candidate_crc, crc_added = merge_crc_rules(prod_crc, target_drugs)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    part3_out = args.out_dir / "reviewed_part3_knowledge.candidate.yaml"
    crc_out = args.out_dir / "crc.candidate.yaml"
    summary_json = args.out_dir / "promotion_summary.json"
    summary_xlsx = args.out_dir / "promotion_summary.xlsx"

    dump_yaml(part3_out, candidate_part3)
    dump_yaml(crc_out, candidate_crc)
    rows = part3_added + crc_added
    summary = {
        "status": "candidate_only_pending_medical_review",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "part3_added": len(part3_added),
        "crc_rule_added": len(crc_added),
        "outputs": {
            "part3": str(part3_out),
            "crc_rules": str(crc_out),
            "summary_xlsx": str(summary_xlsx),
        },
        "rows": rows,
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary_xlsx(summary_xlsx, rows)

    print(f"part3_added={len(part3_added)} crc_rule_added={len(crc_added)}")
    print(f"part3_out={part3_out}")
    print(f"crc_out={crc_out}")
    print(f"summary={summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
