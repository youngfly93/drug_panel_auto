#!/usr/bin/env python3
"""Promote approved CRC358 candidate knowledge rows into a reviewed overlay draft.

The default output is a draft YAML under ``tmp/``. Production overlay updates
must be explicit via ``--output panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml``.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import yaml
from openpyxl import load_workbook


DEFAULT_REVIEW = Path("tmp/knowledge_buildout/CRC358_医学知识库候选审核表_v0.1.xlsx")
DEFAULT_OUTPUT = Path("tmp/knowledge_buildout/reviewed_part3_knowledge_from_candidates.yaml")
APPROVED = {"通过", "修改后通过"}
PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]


def clean(value: object) -> str:
    return str(value or "").strip()


def has_pii(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def read_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "候选审核表" not in wb.sheetnames:
        raise ValueError("workbook missing sheet: 候选审核表")
    ws = wb["候选审核表"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(x) for x in rows[0]]
    out = []
    for raw in rows[1:]:
        item = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        if item.get("review_status") in APPROVED:
            text = item.get("reviewed_text") or item.get("candidate_text") or ""
            if has_pii(text):
                raise ValueError(f"approved row has PII risk: {item.get('candidate_id')}")
            item["final_text"] = text
            out.append(item)
    return out


def build_overlay(rows: list[dict[str, str]]) -> dict:
    gene_sections: dict[tuple[str, str, str], dict[str, str]] = defaultdict(dict)
    drug_sections: dict[tuple[str, str, str, str, str], dict[str, str]] = defaultdict(dict)

    for row in rows:
        gene = row.get("gene", "").upper()
        content_type = row.get("content_type", "")
        if content_type in {"gene_intro", "public_gene_intro"}:
            c_hgvs = ""
            p_hgvs = ""
        else:
            c_hgvs = row.get("c_hgvs", "")
            p_hgvs = row.get("p_hgvs", "")
        text = row.get("final_text", "")
        if not gene or not text:
            continue
        if content_type in {"gene_intro", "public_gene_intro"}:
            item = gene_sections[(gene, c_hgvs, p_hgvs)]
            item.update({"gene": gene, "c_hgvs": c_hgvs, "p_hgvs": p_hgvs})
            item.setdefault("intro", text)
        elif content_type in {"mutation_analysis", "public_mutation_analysis"}:
            item = gene_sections[(gene, c_hgvs, p_hgvs)]
            item.update({"gene": gene, "c_hgvs": c_hgvs, "p_hgvs": p_hgvs})
            item.setdefault("mutation_analysis", text)
        elif content_type in {"drug_relation", "drug_clinical"}:
            key = (
                gene,
                c_hgvs,
                p_hgvs,
                row.get("drug_type") or "benefit",
                row.get("drug_name") or row.get("header") or "",
            )
            item = drug_sections[key]
            item.update(
                {
                    "gene": gene,
                    "c_hgvs": c_hgvs,
                    "p_hgvs": p_hgvs,
                    "type": row.get("drug_type") or "benefit",
                    "header": row.get("header", ""),
                    "drug_name": row.get("drug_name", ""),
                }
            )
            if content_type == "drug_relation":
                item["relation"] = text
            else:
                item["clinical"] = text

    def strip_empty(row: dict[str, str]) -> dict[str, str]:
        return {k: v for k, v in row.items() if v}

    return {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "purpose": (
                "Reviewed Part 3 knowledge generated from CRC358 candidate review workbook. "
                "Only rows marked 通过/修改后通过 are included."
            ),
            "source_type": "candidate_review_workbook",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "gene_sections": [strip_empty(v) for v in gene_sections.values()],
        "drug_sections": [strip_empty(v) for v in drug_sections.values()],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--review-workbook", type=Path, default=DEFAULT_REVIEW)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows = read_rows(args.review_workbook)
    overlay = build_overlay(rows)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(yaml.safe_dump(overlay, allow_unicode=True, sort_keys=False), encoding="utf-8")
    print(f"approved_rows={len(rows)}")
    print(f"gene_sections={len(overlay['gene_sections'])} drug_sections={len(overlay['drug_sections'])}")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
