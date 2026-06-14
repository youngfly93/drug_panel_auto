#!/usr/bin/env python3
"""Audit CRC358 Part3 knowledge coverage for a batch of Excel inputs.

The output is an XLSX workbook for report-team review. It does not change the
knowledge base; reviewed content should still be curated in
``panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml``.
"""

from __future__ import annotations

import argparse
import shutil
import tempfile
import zipfile
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from reportgen.knowledge.gene_knowledge import GeneKnowledgeProvider


CLASS_LABELS = {"Ⅰ类", "Ⅱ类", "Ⅲ类", "I类", "II类", "III类"}
FEEDBACK_VARIANTS = [
    ("FGFR1", "c.1648G>T", "p.A550S"),
    ("PCLO", "c.11722C>A", "p.H3908N"),
    ("DNMT3A", "c.2322+1G>A", ""),
    ("EGFR", "c.2387G>A", "p.G796D"),
    ("TSC1", "c.1963C>T", "p.Q655*"),
]


@dataclass(frozen=True)
class VariantRow:
    sample_id: str
    gene: str
    c_hgvs: str
    p_hgvs: str
    frequency: str
    class_label: str
    mutation_type: str
    excel_path: str


def _clean(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _normalize_p_hgvs(value: Any) -> str:
    text = _clean(value)
    if text in {"*", "--", "-"}:
        return ""
    return text


def _normalize_class(value: Any) -> str:
    text = _clean(value)
    mapping = {
        "Ⅰ": "Ⅰ类",
        "Ⅱ": "Ⅱ类",
        "Ⅲ": "Ⅲ类",
        "i": "I类",
        "ii": "II类",
        "iii": "III类",
    }
    return mapping.get(text, text)


def _extract_input(input_path: Path) -> tuple[Path, tempfile.TemporaryDirectory[str] | None]:
    if input_path.is_dir():
        return input_path, None
    if input_path.suffix.lower() != ".zip":
        return input_path.parent, None
    tmp = tempfile.TemporaryDirectory(prefix="crc_part3_gap_")
    with zipfile.ZipFile(input_path) as zf:
        zf.extractall(tmp.name)
    return Path(tmp.name), tmp


def _excel_files(input_path: Path) -> list[Path]:
    root, tmp = _extract_input(input_path)
    try:
        files = sorted(
            p
            for p in root.rglob("*.xlsx")
            if not p.name.startswith("._") and not p.name.startswith("~$")
        )
        return [Path(str(p)) for p in files]
    finally:
        if tmp is not None:
            # Keep a materialized copy under a second temp dir while callers read
            # the returned paths would be unsafe, so only ZIP inputs are handled
            # directly by audit_input below.
            pass


def _read_variants_from_file(path: Path) -> list[VariantRow]:
    try:
        df = pd.read_excel(path, sheet_name="Variations")
    except Exception:
        return []

    required = {"Gene_Symbol", "cHGVS"}
    if not required.issubset(set(df.columns)):
        return []

    rows: list[VariantRow] = []
    sample_id = path.stem
    for _, row in df.iterrows():
        class_label = _normalize_class(row.get("ExistIn552"))
        if class_label not in CLASS_LABELS:
            continue
        gene = _clean(row.get("Gene_Symbol")).upper()
        c_hgvs = _clean(row.get("cHGVS"))
        p_hgvs = _normalize_p_hgvs(row.get("pHGVS_S") or row.get("pHGVS_A"))
        if not gene or not c_hgvs:
            continue
        rows.append(
            VariantRow(
                sample_id=sample_id,
                gene=gene,
                c_hgvs=c_hgvs,
                p_hgvs=p_hgvs,
                frequency=_clean(row.get("Freq(%)")),
                class_label=class_label,
                mutation_type=_clean(row.get("Function")),
                excel_path=str(path),
            )
        )
    return rows


def _load_provider(project_root: Path) -> GeneKnowledgeProvider:
    settings_path = project_root / "config/settings.yaml"
    settings = yaml.safe_load(settings_path.read_text(encoding="utf-8"))
    gene_cfg = dict(settings["knowledge_bases"]["gene_knowledge_db"])
    gene_cfg["reviewed_part3_overlay_path"] = (
        "panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml"
    )
    provider = GeneKnowledgeProvider(
        {
            "enabled": True,
            "gene_knowledge_db": gene_cfg,
            "gene_transcript_db": {"enabled": False},
        }
    )
    provider.load(str(project_root))
    return provider


def _coverage_status(provider: GeneKnowledgeProvider, row: VariantRow) -> dict[str, str]:
    exact_key = provider._variant_key(row.gene, row.c_hgvs, row.p_hgvs)
    no_p_key = provider._variant_key(row.gene, row.c_hgvs, "")
    gene_key = provider._hgvs_key(row.gene)

    if exact_key in provider._reviewed_gene_section_overrides:
        return {
            "status": "位点级已审核覆盖",
            "source": "reviewed_part3_knowledge.yaml",
            "risk": "低",
            "action": "已覆盖，复测确认即可",
        }
    if no_p_key in provider._reviewed_gene_section_overrides:
        return {
            "status": "位点级已审核覆盖",
            "source": "reviewed_part3_knowledge.yaml（无pHGVS匹配）",
            "risk": "低",
            "action": "已覆盖，复测确认即可",
        }
    if gene_key in provider._gene_level_section_overrides:
        return {
            "status": "基因级已审核覆盖",
            "source": "reviewed_part3_knowledge.yaml",
            "risk": "中",
            "action": "可先用；高频位点建议补位点级内容",
        }
    if row.gene in provider._reviewed_gene_analysis_cache:
        return {
            "status": "基础库可自动拼接",
            "source": "gene_knowledge_db.xlsx reviewed列",
            "risk": "中",
            "action": "建议报告组抽查措辞，必要时升级为已审核覆盖",
        }
    if row.gene in provider._gene_analysis_cache:
        return {
            "status": "基础库通用内容",
            "source": "gene_knowledge_db.xlsx 通用列",
            "risk": "高",
            "action": "优先从历史终版报告提取并审核入库",
        }
    return {
        "status": "缺失，可能固定套话",
        "source": "无匹配知识",
        "risk": "高",
        "action": "需补基因级兜底；若为高频/II类则补位点级内容",
    }


def _priority(status: dict[str, str], row: VariantRow, gene_counts: Counter[str]) -> str:
    if status["status"] == "位点级已审核覆盖":
        return "已处理"
    if row.class_label in {"Ⅱ类", "II类"} and status["risk"] == "高":
        return "P0"
    if gene_counts[row.gene] >= 2 and status["risk"] == "高":
        return "P1"
    if status["risk"] == "高":
        return "P2"
    return "P3"


def _append_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                max_len = max(max_len, min(len(str(c.value or "")), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)


def audit(input_path: Path, output_path: Path, project_root: Path) -> dict[str, Any]:
    provider = _load_provider(project_root)
    work_dir: Path
    tmp: tempfile.TemporaryDirectory[str] | None = None
    if input_path.is_dir():
        work_dir = input_path
    elif input_path.suffix.lower() == ".zip":
        tmp = tempfile.TemporaryDirectory(prefix="crc_part3_gap_")
        with zipfile.ZipFile(input_path) as zf:
            zf.extractall(tmp.name)
        work_dir = Path(tmp.name)
    else:
        work_dir = input_path.parent

    try:
        files = sorted(
            p
            for p in work_dir.rglob("*.xlsx")
            if not p.name.startswith("._") and not p.name.startswith("~$")
        )
        variants: list[VariantRow] = []
        for file in files:
            variants.extend(_read_variants_from_file(file))

        gene_counts = Counter(v.gene for v in variants)
        records: list[dict[str, Any]] = []
        for row in variants:
            status = _coverage_status(provider, row)
            records.append(
                {
                    "样本编号": row.sample_id,
                    "基因": row.gene,
                    "cHGVS": row.c_hgvs,
                    "pHGVS": row.p_hgvs,
                    "频率": row.frequency,
                    "等级": row.class_label,
                    "变异类型": row.mutation_type,
                    "覆盖状态": status["status"],
                    "来源": status["source"],
                    "固定套话风险": status["risk"],
                    "优先级": _priority(status, row, gene_counts),
                    "建议动作": status["action"],
                    "Excel": Path(row.excel_path).name,
                }
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "汇总"

        status_counts = Counter(r["覆盖状态"] for r in records)
        risk_counts = Counter(r["固定套话风险"] for r in records)
        priority_counts = Counter(r["优先级"] for r in records)
        summary_rows = [
            ["输入Excel数量", len(files)],
            ["纳入统计变异数", len(records)],
            ["位点级已审核覆盖", status_counts.get("位点级已审核覆盖", 0)],
            ["基因级已审核覆盖", status_counts.get("基因级已审核覆盖", 0)],
            ["基础库可自动拼接", status_counts.get("基础库可自动拼接", 0)],
            ["基础库通用内容", status_counts.get("基础库通用内容", 0)],
            ["缺失/固定套话高风险", risk_counts.get("高", 0)],
            ["P0需优先补库", priority_counts.get("P0", 0)],
            ["说明", "P0/P1优先从历史终版报告提取候选解读，人工审核后入库。"],
        ]
        _append_table(ws, ["指标", "结果"], summary_rows)

        ws2 = wb.create_sheet("缺口明细")
        headers = [
            "样本编号",
            "基因",
            "cHGVS",
            "pHGVS",
            "频率",
            "等级",
            "变异类型",
            "覆盖状态",
            "来源",
            "固定套话风险",
            "优先级",
            "建议动作",
            "Excel",
        ]
        detail_rows = [[r[h] for h in headers] for r in records]
        detail_rows.sort(key=lambda x: (str(x[10]), str(x[1]), str(x[2]), str(x[0])))
        _append_table(ws2, headers, detail_rows)

        ws3 = wb.create_sheet("高频缺口基因")
        by_gene: dict[str, Counter[str]] = defaultdict(Counter)
        for r in records:
            by_gene[r["基因"]][r["覆盖状态"]] += 1
            by_gene[r["基因"]]["总数"] += 1
            if r["固定套话风险"] == "高":
                by_gene[r["基因"]]["高风险"] += 1
            if r["优先级"] in {"P0", "P1"}:
                by_gene[r["基因"]]["优先补库"] += 1
        gene_rows = []
        for gene, counts in sorted(
            by_gene.items(), key=lambda item: (-item[1]["高风险"], -item[1]["总数"], item[0])
        ):
            gene_rows.append(
                [
                    gene,
                    counts["总数"],
                    counts["高风险"],
                    counts["优先补库"],
                    counts.get("位点级已审核覆盖", 0),
                    counts.get("基因级已审核覆盖", 0),
                    counts.get("基础库可自动拼接", 0),
                    counts.get("基础库通用内容", 0),
                    "先补位点级" if counts["优先补库"] else "维持抽查/后续补全",
                ]
            )
        _append_table(
            ws3,
            [
                "基因",
                "出现次数",
                "高风险次数",
                "P0/P1次数",
                "位点级已审核",
                "基因级已审核",
                "基础库自动拼接",
                "基础库通用",
                "建议",
            ],
            gene_rows,
        )

        ws4 = wb.create_sheet("报告组反馈核对")
        feedback_rows = []
        for gene, c_hgvs, p_hgvs in FEEDBACK_VARIANTS:
            row = VariantRow(
                sample_id="feedback_case",
                gene=gene,
                c_hgvs=c_hgvs,
                p_hgvs=p_hgvs,
                frequency="",
                class_label="反馈",
                mutation_type="",
                excel_path="",
            )
            status = _coverage_status(provider, row)
            feedback_rows.append(
                [
                    gene,
                    c_hgvs,
                    p_hgvs,
                    status["status"],
                    status["source"],
                    status["risk"],
                    status["action"],
                ]
            )
        _append_table(
            ws4,
            ["基因", "cHGVS", "pHGVS", "当前覆盖状态", "来源", "风险", "建议动作"],
            feedback_rows,
        )

        ws5 = wb.create_sheet("补库流程")
        _append_table(
            ws5,
            ["步骤", "说明"],
            [
                ["1. 统计缺口", "用本表锁定P0/P1位点，先处理报告组已反馈和高频II类变异。"],
                ["2. 提取候选内容", "从近几年终版报告中提取同基因/同位点的基因简介、变异解析、用药说明。"],
                ["3. 人工审核", "报告组确认适用癌种、证据不过期、措辞可以复用。"],
                ["4. 入库上线", "写入reviewed_part3_knowledge.yaml，保留来源和审核记录，重新跑同批Excel验证。"],
                ["5. 覆盖率复测", "比较固定套话高风险数量是否下降，未下降的继续进入下一轮补库。"],
            ],
        )

        candidate_records = [
            r
            for r in records
            if r["优先级"] in {"P0", "P1", "P2"}
            and r["覆盖状态"] in {"基础库通用内容", "缺失，可能固定套话"}
        ]
        ws6 = wb.create_sheet("待审核候选内容")
        candidate_rows = []
        seen_candidate_keys: set[tuple[str, str, str]] = set()
        record_by_key = {(r["基因"], r["cHGVS"], r["pHGVS"] or ""): r for r in records}
        for r in sorted(candidate_records, key=lambda item: (item["优先级"], item["基因"], item["cHGVS"])):
            key = (r["基因"], r["cHGVS"], r["pHGVS"] or "")
            if key in seen_candidate_keys:
                continue
            seen_candidate_keys.add(key)
            section = provider.build_gene_knowledge_section(
                gene=r["基因"],
                c_hgvs=r["cHGVS"],
                p_hgvs=r["pHGVS"] or "",
                frequency=0.0,
                mutation_type=r["变异类型"],
                has_drug=False,
                cancer_type="结直肠癌",
            )
            same_gene_examples = [
                f'{item["样本编号"]}:{item["cHGVS"]}/{item["pHGVS"] or "-"}'
                for item in records
                if item["基因"] == r["基因"]
            ][:5]
            candidate_rows.append(
                [
                    r["优先级"],
                    r["基因"],
                    r["cHGVS"],
                    r["pHGVS"],
                    r["等级"],
                    r["覆盖状态"],
                    "待报告组审核",
                    "未检索到精确历史位点时，先作为基因级候选；审核通过后再入正式库。",
                    section.get("intro", ""),
                    section.get("mutation_analysis", ""),
                    "；".join(same_gene_examples),
                ]
            )
        _append_table(
            ws6,
            [
                "优先级",
                "基因",
                "cHGVS",
                "pHGVS",
                "等级",
                "当前状态",
                "审核状态",
                "入库建议",
                "候选基因简介",
                "候选变异解析",
                "本批相关样本",
            ],
            candidate_rows,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {
            "excel_count": len(files),
            "variant_count": len(records),
            "status_counts": dict(status_counts),
            "risk_counts": dict(risk_counts),
            "priority_counts": dict(priority_counts),
            "output": str(output_path),
        }
    finally:
        if tmp is not None:
            tmp.cleanup()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path, help="Excel directory or ZIP")
    parser.add_argument("--output", required=True, type=Path, help="Output XLSX")
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    result = audit(args.input, args.output, args.project_root.resolve())
    print(yaml.safe_dump(result, allow_unicode=True, sort_keys=False))


if __name__ == "__main__":
    main()
