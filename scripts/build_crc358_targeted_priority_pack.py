#!/usr/bin/env python3
"""Build a targeted CRC358 knowledge review pack for priority gaps.

The outputs are review/draft artifacts under ``tmp/``. They do not update the
production CRC358 overlay or drug rules.
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_CANDIDATES = Path("tmp/knowledge_buildout/CRC358_医学知识库候选审核表_v0.1.xlsx")
DEFAULT_CIVIC = Path(
    "tmp/civic_crc358_public_candidate_pilot_20260611/"
    "CRC358_CIViC公共候选知识库pilot_20260611.xlsx"
)
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout")
TARGET_GENES = ("FGFR1", "PCLO", "DNMT3A", "EGFR", "TSC1")
SAFE_CONFIDENCE = {"高", "中"}
TODAY = "2026-06-11"

TSC1_DRUG_NAME = (
    "依维莫司（Everolimus）、依维莫司（Everolimus）+Buparlisib、"
    "西罗莫司（Sirolimus）、替西罗莫司（Temsirolimus）、Sapanisertib"
)
TSC1_LOF_RELATION = (
    "TSC1/TSC2 缺失或失活可导致 mTOR 信号通路活化，提示携带 TSC1 "
    "功能缺失型变异的肿瘤患者可能从 mTOR 抑制剂及相关通路抑制剂治疗中获益。"
    "mTOR 抑制剂包括依维莫司、西罗莫司和替西罗莫司等；Buparlisib、"
    "Sapanisertib 等相关通路抑制剂仍需结合适应证和临床证据谨慎评估。"
)
TSC1_LOF_CLINICAL = (
    "病例研究及基因组变异与依维莫司敏感性研究提示，MTOR、TSC1、TSC2 等 "
    "mTOR 通路变异患者中可观察到 mTOR 抑制剂临床获益；TSC1/TSC2 失活突变"
    "相关实体瘤中，西罗莫司白蛋白结合纳米颗粒、替西罗莫司、Sapanisertib 等"
    "临床试验正在或曾开展。该结论需结合癌种、变异类型、证据等级和临床适应证"
    "综合判断。"
)

PII_PATTERNS = [
    re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    re.compile(r"报告编号"),
    re.compile(r"姓名[:：]"),
    re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
]


def clean(value: Any) -> str:
    return str(value or "").strip()


def has_pii(text: str) -> bool:
    return any(pattern.search(text or "") for pattern in PII_PATTERNS)


def read_candidate_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["候选审核表"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [clean(value) for value in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        row = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        if row.get("gene", "").upper() in TARGET_GENES:
            out.append(row)
    return out


def read_pclo_civic_gap(path: Path) -> str:
    if not path.exists():
        return "未找到 CIViC pilot 文件；PCLO 仍需从历史终版或内部库补充。"
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in ("候选中文解读", "覆盖率判断"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = [clean(value) for value in rows[0]]
        for raw in rows[1:]:
            row = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
            if row.get("基因", "").upper() == "PCLO":
                return row.get("建议") or row.get("备注") or "CIViC覆盖不足，需补充内部知识库。"
    return "PCLO 未在 CIViC pilot 中形成可用中文候选；需补充内部知识库。"


def select_gene_intro_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    selected = []
    for row in rows:
        if row.get("source_type") != "historical_final_report":
            continue
        if row.get("content_type") != "gene_intro":
            continue
        if row.get("confidence") not in SAFE_CONFIDENCE:
            continue
        if row.get("current_reviewed_status") != "暂无reviewed覆盖":
            continue
        if has_pii(row.get("candidate_text", "")):
            continue
        selected.append(row)
    return selected


def review_action(row: dict[str, str]) -> str:
    gene = row.get("gene", "").upper()
    ctype = row.get("content_type", "")
    if ctype == "gene_intro" and row.get("confidence") in SAFE_CONFIDENCE:
        return "建议通过：基因简介可进入草稿 overlay"
    if gene == "TSC1" and ctype in {"drug_relation", "drug_clinical"}:
        return "重点审核：建议整理为 loss_of_function 条件药物规则"
    if gene == "EGFR" and ctype in {"drug_relation", "drug_clinical"}:
        return "暂不建议自动通过：EGFR 药物需避免误入免疫/靶向栏目"
    return "待医学审核：可作为后续补库素材"


def build_part3_overlay(gene_intro_rows: list[dict[str, str]]) -> dict[str, Any]:
    gene_sections = []
    seen_genes = set()
    for row in gene_intro_rows:
        gene = row.get("gene", "").upper()
        if gene in seen_genes:
            continue
        seen_genes.add(gene)
        gene_sections.append({"gene": gene, "intro": row.get("candidate_text", "")})
    drug_sections = [
        {
            "gene": "TSC1",
            "type": "benefit",
            "applicability": "loss_of_function",
            "drug_name": TSC1_DRUG_NAME,
            "header": "TSC1 功能缺失型变异相关潜在获益药物",
            "relation": TSC1_LOF_RELATION,
            "clinical": TSC1_LOF_CLINICAL,
        }
    ]
    return {
        "schema_version": 1,
        "source": {
            "panel": "crc_358_msi",
            "purpose": "Targeted priority review draft; requires medical review before production.",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "priority_genes": list(TARGET_GENES),
        },
        "gene_sections": gene_sections,
        "drug_sections": drug_sections,
    }


def build_drug_override_snippet() -> dict[str, Any]:
    return {
        "reviewed_variant_overrides": [
            {
                "gene": "TSC1",
                "applicability": "loss_of_function",
                "benefit_drugs": [
                    "依维莫司（C）",
                    "依维莫司+Buparlisib（C）",
                    "西罗莫司（C）",
                    "替西罗莫司（C）",
                    "Sapanisertib（C）",
                ],
                "caution_drugs": "--",
                "review_note": "草稿：仅对 TSC1 无义/移码/剪接等功能缺失型变异触发，需医学审核后再合入 crc.yaml。",
            }
        ]
    }


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 10), 70)
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                width = max(width, min(len(str(item.value or "")) + 2, 70))
                item.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_review_workbook(
    out_path: Path,
    rows: list[dict[str, str]],
    gene_intro_rows: list[dict[str, str]],
    pclo_gap: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(
        ws,
        [
            {"项目": "定位", "说明": "CRC358 重点基因定向补库审核包，不等同生产上线。"},
            {"项目": "目标基因", "说明": "、".join(TARGET_GENES)},
            {"项目": "安全边界", "说明": "不包含患者身份字段；生产 overlay 未修改。"},
            {"项目": "PCLO结论", "说明": pclo_gap},
            {"项目": "TSC1方案", "说明": "建议采用 loss_of_function 条件，避免普通错义变异误触发药物。"},
        ],
    )

    summary = []
    for gene in TARGET_GENES:
        gene_rows = [row for row in rows if row.get("gene", "").upper() == gene]
        intro = [row for row in gene_intro_rows if row.get("gene", "").upper() == gene]
        drug = [row for row in gene_rows if row.get("content_type") in {"drug_relation", "drug_clinical"}]
        summary.append(
            {
                "gene": gene,
                "候选条数": len(gene_rows),
                "可直接进草稿的基因简介": len(intro),
                "药物候选条数": len(drug),
                "本轮建议": (
                    "需补充内部/人工整理"
                    if gene == "PCLO"
                    else "基因简介可用；药物需审核" if gene in {"EGFR", "TSC1"} else "基因简介可用/继续补变异解析"
                ),
            }
        )
    ws = wb.create_sheet("优先结论")
    write_sheet(ws, summary)

    detail_rows = []
    for row in rows:
        detail_rows.append(
            {
                "candidate_id": row.get("candidate_id", ""),
                "gene": row.get("gene", ""),
                "content_type": row.get("content_type", ""),
                "content_type_cn": row.get("content_type_cn", ""),
                "confidence": row.get("confidence", ""),
                "source_count": row.get("source_count", ""),
                "c_hgvs": row.get("c_hgvs", ""),
                "p_hgvs": row.get("p_hgvs", ""),
                "drug_type": row.get("drug_type", ""),
                "drug_name": row.get("drug_name", ""),
                "建议处理": review_action(row),
                "candidate_text": row.get("candidate_text", ""),
                "review_status": "待医学审核",
                "reviewed_text": "",
                "review_notes": "",
            }
        )
    ws = wb.create_sheet("候选明细")
    write_sheet(ws, detail_rows)

    overlay_preview = [
        {"type": "gene_section", "gene": row.get("gene", ""), "applicability": "", "text": row.get("candidate_text", "")}
        for row in gene_intro_rows
    ]
    overlay_preview.append(
        {
            "type": "drug_section",
            "gene": "TSC1",
            "applicability": "loss_of_function",
            "text": f"{TSC1_LOF_RELATION}\n{TSC1_LOF_CLINICAL}",
        }
    )
    ws = wb.create_sheet("草稿overlay预览")
    write_sheet(ws, overlay_preview)

    ws = wb.create_sheet("缺口")
    write_sheet(
        ws,
        [
            {"gene": "PCLO", "缺口": pclo_gap, "建议": "当前历史 Part3 和 CIViC pilot 均无可直接复用中文 reviewed 内容。"},
            {"gene": "TSC1", "缺口": "药物候选为低置信/单来源，但可整理为 LoF 条件草稿。", "建议": "医学审核后再进入生产。"},
        ],
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--candidates", type=Path, default=DEFAULT_CANDIDATES)
    parser.add_argument("--civic", type=Path, default=DEFAULT_CIVIC)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    rows = read_candidate_rows(args.candidates)
    gene_intro_rows = select_gene_intro_rows(rows)
    pclo_gap = read_pclo_civic_gap(args.civic)

    review_xlsx = args.out_dir / "CRC358_重点基因定向补库审核表_v0.1.xlsx"
    part3_yaml = args.out_dir / "reviewed_part3_knowledge_targeted_priority_v0.1.yaml"
    drug_yaml = args.out_dir / "crc358_targeted_drug_overrides_priority_v0.1.yaml"

    write_review_workbook(review_xlsx, rows, gene_intro_rows, pclo_gap)
    part3_yaml.write_text(
        yaml.safe_dump(build_part3_overlay(gene_intro_rows), allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    drug_yaml.write_text(
        yaml.safe_dump(build_drug_override_snippet(), allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )

    print(f"target_rows={len(rows)} gene_intro_draft={len(gene_intro_rows)}")
    print(f"review_xlsx={review_xlsx}")
    print(f"part3_yaml={part3_yaml}")
    print(f"drug_yaml={drug_yaml}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
