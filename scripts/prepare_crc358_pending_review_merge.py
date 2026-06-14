#!/usr/bin/env python3
"""Build a CRC358 pending-review merge candidate from reviewed overlay drafts.

This script never overwrites production files. It appends pending medical-review
rows to a candidate YAML only when the production overlay does not already have
the same key, and writes an Excel summary for review.
"""

from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.build_crc358_knowledge_buildout import candidate_text_matches_context


DEFAULT_PROD_OVERLAY = Path("panels/crc_358_msi/rules/reviewed_part3_knowledge.yaml")
DEFAULT_PENDING_OVERLAYS = [
    Path("tmp/knowledge_buildout_after_batch7_gene_normalize_20260614/reviewed_part3_gene_gap_pending_review_batch6.yaml"),
    Path("tmp/knowledge_buildout_after_batch8_cross_cancer_gene_support_20260614/reviewed_part3_cross_cancer_gene_intro_pending_review_batch8.yaml"),
    Path("tmp/knowledge_buildout_after_batch7_gene_normalize_20260614/reviewed_part3_drug_pairs_pending_review_batch5.yaml"),
]
DEFAULT_OUT_DIR = Path("tmp/knowledge_buildout_after_batch9_pending_merge_20260614")
PII_PATTERNS = {
    "sample_id": re.compile(r"\b(?:LZ|LW|lz|lw)\d{5,}\b"),
    "report_no": re.compile(r"报告编号"),
    "name_label": re.compile(r"姓名[:：]"),
    "sender": re.compile(r"送检者"),
    "date": re.compile(r"\b20\d{2}[./-]\d{1,2}[./-]\d{1,2}\b"),
}


def clean(value: Any) -> str:
    return str(value or "").strip()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def dump_yaml(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8")


def gene_key(row: dict[str, Any]) -> tuple[str, str, str]:
    return (clean(row.get("gene")).upper(), clean(row.get("c_hgvs")), clean(row.get("p_hgvs")))


def drug_key(row: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        clean(row.get("gene")).upper(),
        clean(row.get("c_hgvs")),
        clean(row.get("p_hgvs")),
        clean(row.get("type") or "benefit"),
        clean(row.get("applicability")),
        clean(row.get("drug_name")),
        clean(row.get("header")),
    )


def source_label(path: Path, data: dict[str, Any]) -> str:
    source = data.get("source") or {}
    return clean(source.get("source_type")) or path.stem


def text_for_scan(row: dict[str, Any]) -> str:
    parts = []
    for key in ("gene", "c_hgvs", "p_hgvs", "intro", "mutation_analysis", "drug_name", "relation", "clinical"):
        parts.append(clean(row.get(key)))
    return "\n".join(parts)


def pii_hits(text: str) -> list[str]:
    return [name for name, pattern in PII_PATTERNS.items() if pattern.search(text)]


def gene_context_mismatch_reason(row: dict[str, Any]) -> str:
    gene = clean(row.get("gene")).upper()
    intro = clean(row.get("intro"))
    mutation_analysis = clean(row.get("mutation_analysis"))
    if intro and not candidate_text_matches_context(gene, "gene_intro", intro):
        return "intro正文首个基因与gene不一致"
    if mutation_analysis and not candidate_text_matches_context(gene, "mutation_analysis", mutation_analysis):
        return "mutation_analysis正文首个基因与gene不一致"
    return ""


def row_summary(
    row: dict[str, Any],
    section: str,
    review_source: str,
    action: str,
    reason: str = "",
) -> dict[str, Any]:
    intro = clean(row.get("intro"))
    mutation_analysis = clean(row.get("mutation_analysis"))
    relation = clean(row.get("relation"))
    clinical = clean(row.get("clinical"))
    return {
        "action": action,
        "reason": reason,
        "review_source": review_source,
        "section": section,
        "gene": clean(row.get("gene")).upper(),
        "c_hgvs": clean(row.get("c_hgvs")),
        "p_hgvs": clean(row.get("p_hgvs")),
        "type": clean(row.get("type") or row.get("drug_type")),
        "applicability": clean(row.get("applicability")),
        "header": clean(row.get("header")),
        "drug_name": clean(row.get("drug_name")),
        "has_intro": "是" if intro else "否",
        "has_mutation_analysis": "是" if mutation_analysis else "否",
        "has_relation": "是" if relation else "否",
        "has_clinical": "是" if clinical else "否",
        "intro": intro,
        "mutation_analysis": mutation_analysis,
        "relation": relation,
        "clinical": clinical,
        "review_status": "待医学审核" if action == "added" else "",
        "reviewed_intro": "",
        "reviewed_mutation_analysis": "",
        "reviewed_relation": "",
        "reviewed_clinical": "",
        "review_notes": "",
        "text_preview": clean(
            intro
            or mutation_analysis
            or relation
            or clinical
            or row.get("drug_name")
        )[:300],
    }


def merge_pending_overlays(
    prod: dict[str, Any],
    pending: list[tuple[Path, dict[str, Any]]],
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    merged = copy.deepcopy(prod)
    gene_rows = copy.deepcopy(prod.get("gene_sections") or [])
    drug_rows = copy.deepcopy(prod.get("drug_sections") or [])
    seen_gene = {gene_key(row) for row in gene_rows}
    seen_drug = {drug_key(row) for row in drug_rows}
    added: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []

    for path, data in pending:
        review_source = source_label(path, data)
        status = clean((data.get("source") or {}).get("status"))
        if status != "pending_medical_review":
            issues.append(
                {
                    "level": "warning",
                    "file": str(path),
                    "issue": "pending overlay source.status is not pending_medical_review",
                    "detail": status,
                }
            )

        for row in data.get("gene_sections") or []:
            key = gene_key(row)
            hits = pii_hits(text_for_scan(row))
            if hits:
                skipped.append(row_summary(row, "gene_sections", review_source, "skipped", f"PII风险:{','.join(hits)}"))
                continue
            mismatch_reason = gene_context_mismatch_reason(row)
            if mismatch_reason:
                skipped.append(row_summary(row, "gene_sections", review_source, "skipped", mismatch_reason))
                continue
            if clean(row.get("mutation_analysis")) and not clean(row.get("c_hgvs")):
                skipped.append(row_summary(row, "gene_sections", review_source, "skipped", "mutation_analysis缺少c_hgvs"))
                continue
            if key in seen_gene:
                skipped.append(row_summary(row, "gene_sections", review_source, "skipped", "生产或前序pending已存在同key"))
                continue
            row_copy = copy.deepcopy(row)
            gene_rows.append(row_copy)
            seen_gene.add(key)
            added.append(row_summary(row_copy, "gene_sections", review_source, "added"))

        for row in data.get("drug_sections") or []:
            key = drug_key(row)
            hits = pii_hits(text_for_scan(row))
            if hits:
                skipped.append(row_summary(row, "drug_sections", review_source, "skipped", f"PII风险:{','.join(hits)}"))
                continue
            if not clean(row.get("drug_name")):
                skipped.append(row_summary(row, "drug_sections", review_source, "skipped", "drug_name缺失"))
                continue
            if not (clean(row.get("relation")) and clean(row.get("clinical"))):
                skipped.append(row_summary(row, "drug_sections", review_source, "skipped", "relation/clinical未成对"))
                continue
            if key in seen_drug:
                skipped.append(row_summary(row, "drug_sections", review_source, "skipped", "生产或前序pending已存在同key"))
                continue
            row_copy = copy.deepcopy(row)
            drug_rows.append(row_copy)
            seen_drug.add(key)
            added.append(row_summary(row_copy, "drug_sections", review_source, "added"))

    source = dict(merged.get("source") or {})
    source["candidate_status"] = "pending_medical_review"
    source["candidate_merged_at"] = datetime.now(timezone.utc).isoformat()
    source["candidate_merge_policy"] = "production rows win; append valid pending rows only when key is absent"
    source["candidate_pending_sources"] = [str(path) for path, _ in pending]
    merged["source"] = source
    merged["gene_sections"] = gene_rows
    merged["drug_sections"] = drug_rows
    return merged, added, skipped, issues


def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 2, 12), 80)
        for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in col:
                width = max(width, min(len(str(cell.value or "")) + 2, 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_review_workbook(
    path: Path,
    summary_rows: list[dict[str, Any]],
    added: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    issues: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    write_sheet(ws, summary_rows)
    ws = wb.create_sheet("新增候选")
    write_sheet(ws, added)
    ws = wb.create_sheet("新增gene完整审核")
    write_sheet(ws, [row for row in added if row.get("section") == "gene_sections"])
    ws = wb.create_sheet("新增drug完整审核")
    write_sheet(ws, [row for row in added if row.get("section") == "drug_sections"])
    ws = wb.create_sheet("跳过候选")
    write_sheet(ws, skipped)
    ws = wb.create_sheet("问题检查")
    write_sheet(ws, issues)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--prod-overlay", type=Path, default=DEFAULT_PROD_OVERLAY)
    parser.add_argument("--pending-overlay", type=Path, action="append")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    pending_paths = args.pending_overlay or DEFAULT_PENDING_OVERLAYS
    prod = load_yaml(args.prod_overlay)
    pending = [(path, load_yaml(path)) for path in pending_paths]
    merged, added, skipped, issues = merge_pending_overlays(prod, pending)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    candidate_yaml = args.out_dir / "reviewed_part3_knowledge.pending_review_candidate_batch9.yaml"
    summary_json = args.out_dir / "pending_review_merge_summary_batch9.json"
    summary_xlsx = args.out_dir / "CRC358_batch9_待医学审核合入包_20260614.xlsx"

    dump_yaml(candidate_yaml, merged)
    summary = {
        "status": "candidate_only_pending_medical_review",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "prod_overlay": str(args.prod_overlay),
        "pending_overlays": [str(path) for path in pending_paths],
        "added": len(added),
        "skipped": len(skipped),
        "issues": len(issues),
        "candidate_yaml": str(candidate_yaml),
        "summary_xlsx": str(summary_xlsx),
        "added_by_section": {
            "gene_sections": sum(1 for row in added if row.get("section") == "gene_sections"),
            "drug_sections": sum(1 for row in added if row.get("section") == "drug_sections"),
        },
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_rows = [
        {"项目": "定位", "结果": "待医学审核合入包；不覆盖生产规则"},
        {"项目": "生产overlay", "结果": str(args.prod_overlay)},
        {"项目": "pending来源数", "结果": len(pending_paths)},
        {"项目": "新增候选", "结果": len(added)},
        {"项目": "跳过候选", "结果": len(skipped)},
        {"项目": "问题检查", "结果": len(issues)},
        {"项目": "candidate_yaml", "结果": str(candidate_yaml)},
        {"项目": "审核工作表", "结果": "优先审核“新增gene完整审核”和“新增drug完整审核”两个sheet"},
        {"项目": "允许审核结论", "结果": "通过 / 修改后通过 / 不通过 / 暂缓 / 待医学审核"},
        {"项目": "通过", "结果": "表示原始 intro/mutation_analysis/relation/clinical 正文可直接入库；reviewed_* 可留空"},
        {"项目": "修改后通过", "结果": "必须在对应 reviewed_* 字段填写最终定稿正文；否则自动合入工具会拒绝该行"},
        {"项目": "不合入状态", "结果": "不通过、暂缓、待医学审核、空白均不会进入生产候选"},
        {"项目": "不建议修改字段", "结果": "section/gene/c_hgvs/p_hgvs/type/applicability/drug_name/header 如需修改，建议退回重新生成候选"},
        {"项目": "落地工具", "结果": "scripts/apply_crc358_pending_review_decisions.py"},
    ]
    write_review_workbook(summary_xlsx, summary_rows, added, skipped, issues)

    print(f"added={len(added)} skipped={len(skipped)} issues={len(issues)}")
    print(f"candidate_yaml={candidate_yaml}")
    print(f"summary_json={summary_json}")
    print(f"summary_xlsx={summary_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
