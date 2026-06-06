from __future__ import annotations

import importlib.util
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "scripts" / "build_historical_report_inventory.py"


def _load_inventory_module():
    spec = importlib.util.spec_from_file_location("historical_inventory", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _write_minimal_docx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        "<w:p><w:r><w:t>第三部分：基因变异解析</w:t></w:r></w:p>"
        "<w:tbl><w:tr><w:tc><w:p><w:r><w:t>基因</w:t></w:r></w:p></w:tc></w:tr></w:tbl>"
        "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("word/document.xml", xml)


def test_historical_report_inventory_is_deidentified(tmp_path):
    module = _load_inventory_module()
    corpus = tmp_path / "各癌种基因报告近年汇总"
    report = corpus / "肠癌" / "测试患者-直肠癌-结直肠癌358基因+msi-mljy-case001-终版.docx"
    _write_minimal_docx(report)
    (corpus / "肠癌" / "._测试患者-直肠癌-结直肠癌358基因+msi-mljy-case001-终版.docx").write_bytes(
        b"resource"
    )

    records = module.build_inventory(corpus)
    output = tmp_path / "inventory.xlsx"
    summary = module.write_workbook(records, output, corpus)

    assert summary["valid_docx_non_resource"] == 1
    assert summary["parseable_docx"] == 1
    assert summary["resource_files_ignored"] == 1

    workbook_text = []
    wb = load_workbook(output, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            workbook_text.extend(str(value) for value in row if value is not None)
    combined = "\n".join(workbook_text)

    assert "测试患者" not in combined
    assert "case001" not in combined
    assert report.name not in combined
    assert "结直肠癌358基因+msi" in combined


def test_infer_product_family_never_returns_unrecognized_raw_filename():
    module = _load_inventory_module()
    filename = "测试患者-未知诊断-mljy-case001-终版.docx"

    assert module.infer_product_family(filename) == "未识别产品族"
