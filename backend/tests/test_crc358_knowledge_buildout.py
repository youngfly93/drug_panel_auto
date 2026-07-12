from pathlib import Path

from scripts.build_crc358_knowledge_buildout import (
    classify_product_family,
    candidate_text_matches_context,
    dedupe_candidates,
    extract_candidates_from_paragraphs,
    is_composite_context,
    make_candidate,
    normalize_gene_symbol,
    parse_variant_heading,
    should_extract_candidates,
    VariantContext,
)
from scripts.preapprove_crc358_candidates import should_preapprove
from scripts.promote_crc358_reviewed_candidates import build_overlay
from scripts.merge_reviewed_part3_overlays import merge
from scripts.build_crc358_targeted_priority_pack import (
    build_drug_override_snippet,
    build_part3_overlay,
)
from scripts.prepare_crc358_targeted_priority_promotion import (
    merge_crc_rules,
    merge_part3,
)
from scripts.prepare_crc358_drug_pair_review import (
    build_pending_overlay,
    split_drug_pairs,
)
from scripts.prepare_crc358_gene_gap_review import (
    build_gene_gap_review,
    build_pending_overlay as build_gene_gap_pending_overlay,
)
from scripts.prepare_crc358_cross_cancer_gene_support import (
    build_intro_review,
    build_pending_overlay as build_cross_cancer_intro_pending_overlay,
)
from scripts.prepare_crc358_pending_review_merge import merge_pending_overlays
from scripts.apply_crc358_pending_review_decisions import (
    apply_review_decisions,
    read_review_rows,
)
from scripts.check_crc358_knowledge_release_ready import (
    build_report as build_release_readiness_report,
    overlay_stats,
)
from scripts.triage_crc358_pending_medical_review import triage_rows
from scripts.prepare_crc358_triage_subset_release import release_row, workbook_row
from openpyxl import Workbook
import yaml


def test_release_checker_allows_gene_level_analysis_but_rejects_orphan_protein_hgvs(tmp_path):
    overlay = tmp_path / "reviewed_part3_knowledge.yaml"
    overlay.write_text(
        yaml.safe_dump(
            {
                "gene_sections": [
                    {
                        "gene": "ABCB1",
                        "mutation_analysis": "仅提供保守的基因级解释。",
                    },
                    {
                        "gene": "TP53",
                        "p_hgvs": "p.R175H",
                        "mutation_analysis": "这是缺少c.HGVS的变异级记录。",
                    },
                ],
                "drug_sections": [],
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    stats = overlay_stats(overlay)

    assert stats["mutation_analysis_without_c_hgvs"] == 1
    assert stats["details"]["mutation_analysis_without_c_hgvs"] == [
        {"gene": "TP53", "p_hgvs": "p.R175H"}
    ]


def test_classify_crc358_msi_final_report_name():
    product, label, has_msi, is_final = classify_product_family(
        "苏雨起-乙状结肠癌-结直肠癌358基因+msi-mljy-lz258792-终版.docx"
    )
    assert product == "crc_358_msi"
    assert label == "结直肠癌358基因+MSI"
    assert has_msi is True
    assert is_final is True


def test_candidate_extraction_scope_defaults_to_crc358_msi():
    assert should_extract_candidates("crc_358_msi", True, include_all_crc_panels=False)
    assert not should_extract_candidates("crc_301_msi", True, include_all_crc_panels=False)
    assert not should_extract_candidates("crc_358_msi", False, include_all_crc_panels=True)
    assert should_extract_candidates("crc_301_msi", True, include_all_crc_panels=True)
    assert not should_extract_candidates("crc_other", True, include_all_crc_panels=True)


def test_parse_variant_heading_extracts_gene_hgvs_and_frequency():
    ctx = parse_variant_heading("u TP53：c.821T>A，p.V274D；35.08%")
    assert ctx is not None
    assert ctx.gene == "TP53"
    assert ctx.c_hgvs == "c.821T>A"
    assert ctx.p_hgvs == "p.V274D"
    assert ctx.frequency == "35.08%"


def test_parse_variant_heading_normalizes_known_gene_typos():
    ctx = parse_variant_heading("u DMNT3A：c.2207G>A，p.R736H；0.62%")
    assert ctx is not None
    assert ctx.gene == "DNMT3A"
    assert normalize_gene_symbol("DNM3TA") == "DNMT3A"


def test_candidate_text_must_match_gene_context_when_text_names_gene():
    assert candidate_text_matches_context(
        "DNMT3A",
        "gene_intro",
        "DNMT3A基因编码DNA甲基转移酶。",
    )
    assert not candidate_text_matches_context(
        "DNMT3A",
        "gene_intro",
        "TSC2基因是一种肿瘤抑制基因。",
    )
    assert not candidate_text_matches_context(
        "MYH11",
        "gene_intro",
        "PALB2编码一种BRCA2连接蛋白。",
    )
    assert not candidate_text_matches_context(
        "DNMT3A",
        "mutation_analysis",
        "该突变在COSMIC数据库中暂无记载。TSC2基因编码的蛋白全长为1807个氨基酸。",
    )


def test_composite_wildtype_heading_is_not_single_gene_context():
    ctx = parse_variant_heading("KRAS/NRAS/BRAF：未突变")
    assert ctx is not None
    assert ctx.gene == "KRAS/NRAS/BRAF"
    assert is_composite_context(ctx) is True


def test_extract_part3_gene_and_drug_candidates_from_paragraphs():
    paragraphs = [
        "封面",
        "第三部分：基因变异及相应靶向/免疫药物解析",
        "基因变异解析",
        "u TP53：c.821T>A，p.V274D；35.08%",
        "基因简介：",
        "TP53基因是重要的抑癌基因之一，参与细胞周期调控。",
        "基因变异说明：",
        "该样本检出TP53基因c.821T>A错义突变，突变丰度为35.08%。",
        "基因变异解析：",
        "TP53突变在结直肠癌中较常见，可能与疾病发生发展相关。",
        "靶向药物/免疫用药提示解析",
        "潜在获益靶向/免疫药物解析",
        "TP53：c.821T>A，p.V274D突变相应靶向药物",
        "AZD1775",
        "基因变异与药物关联分析：",
        "TP53缺失肿瘤细胞可能对WEE1抑制剂敏感。",
        "药物疗效临床解析：",
        "一项临床研究显示AZD1775在实体瘤中具有潜在活性[27601554]。",
        "第四部分：附录",
    ]
    rows = extract_candidates_from_paragraphs(paragraphs, "hash1", "crc_358_msi")
    content_types = {row.content_type for row in rows}
    assert {"gene_intro", "variant_description", "mutation_analysis", "drug_relation", "drug_clinical"} <= content_types
    drug = next(row for row in rows if row.content_type == "drug_relation")
    assert drug.gene == "TP53"
    assert drug.drug_name == "AZD1775"
    assert drug.drug_type == "benefit"


def test_extract_part3_skips_orphan_drug_text_without_gene_context():
    paragraphs = [
        "封面",
        "第三部分：基因变异及相应靶向/免疫药物解析",
        "靶向药物/免疫用药提示解析",
        "潜在负相关靶向/免疫药物解析",
        "药物疗效临床解析：",
        "一项临床研究显示KRAS野生型患者可能从anti-EGFR治疗中获益。",
        "第四部分：附录",
    ]
    rows = extract_candidates_from_paragraphs(paragraphs, "hash1", "crc_358_msi")
    assert rows == []


def test_extract_part3_resets_composite_context_before_gene_intro():
    paragraphs = [
        "封面",
        "第三部分：基因变异及相应靶向/免疫药物解析",
        "基因变异解析",
        "KRAS/NRAS/BRAF：未突变",
        "基因简介：",
        "KRAS基因是致癌基因，属于RAS基因家族成员。",
        "基因变异解析：",
        "KRAS/NRAS/BRAF均未突变时可结合临床评估anti-EGFR治疗。",
        "第四部分：附录",
    ]
    rows = extract_candidates_from_paragraphs(paragraphs, "hash1", "crc_301_msi")
    assert rows == []


def test_dedupe_candidates_counts_distinct_sources():
    paragraphs = [
        "第三部分：基因变异及相应靶向/免疫药物解析",
        "基因变异解析",
        "u TP53：c.821T>A，p.V274D；35.08%",
        "基因简介：",
        "TP53基因是重要的抑癌基因之一，参与细胞周期调控。",
        "第四部分：附录",
    ]
    a = extract_candidates_from_paragraphs(paragraphs, "hash1", "crc_358_msi")
    b = extract_candidates_from_paragraphs(paragraphs, "hash2", "crc_358_msi")
    rows = dedupe_candidates(a + b)
    assert len(rows) == 1
    assert rows[0].source_count == 2
    assert rows[0].confidence == "中"


def test_dedupe_gene_intro_counts_cross_crc_panel_sources():
    paragraphs = [
        "第三部分：基因变异及相应靶向/免疫药物解析",
        "基因变异解析",
        "u ROS1：c.1A>T；",
        "基因简介：",
        "ROS1基因是受体酪氨酸激酶家族成员之一。",
        "第四部分：附录",
    ]
    a = extract_candidates_from_paragraphs(paragraphs, "hash1", "crc_301_msi")
    b = extract_candidates_from_paragraphs(paragraphs, "hash2", "crc_358_msi")
    rows = dedupe_candidates(a + b)
    assert len(rows) == 1
    assert rows[0].source_count == 2
    assert rows[0].confidence == "中"
    assert rows[0].product_family == "crc_301_msi;crc_358_msi"


def test_variant_description_is_classified_as_dynamic_not_reviewed_gap():
    candidate = make_candidate(
        "hash1",
        "historical_final_report",
        "crc_358_msi",
        VariantContext(gene="TP53", c_hgvs="c.821T>A", p_hgvs="p.V274D"),
        "variant_description",
        "该样本检出TP53基因c.821T>A错义突变。",
    )
    assert candidate is not None
    candidate.current_reviewed_status = "暂无reviewed覆盖"
    candidate.confidence = "低"
    row = candidate.as_row()
    assert row["kb_gap_class"] == "动态生成项_不入reviewed知识库"


def test_drug_candidate_without_drug_name_requires_manual_cleanup():
    candidate = make_candidate(
        "hash1",
        "historical_final_report",
        "crc_358_msi",
        VariantContext(gene="DDR2", c_hgvs="c.1466G>A", p_hgvs="p.R489Q"),
        "drug_relation",
        "该样本检出DDR2突变，可能与某些药物相关。",
        drug_type="benefit",
    )
    assert candidate is not None
    candidate.current_reviewed_status = "暂无reviewed覆盖"
    row = candidate.as_row()
    assert row["kb_gap_class"] == "药物解析缺药名_需人工整理"


def test_machine_preapprove_only_allows_conservative_historical_rows():
    row = {
        "source_type": "historical_final_report",
        "confidence": "高",
        "content_type": "mutation_analysis",
        "current_reviewed_status": "暂无reviewed覆盖",
        "gene": "TSC1",
        "c_hgvs": "c.1963C>T",
        "p_hgvs": "p.Q655*",
        "candidate_text": "TSC1突变可能影响mTOR信号通路。",
    }
    ok, _ = should_preapprove(row)
    assert ok is True

    row["source_type"] = "public_civic_candidate"
    ok, reason = should_preapprove(row)
    assert ok is False
    assert "非历史终版来源" in reason


def test_machine_preapprove_rejects_incomplete_or_medium_drug_rows():
    row = {
        "source_type": "historical_final_report",
        "confidence": "高",
        "source_count": "4",
        "content_type": "drug_relation",
        "current_reviewed_status": "暂无reviewed覆盖",
        "gene": "KRAS",
        "c_hgvs": "c.35G>T",
        "p_hgvs": "p.G12V",
        "drug_type": "caution",
        "drug_name": "",
        "candidate_text": "KRAS激活突变与anti-EGFR抗体药物疗效负相关。",
    }
    ok, reason = should_preapprove(row)
    assert ok is False
    assert "缺药名" in reason

    row["drug_name"] = "西妥昔单抗（Cetuximab）"
    row["confidence"] = "中"
    ok, reason = should_preapprove(row)
    assert ok is False
    assert "药物解析需高置信" in reason


def test_machine_preapprove_limits_cross_panel_rows_to_gene_intro():
    row = {
        "source_type": "historical_final_report",
        "product_family": "crc_301_msi",
        "confidence": "高",
        "source_count": "4",
        "content_type": "mutation_analysis",
        "current_reviewed_status": "暂无reviewed覆盖",
        "gene": "PTEN",
        "c_hgvs": "c.388C>T",
        "p_hgvs": "p.R130*",
        "candidate_text": "PTEN突变可能影响PI3K/AKT信号通路。",
    }
    ok, reason = should_preapprove(row)
    assert ok is False
    assert "跨panel候选仅允许基因简介" in reason

    row["content_type"] = "gene_intro"
    row["c_hgvs"] = ""
    row["p_hgvs"] = ""
    ok, _ = should_preapprove(row)
    assert ok is True


def test_promote_gene_intro_as_gene_level_overlay():
    overlay = build_overlay(
        [
            {
                "gene": "TSC1",
                "c_hgvs": "c.1963C>T",
                "p_hgvs": "p.Q655*",
                "content_type": "gene_intro",
                "final_text": "TSC1 gene intro",
            }
        ]
    )
    assert overlay["gene_sections"] == [{"gene": "TSC1", "intro": "TSC1 gene intro"}]


def test_promote_keeps_first_gene_intro_when_duplicate_gene_rows_exist():
    overlay = build_overlay(
        [
            {
                "gene": "RB1",
                "content_type": "gene_intro",
                "final_text": "high-confidence RB1 intro",
            },
            {
                "gene": "RB1",
                "content_type": "gene_intro",
                "final_text": "lower-confidence RB1 intro",
            },
        ]
    )
    assert overlay["gene_sections"] == [{"gene": "RB1", "intro": "high-confidence RB1 intro"}]


def test_merge_overlay_keeps_base_precedence_and_appends_absent_keys():
    base = {
        "source": {"panel": "crc_358_msi"},
        "gene_sections": [{"gene": "TP53", "intro": "base"}],
        "drug_sections": [{"gene": "KRAS", "type": "caution", "drug_name": "DrugA", "relation": "base"}],
    }
    additive = {
        "source": {"source_type": "machine"},
        "gene_sections": [{"gene": "TP53", "intro": "new"}, {"gene": "TSC1", "intro": "added"}],
        "drug_sections": [
            {"gene": "KRAS", "type": "caution", "drug_name": "DrugA", "relation": "new"},
            {"gene": "TSC1", "type": "benefit", "drug_name": "DrugB", "relation": "added"},
        ],
    }
    merged = merge(base, additive)
    assert merged["gene_sections"] == [
        {"gene": "TP53", "intro": "base"},
        {"gene": "TSC1", "intro": "added"},
    ]
    assert merged["drug_sections"] == [
        {"gene": "KRAS", "type": "caution", "drug_name": "DrugA", "relation": "base"},
        {"gene": "TSC1", "type": "benefit", "drug_name": "DrugB", "relation": "added"},
    ]


def test_targeted_priority_part3_overlay_limits_tsc1_drug_to_lof():
    overlay = build_part3_overlay(
        [
            {
                "gene": "FGFR1",
                "candidate_text": "FGFR1 intro",
            }
        ]
    )
    assert overlay["gene_sections"] == [{"gene": "FGFR1", "intro": "FGFR1 intro"}]
    tsc1 = overlay["drug_sections"][0]
    assert tsc1["gene"] == "TSC1"
    assert tsc1["applicability"] == "loss_of_function"
    assert "依维莫司" in tsc1["drug_name"]


def test_targeted_priority_drug_override_snippet_is_conditioned():
    snippet = build_drug_override_snippet()
    row = snippet["reviewed_variant_overrides"][0]
    assert row["gene"] == "TSC1"
    assert row["applicability"] == "loss_of_function"
    assert "依维莫司（C）" in row["benefit_drugs"]


def test_prepare_promotion_appends_only_absent_part3_rows():
    base = {
        "source": {"panel": "crc_358_msi"},
        "gene_sections": [{"gene": "FGFR1", "intro": "base"}],
        "drug_sections": [],
    }
    additive = {
        "gene_sections": [{"gene": "FGFR1", "intro": "new"}, {"gene": "TSC1", "intro": "added"}],
        "drug_sections": [
            {
                "gene": "TSC1",
                "type": "benefit",
                "applicability": "loss_of_function",
                "drug_name": "依维莫司",
            }
        ],
    }
    merged, added = merge_part3(base, additive)
    assert merged["gene_sections"] == [
        {"gene": "FGFR1", "intro": "base"},
        {"gene": "TSC1", "intro": "added"},
    ]
    assert len(merged["drug_sections"]) == 1
    assert len(added) == 2


def test_prepare_promotion_appends_only_absent_crc_rule_rows():
    existing = {
        "reviewed_variant_overrides": [
            {
                "gene": "TSC1",
                "applicability": "loss_of_function",
                "benefit_drugs": ["依维莫司（C）"],
                "caution_drugs": "--",
            }
        ]
    }
    additive = {
        "reviewed_variant_overrides": [
            {
                "gene": "TSC1",
                "applicability": "loss_of_function",
                "benefit_drugs": ["依维莫司（C）"],
                "caution_drugs": "--",
            },
            {
                "gene": "TSC1",
                "applicability": "loss_of_function",
                "benefit_drugs": ["Sapanisertib（C）"],
                "caution_drugs": "--",
            },
        ]
    }
    merged, added = merge_crc_rules(existing, additive)
    assert len(merged["reviewed_variant_overrides"]) == 2
    assert added == [
        {
            "file": "crc.yaml",
            "section": "reviewed_variant_overrides",
            "gene": "TSC1",
            "detail": "loss_of_function / Sapanisertib（C）",
        }
    ]


def test_drug_pair_review_keeps_only_complete_relation_and_clinical_pairs():
    rows = [
        {
            "candidate_id": "r1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "content_type": "drug_relation",
            "gene": "XRCC2",
            "c_hgvs": "c.190C>T",
            "p_hgvs": "p.R64*",
            "drug_type": "benefit",
            "drug_name": "尼拉帕利（Niraparib）",
            "candidate_text": "XRCC2突变可能提示PARP抑制剂获益。",
        },
        {
            "candidate_id": "c1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "content_type": "drug_clinical",
            "gene": "XRCC2",
            "c_hgvs": "c.190C>T",
            "p_hgvs": "p.R64*",
            "drug_type": "benefit",
            "drug_name": "尼拉帕利（Niraparib）",
            "candidate_text": "临床研究提示PARP抑制剂在相关DNA修复缺陷中可能获益。",
        },
    ]
    reviewable, cleanup = split_drug_pairs(rows)
    assert cleanup == []
    assert len(reviewable) == 1
    assert reviewable[0]["gene"] == "XRCC2"
    overlay = build_pending_overlay(reviewable)
    assert overlay["source"]["status"] == "pending_medical_review"
    assert overlay["drug_sections"][0]["drug_name"] == "尼拉帕利（Niraparib）"


def test_drug_pair_review_sends_missing_drug_name_to_cleanup():
    rows = [
        {
            "candidate_id": "r1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "content_type": "drug_relation",
            "gene": "DDR2",
            "c_hgvs": "c.1466G>A",
            "p_hgvs": "p.R489Q",
            "drug_type": "benefit",
            "drug_name": "",
            "candidate_text": "DDR2突变可能与药物响应相关。",
        }
    ]
    reviewable, cleanup = split_drug_pairs(rows)
    assert reviewable == []
    assert cleanup[0]["整理原因"] == "缺少明确药名"


def test_drug_pair_review_sends_sample_context_to_cleanup():
    rows = [
        {
            "candidate_id": "r1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "content_type": "drug_relation",
            "gene": "XRCC2",
            "c_hgvs": "c.190C>T",
            "p_hgvs": "p.R64*",
            "drug_type": "benefit",
            "drug_name": "奥拉帕利（Olaparib）",
            "candidate_text": "该样本同时检出MSH6基因突变，因此可能从PARP抑制剂获益。",
        },
        {
            "candidate_id": "c1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "content_type": "drug_clinical",
            "gene": "XRCC2",
            "c_hgvs": "c.190C>T",
            "p_hgvs": "p.R64*",
            "drug_type": "benefit",
            "drug_name": "奥拉帕利（Olaparib）",
            "candidate_text": "临床研究提示PARP抑制剂可能获益。",
        },
    ]
    reviewable, cleanup = split_drug_pairs(rows)
    assert reviewable == []
    assert any("包含同样本其他变异上下文" in row["整理原因"] for row in cleanup)


def test_gene_gap_review_uses_base_kb_as_supporting_evidence():
    rows = [
        {
            "candidate_id": "i1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "kb_gap_class": "低置信待补证据",
            "content_type": "gene_intro",
            "gene": "RAD52",
            "source_count": "1",
            "candidate_text": "RAD52基因参与DNA双链断裂修复过程。",
        },
        {
            "candidate_id": "m1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "kb_gap_class": "低置信待补证据",
            "content_type": "mutation_analysis",
            "gene": "RAD52",
            "c_hgvs": "c.1A>T",
            "p_hgvs": "p.K1*",
            "source_count": "1",
            "candidate_text": "RAD52基因突变可能影响DNA损伤修复能力。",
        },
    ]
    base = {
        "RAD52": {
            "base_intro": "RAD52基因参与DNA双链断裂修复过程。",
            "base_mutation_analysis": "RAD52基因突变可能影响DNA损伤修复能力。",
        }
    }
    supported, missing = build_gene_gap_review(rows, base)
    assert missing == []
    assert len(supported) == 1
    overlay = build_gene_gap_pending_overlay(supported)
    assert overlay["source"]["status"] == "pending_medical_review"
    assert overlay["gene_sections"] == [
        {
            "gene": "RAD52",
            "c_hgvs": "",
            "p_hgvs": "",
            "intro": "RAD52基因参与DNA双链断裂修复过程。",
            "mutation_analysis": "",
        },
        {
            "gene": "RAD52",
            "c_hgvs": "c.1A>T",
            "p_hgvs": "p.K1*",
            "intro": "",
            "mutation_analysis": "RAD52基因突变可能影响DNA损伤修复能力。",
        }
    ]


def test_gene_gap_review_keeps_base_missing_rows_out_of_overlay():
    rows = [
        {
            "candidate_id": "i1",
            "source_type": "historical_final_report",
            "current_reviewed_status": "暂无reviewed覆盖",
            "kb_gap_class": "低置信待补证据",
            "content_type": "gene_intro",
            "gene": "PCLO",
            "source_count": "1",
            "candidate_text": "PCLO基因功能仍需补充证据。",
        }
    ]
    supported, missing = build_gene_gap_review(rows, {})
    assert supported == []
    assert missing[0]["gene"] == "PCLO"
    overlay = build_gene_gap_pending_overlay(supported)
    assert overlay["gene_sections"] == []


def test_cross_cancer_intro_review_requires_multiple_sources_for_overlay():
    rows = [
        {
            "gene": "ARID5B",
            "content_type": "gene_intro",
            "candidate_text": "ARID5B基因参与染色质调控过程。",
            "source_hash": "h1",
            "source_family": "肺癌",
        },
        {
            "gene": "ARID5B",
            "content_type": "gene_intro",
            "candidate_text": "ARID5B基因参与染色质调控过程。",
            "source_hash": "h2",
            "source_family": "肠癌",
        },
        {
            "gene": "PAX3",
            "content_type": "gene_intro",
            "candidate_text": "PAX3基因参与发育调控过程。",
            "source_hash": "h3",
            "source_family": "肉瘤",
        },
        {
            "gene": "ARID5B",
            "content_type": "mutation_analysis",
            "candidate_text": "ARID5B变异解析只作参考。",
            "c_hgvs": "c.1A>T",
            "p_hgvs": "p.K1*",
            "source_hash": "h4",
            "source_family": "肺癌",
        },
    ]
    intro_rows, single_source_rows, mutation_reference = build_intro_review(rows)
    assert [row["gene"] for row in intro_rows] == ["ARID5B"]
    assert [row["gene"] for row in single_source_rows] == ["PAX3"]
    assert mutation_reference[0]["usage_policy"].startswith("参考材料")
    overlay = build_cross_cancer_intro_pending_overlay(intro_rows)
    assert overlay["gene_sections"] == [
        {
            "gene": "ARID5B",
            "c_hgvs": "",
            "p_hgvs": "",
            "intro": "ARID5B基因参与染色质调控过程。",
            "mutation_analysis": "",
        }
    ]


def test_pending_review_merge_appends_only_absent_valid_rows():
    prod = {
        "gene_sections": [{"gene": "TP53", "c_hgvs": "", "p_hgvs": "", "intro": "prod"}],
        "drug_sections": [],
    }
    pending = [
        (
            Path("pending.yaml"),
            {
                "source": {"status": "pending_medical_review", "source_type": "unit_test"},
                "gene_sections": [
                    {"gene": "TP53", "c_hgvs": "", "p_hgvs": "", "intro": "skip existing"},
                    {"gene": "RAD52", "c_hgvs": "", "p_hgvs": "", "intro": "add intro"},
                    {"gene": "RAD52", "c_hgvs": "", "p_hgvs": "", "mutation_analysis": "bad analysis"},
                    {"gene": "MYH11", "c_hgvs": "", "p_hgvs": "", "intro": "PALB2基因参与DNA损伤修复。"},
                ],
                "drug_sections": [
                    {
                        "gene": "XRCC2",
                        "c_hgvs": "c.1A>T",
                        "p_hgvs": "p.K1*",
                        "type": "benefit",
                        "drug_name": "尼拉帕利",
                        "relation": "relation",
                        "clinical": "clinical",
                    },
                    {
                        "gene": "DDR2",
                        "type": "benefit",
                        "drug_name": "",
                        "relation": "relation",
                        "clinical": "clinical",
                    },
                ],
            },
        )
    ]
    merged, added, skipped, issues = merge_pending_overlays(prod, pending)
    assert issues == []
    assert len(added) == 2
    assert len(skipped) == 4
    assert any(row["gene"] == "MYH11" and "intro正文首个基因" in row["reason"] for row in skipped)
    added_intro = next(row for row in added if row["gene"] == "RAD52")
    assert added_intro["intro"] == "add intro"
    assert added_intro["review_status"] == "待医学审核"
    added_drug = next(row for row in added if row["section"] == "drug_sections")
    assert added_drug["relation"] == "relation"
    assert added_drug["clinical"] == "clinical"
    assert len(merged["gene_sections"]) == 2
    assert len(merged["drug_sections"]) == 1
    assert merged["source"]["candidate_status"] == "pending_medical_review"


def _write_batch9_review_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    headers = [
        "action",
        "reason",
        "review_source",
        "section",
        "gene",
        "c_hgvs",
        "p_hgvs",
        "type",
        "applicability",
        "header",
        "drug_name",
        "intro",
        "mutation_analysis",
        "relation",
        "clinical",
        "review_status",
        "reviewed_intro",
        "reviewed_mutation_analysis",
        "reviewed_relation",
        "reviewed_clinical",
        "review_notes",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "新增gene完整审核"
    ws.append(headers)
    for row in rows:
        if row.get("section") == "gene_sections":
            ws.append([row.get(header, "") for header in headers])
    ws = wb.create_sheet("新增drug完整审核")
    ws.append(headers)
    for row in rows:
        if row.get("section") == "drug_sections":
            ws.append([row.get(header, "") for header in headers])
    wb.save(path)


def test_apply_pending_review_decisions_promotes_only_approved_rows(tmp_path):
    workbook = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        workbook,
        [
            {
                "section": "gene_sections",
                "gene": "RAD52",
                "intro": "RAD52基因参与DNA损伤修复。",
                "review_status": "通过",
            },
            {
                "section": "gene_sections",
                "gene": "XRCC2",
                "c_hgvs": "c.190C>T",
                "p_hgvs": "p.R64*",
                "mutation_analysis": "原始解析。",
                "review_status": "修改后通过",
                "reviewed_mutation_analysis": "XRCC2变异可能影响同源重组修复能力。",
            },
            {
                "section": "drug_sections",
                "gene": "XRCC2",
                "c_hgvs": "c.190C>T",
                "p_hgvs": "p.R64*",
                "type": "benefit",
                "drug_name": "尼拉帕利（Niraparib）",
                "relation": "原始关联。",
                "clinical": "原始临床。",
                "review_status": "修改后通过",
                "reviewed_relation": "XRCC2功能缺失可能提示PARP抑制剂潜在获益。",
                "reviewed_clinical": "相关DNA修复缺陷肿瘤中可结合临床评估PARP抑制剂证据。",
            },
            {
                "section": "gene_sections",
                "gene": "PCLO",
                "intro": "PCLO待审核文本。",
                "review_status": "待医学审核",
            },
        ],
    )
    rows = read_review_rows(workbook)
    merged, added, skipped, issues = apply_review_decisions(
        {"gene_sections": [{"gene": "TP53", "intro": "prod"}], "drug_sections": []},
        rows,
    )
    assert issues == []
    assert len(added) == 3
    assert any(row["gene"] == "PCLO" and row["reason"] == "未审核通过" for row in skipped)
    assert {"gene": "RAD52", "intro": "RAD52基因参与DNA损伤修复。"} in merged["gene_sections"]
    assert {
        "gene": "XRCC2",
        "c_hgvs": "c.190C>T",
        "p_hgvs": "p.R64*",
        "mutation_analysis": "XRCC2变异可能影响同源重组修复能力。",
    } in merged["gene_sections"]
    assert merged["drug_sections"][0]["relation"] == "XRCC2功能缺失可能提示PARP抑制剂潜在获益。"
    assert merged["drug_sections"][0]["clinical"] == "相关DNA修复缺陷肿瘤中可结合临床评估PARP抑制剂证据。"


def test_apply_pending_review_decisions_requires_final_text_for_modified_rows(tmp_path):
    workbook = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        workbook,
        [
            {
                "section": "gene_sections",
                "gene": "RAD52",
                "intro": "原始简介。",
                "review_status": "修改后通过",
            }
        ],
    )
    merged, added, skipped, issues = apply_review_decisions(
        {"gene_sections": [], "drug_sections": []},
        read_review_rows(workbook),
    )
    assert merged["gene_sections"] == []
    assert added == []
    assert skipped == []
    assert issues[0]["issue"] == "reviewed_intro缺失：修改后通过必须填写最终定稿文本"


def test_apply_pending_review_decisions_keeps_production_key_precedence(tmp_path):
    workbook = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        workbook,
        [
            {
                "section": "gene_sections",
                "gene": "TP53",
                "intro": "new intro",
                "review_status": "通过",
            }
        ],
    )
    merged, added, skipped, issues = apply_review_decisions(
        {"gene_sections": [{"gene": "TP53", "intro": "prod intro"}], "drug_sections": []},
        read_review_rows(workbook),
    )
    assert issues == []
    assert added == []
    assert skipped[0]["reason"] == "生产或前序审核行已存在同key"
    assert merged["gene_sections"] == [{"gene": "TP53", "intro": "prod intro"}]


def _write_context_retest(path: Path, statuses: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["指标", "结果"])
    ws.append(["输入Excel数", len(statuses)])
    ws.append(["PASS样本数", sum(1 for status in statuses if status == "PASS")])
    ws.append(["FAIL样本数", sum(1 for status in statuses if status != "PASS")])
    ws = wb.create_sheet("样本对比")
    ws.append(["sample_id", "status"])
    for idx, status in enumerate(statuses, 1):
        ws.append([f"LZ{idx:06d}", status])
    wb.save(path)


def test_release_readiness_gate_rejects_pending_review_rows(tmp_path):
    review = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        review,
        [
            {
                "section": "gene_sections",
                "gene": "RAD52",
                "intro": "RAD52基因参与DNA损伤修复。",
                "review_status": "待医学审核",
            }
        ],
    )
    overlay = tmp_path / "overlay.yaml"
    overlay.write_text("gene_sections: []\ndrug_sections: []\n", encoding="utf-8")
    summary = tmp_path / "summary.json"
    summary.write_text('{"issues": 0, "added": 0, "skipped": 1}', encoding="utf-8")
    retest = tmp_path / "retest.xlsx"
    _write_context_retest(retest, ["PASS"])

    class Args:
        prod_overlay = overlay
        approved_overlay = overlay
        review_workbook = review
        approved_summary = summary
        context_retest = retest

    report = build_release_readiness_report(Args)
    assert report["status"] == "not_release_ready"
    failed = [check["name"] for check in report["checks"] if not check["passed"]]
    assert "review_workbook_has_no_pending_rows" in failed


def test_release_readiness_gate_accepts_clean_approved_package(tmp_path):
    review = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        review,
        [
            {
                "section": "gene_sections",
                "gene": "RAD52",
                "intro": "RAD52基因参与DNA损伤修复。",
                "review_status": "通过",
            }
        ],
    )
    overlay = tmp_path / "overlay.yaml"
    overlay.write_text(
        "gene_sections:\n"
        "- gene: RAD52\n"
        "  intro: RAD52基因参与DNA损伤修复。\n"
        "drug_sections: []\n",
        encoding="utf-8",
    )
    summary = tmp_path / "summary.json"
    summary.write_text('{"issues": 0, "added": 1, "skipped": 0}', encoding="utf-8")
    retest = tmp_path / "retest.xlsx"
    _write_context_retest(retest, ["PASS", "PASS"])

    class Args:
        prod_overlay = overlay
        approved_overlay = overlay
        review_workbook = review
        approved_summary = summary
        context_retest = retest

    report = build_release_readiness_report(Args)
    assert report["status"] == "release_ready"
    assert all(check["passed"] for check in report["checks"])


def test_release_readiness_gate_accepts_deferred_non_approved_rows(tmp_path):
    review = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        review,
        [
            {
                "section": "gene_sections",
                "gene": "RAD52",
                "intro": "RAD52基因参与DNA损伤修复。",
                "review_status": "通过",
            },
            {
                "section": "gene_sections",
                "gene": "AXIN2",
                "intro": "AXIN2基因参与Wnt信号通路调控。",
                "review_status": "暂缓",
            },
        ],
    )
    overlay = tmp_path / "overlay.yaml"
    overlay.write_text(
        "gene_sections:\n"
        "- gene: RAD52\n"
        "  intro: RAD52基因参与DNA损伤修复。\n"
        "drug_sections: []\n",
        encoding="utf-8",
    )
    summary = tmp_path / "summary.json"
    summary.write_text(
        '{"issues": 0, "added": 1, "skipped": 1, "approved_skipped": 0, "non_approved_skipped": 1}',
        encoding="utf-8",
    )
    retest = tmp_path / "retest.xlsx"
    _write_context_retest(retest, ["PASS"])

    class Args:
        prod_overlay = overlay
        approved_overlay = overlay
        review_workbook = review
        approved_summary = summary
        context_retest = retest

    report = build_release_readiness_report(Args)
    assert report["status"] == "release_ready"
    assert all(check["passed"] for check in report["checks"])


def test_release_readiness_gate_rejects_gene_context_mismatch(tmp_path):
    review = tmp_path / "review.xlsx"
    _write_batch9_review_workbook(
        review,
        [
            {
                "section": "gene_sections",
                "gene": "MYH11",
                "intro": "PALB2基因参与DNA损伤修复。",
                "review_status": "通过",
            }
        ],
    )
    overlay = tmp_path / "overlay.yaml"
    overlay.write_text(
        "gene_sections:\n"
        "- gene: MYH11\n"
        "  intro: PALB2基因参与DNA损伤修复。\n"
        "drug_sections: []\n",
        encoding="utf-8",
    )
    summary = tmp_path / "summary.json"
    summary.write_text('{"issues": 0, "added": 1, "skipped": 0}', encoding="utf-8")
    retest = tmp_path / "retest.xlsx"
    _write_context_retest(retest, ["PASS"])

    class Args:
        prod_overlay = overlay
        approved_overlay = overlay
        review_workbook = review
        approved_summary = summary
        context_retest = retest

    report = build_release_readiness_report(Args)
    assert report["status"] == "not_release_ready"
    failed = [check["name"] for check in report["checks"] if not check["passed"]]
    assert "approved_overlay_has_no_gene_context_mismatches" in failed


def test_medical_triage_suggests_clean_base_kb_gene_row_for_approval():
    rows = triage_rows(
        [
            {
                "review_source": "base_gene_kb_supported_gap_review",
                "section": "gene_sections",
                "gene": "RAD52",
                "has_intro": "是",
                "intro": "RAD52基因参与DNA双链断裂修复过程。",
                "review_status": "待医学审核",
            }
        ]
    )
    assert rows[0]["machine_suggestion"] == "建议通过"
    assert rows[0]["suggested_review_status"] == "通过"
    assert rows[0]["risk_flags"] == ""


def test_medical_triage_routes_sample_context_to_defer():
    rows = triage_rows(
        [
            {
                "review_source": "historical_final_report_drug_pair_review",
                "section": "drug_sections",
                "gene": "XRCC2",
                "c_hgvs": "c.190C>T",
                "p_hgvs": "p.R64*",
                "drug_name": "奥拉帕利",
                "relation": "该样本同时检出MSH6基因突变，因此可能从PARP抑制剂获益。",
                "clinical": "临床研究提示PARP抑制剂可能获益。",
                "has_relation": "是",
                "has_clinical": "是",
                "review_status": "待医学审核",
            }
        ]
    )
    assert rows[0]["machine_suggestion"] == "建议暂缓"
    assert "样本上下文" in rows[0]["risk_flags"]


def test_medical_triage_routes_incomplete_drug_row_to_modify():
    rows = triage_rows(
        [
            {
                "review_source": "historical_final_report_drug_pair_review",
                "section": "drug_sections",
                "gene": "XRCC2",
                "drug_name": "尼拉帕利",
                "relation": "XRCC2突变可能提示PARP抑制剂获益。",
                "clinical": "",
                "review_status": "待医学审核",
            }
        ]
    )
    assert rows[0]["machine_suggestion"] == "建议修改后通过"
    assert "clinical缺失" in rows[0]["risk_flags"]


def test_triage_subset_release_maps_only_suggested_pass_to_approved():
    approved = release_row(
        {
            "machine_suggestion": "建议通过",
            "machine_reason": "基础知识库支撑",
            "section": "gene_sections",
            "gene": "RAD52",
            "intro": "RAD52基因参与DNA损伤修复。",
        }
    )
    deferred = release_row(
        {
            "machine_suggestion": "建议人工精审",
            "machine_reason": "需确认CRC358语境",
            "section": "gene_sections",
            "gene": "AXIN2",
            "intro": "AXIN2基因参与Wnt信号通路调控。",
        }
    )
    assert approved["review_status"] == "通过"
    assert deferred["review_status"] == "暂缓"
    assert workbook_row(approved)["has_intro"] == "是"
    assert workbook_row(deferred)["review_status"] == "暂缓"


def test_triage_subset_release_can_explicitly_approve_cross_cancer_gene_intro():
    row = {
        "machine_suggestion": "建议人工精审",
        "review_source": "all_cancer_final_report_gene_intro_support",
        "section": "gene_sections",
        "gene": "AXIN2",
        "intro": "AXIN2基因参与Wnt信号通路调控。",
        "risk_flags": "",
    }
    assert release_row(row)["review_status"] == "暂缓"
    assert release_row(row, approve_cross_cancer_gene_intros=True)["review_status"] == "通过"


def test_triage_subset_release_can_finalize_historical_drug_pair_as_not_in_kb():
    row = {
        "machine_suggestion": "建议人工精审",
        "review_source": "historical_final_report_drug_pair_review",
        "section": "drug_sections",
        "gene": "XRCC2",
        "drug_name": "尼拉帕利（Niraparib）",
        "relation": "XRCC2功能缺失可能与PARP抑制剂相关。",
        "clinical": "相关篮子试验正在开展。",
        "risk_flags": "",
    }
    out = release_row(row, reject_historical_drug_pairs=True)
    assert out["review_status"] == "不入库"
    assert "单个XRCC2位点不足以作为通用用药提示" in out["review_notes"]
