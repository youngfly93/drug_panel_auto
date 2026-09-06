"""Membership derivation is lossless and deliberately independent of grading."""

import copy
import hashlib
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from scripts.derive_panel_input import derive_panel_input

GENES_13 = "EGFR AKT1 ALK BRAF ERBB2 KRAS MAP2K1 MET NRAS PIK3CA RET ROS1 TP53".split()


def source_book(tmp_path, flags=("ExistInsmall588",)):
    path = tmp_path / "SYN-SUPERSET.xlsx"
    book = Workbook()
    book.remove(book.active)
    for name in ("Variations", "Hereditary_tumor"):
        sheet = book.create_sheet(name)
        sheet.append(["ExistIn552", *flags, "Gene_Symbol", "cHGVS", "pHGVS_S"])
        for gene, c, p, grade in [
            ("BRAF", "c.1799T>A", "p.V600E", "Ⅰ类"),
            ("ERBB2", "c.1979G>A", "p.G660D", "Ⅰ类"),
            ("TP53", "c.734G>A", "p.G245D", "Ⅱ类"),
            ("PTEN", "c.802-2A>T", "", "Ⅱ类"),
            ("ALK", "c.1837G>A", "p.A613T", 1),
            ("PIK3CA", "c.3197C>T", "p.A1066V", "Ⅱ类"),
        ]:
            sheet.append([grade, *[1 for _ in flags], gene, c, p])
        sheet["A1"].font = Font(name="Arial", bold=True)
        sheet.freeze_panes = "D2"
    for name in (
        "Cnv",
        "TMB",
        "Msisensor",
        "QC",
        "CtDrug",
        "Hotspot",
        "Fusion",
        "HLA",
        "Ct1000",
        "Ct15",
    ):
        sheet = book.create_sheet(name)
        sheet.append(["ExistIn137", "raw", "formula"])
        sheet.append([1, "unchanged", "=1+2"])
    book.save(path)
    return path


@pytest.mark.parametrize(
    "flags", [("ExistInsmall588",), ("ExistInsmall588", "ExistInsmall358"), ()]
)
def test_only_membership_columns_change_and_all_other_parts_are_identical(tmp_path, flags):
    source = source_book(tmp_path, flags)
    original = hashlib.sha256(source.read_bytes()).hexdigest()
    receipt = derive_panel_input(
        source,
        panel_id="lung_13",
        genes=GENES_13,
        flag_column="ExistInsmall13",
        output_dir=tmp_path / ".work",
    )
    assert hashlib.sha256(source.read_bytes()).hexdigest() == original
    derived = Path(receipt["output"])
    assert derived.name == "SYN-SUPERSET-derived-lung_13.xlsx"
    before, after = load_workbook(source), load_workbook(derived)
    assert before.sheetnames == after.sheetnames
    for name in before.sheetnames:
        raw, new = list(before[name].values), list(after[name].values)
        if name not in ("Variations", "Hereditary_tumor"):
            assert raw == new
            continue
        assert "ExistInsmall13" in new[0]
        assert not any(flag in new[0] for flag in flags)
        actual = [dict(zip(new[0], r)) for r in new[1:]]
        for old, result in zip(raw[1:], actual):
            original_row = dict(zip(raw[0], old))
            assert result["ExistInsmall13"] == (
                1 if original_row["Gene_Symbol"] in GENES_13 else None
            )
            assert {k: v for k, v in result.items() if k != "ExistInsmall13"} == {
                k: v for k, v in original_row.items() if k not in flags
            }
        # A membership flag must not rewrite the numeric/classification distinction.
        assert next(r for r in actual if r["Gene_Symbol"] == "ALK")["ExistIn552"] == 1
        assert next(r for r in actual if r["Gene_Symbol"] == "PIK3CA")["ExistInsmall13"] == 1
        assert copy.copy(after[name]["A1"].font) == copy.copy(before[name]["A1"].font)
    before.close()
    after.close()
    with zipfile.ZipFile(source) as raw, zipfile.ZipFile(derived) as new:
        changed = {name for name in raw.namelist() if raw.read(name) != new.read(name)}
    assert changed == {"xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"}
    assert receipt["clinical_values_inferred"] is False


def test_custom_product_flag_and_no_overwrite(tmp_path):
    source = source_book(tmp_path)
    kwargs = dict(
        panel_id="lung_62_pdl1",
        genes=GENES_13,
        flag_column="ExistInsmall62pdl1",
        output_dir=tmp_path / ".work",
    )
    receipt = derive_panel_input(source, **kwargs)
    assert receipt["membership_column"] == "ExistInsmall62pdl1"
    with pytest.raises(FileExistsError):
        derive_panel_input(source, **kwargs)


def test_refuses_nonprivate_output_and_empty_gene_set(tmp_path):
    source = source_book(tmp_path)
    with pytest.raises(ValueError, match=".work"):
        derive_panel_input(
            source,
            panel_id="lung_13",
            genes=GENES_13,
            flag_column="ExistInsmall13",
            output_dir=tmp_path / "tracked",
        )
    with pytest.raises(ValueError, match="gene-symbol"):
        derive_panel_input(
            source,
            panel_id="lung_13",
            genes=[],
            flag_column="ExistInsmall13",
            output_dir=tmp_path / ".work",
        )


def test_membership_is_not_silently_reduced_to_three_expected_report_rows(tmp_path):
    source = source_book(tmp_path)
    receipt = derive_panel_input(
        source,
        panel_id="lung_13",
        genes=GENES_13,
        flag_column="ExistInsmall13",
        output_dir=tmp_path / ".work",
    )
    assert receipt["worksheets"]["Variations"]["flagged_rows"] == 5
    # PIK3CA belongs to the declared 13-gene set; removing it to obtain three
    # report rows would violate the derivation contract, not fix this test.
