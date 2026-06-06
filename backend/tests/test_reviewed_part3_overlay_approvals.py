from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "scripts" / "build_reviewed_part3_overlay_from_approvals.py"


def _load_module():
    spec = importlib.util.spec_from_file_location("reviewed_part3_overlay_approvals", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _append_table(ws, headers: list[str], rows: list[list[str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _write_review_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "历史基因级候选"
    _append_table(
        ws,
        [
            "基因",
            "候选类型",
            "source_id",
            "content_hash",
            "产品族",
            "基因数",
            "命中段落",
            "候选上下文",
            "审核结论",
            "备注",
        ],
        [
            [
                "ERBB2",
                "基因简介",
                "SRC001",
                "hash001",
                "结直肠癌358",
                "358",
                "",
                "ERBB2基因编码受体酪氨酸激酶家族成员，参与细胞增殖和信号通路调控。",
                "通过",
                "",
            ],
            [
                "ERBB2",
                "基因变异解析",
                "SRC002",
                "hash002",
                "结直肠癌358",
                "358",
                "",
                "ERBB2异常可影响下游信号通路，并可能与结直肠癌发生发展及用药评估相关。",
                "修改后通过",
                "",
            ],
            [
                "ALK",
                "基因简介",
                "SRC003",
                "hash003",
                "结直肠癌358",
                "358",
                "",
                "ALK基因编码受体酪氨酸激酶。",
                "待审核",
                "",
            ],
            [
                "DNMT3A",
                "基因变异解析",
                "SRC004",
                "hash004",
                "结直肠癌358",
                "358",
                "",
                "该样本检出DNMT3A基因c.2322+1G>A突变，此突变在样本中的突变丰度为0.84%。",
                "通过",
                "个案句不应入库",
            ],
            [
                "KMT2A",
                "基因简介",
                "SRC005",
                "hash005",
                "结直肠癌358",
                "358",
                "",
                "KMT2A基因参与染色质调控。",
                "不通过",
                "",
            ],
        ],
    )

    ws2 = wb.create_sheet("历史精确位点候选")
    _append_table(
        ws2,
        [
            "优先级",
            "基因",
            "cHGVS",
            "pHGVS",
            "候选类型",
            "source_id",
            "content_hash",
            "产品族",
            "基因数",
            "命中段落",
            "候选上下文",
            "审核结论",
            "备注",
        ],
        [
            [
                "P1",
                "TP53",
                "c.821T>A",
                "p.V274D",
                "基因变异解析",
                "SRC006",
                "hash006",
                "结直肠癌358",
                "358",
                "",
                "TP53 p.V274D突变位于DNA结合结构域，相关异常可能影响蛋白功能。",
                "通过",
                "",
            ],
            [
                "P1",
                "FGFR1",
                "c.1648G>T",
                "p.A550S",
                "基因变异解析",
                "SRC007",
                "hash007",
                "结直肠癌358",
                "358",
                "",
                "此突变在样本中的突变丰度为35.08%。",
                "通过",
                "个案丰度不应入库",
            ],
        ],
    )
    wb.save(path)


def _write_gap_review_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "需补库位点"
    _append_table(
        ws,
        [
            "优先级",
            "基因",
            "cHGVS",
            "pHGVS",
            "等级",
            "当前状态",
            "基础候选简介",
            "基础候选解析",
            "审核后简介",
            "审核后解析",
            "审核结论",
            "备注",
        ],
        [
            [
                "P1",
                "ERBB2",
                "c.1133C>T",
                "p.P378L",
                "Ⅱ类",
                "基础库通用内容",
                "ERBB2基因的功能与肿瘤发生发展密切相关。",
                "ERBB2基因突变在多种肿瘤中被报道，其临床意义是当前研究热点。",
                "",
                "",
                "通过",
                "基础候选不应直接入库",
            ],
            [
                "P1",
                "KMT2A",
                "c.3950_3954del",
                "p.K1317Sfs*7",
                "Ⅱ类",
                "基础库通用内容",
                "KMT2A基因的功能与肿瘤发生发展密切相关。",
                "KMT2A基因突变在多种肿瘤中被报道。",
                "KMT2A基因编码组蛋白甲基转移酶相关蛋白，参与染色质修饰和转录调控。",
                "KMT2A异常可能影响表观遗传调控和细胞增殖过程，具体临床意义需结合变异类型和结直肠癌背景综合判断。",
                "通过",
                "审核后文本可入库",
            ],
        ],
    )
    wb.save(path)


def _write_overlay(path: Path, sections: list[dict[str, str]] | None = None) -> None:
    path.write_text(
        yaml.safe_dump({"schema_version": 1, "gene_sections": sections or []}, allow_unicode=True),
        encoding="utf-8",
    )


def test_build_draft_accepts_only_reviewed_safe_rows(tmp_path):
    module = _load_module()
    review_xlsx = tmp_path / "review.xlsx"
    existing = tmp_path / "existing.yaml"
    output = tmp_path / "draft.yaml"
    _write_review_workbook(review_xlsx)
    _write_overlay(existing)

    result = module.build_draft(
        review_xlsx=review_xlsx,
        existing_overlay=existing,
        output=output,
        apply=False,
        approve_all=False,
    )

    assert result["approved_historical_gene_rows"] == 1
    assert result["approved_exact_rows"] == 1
    assert result["new_sections"] == 2

    data = yaml.safe_load(output.read_text(encoding="utf-8"))
    sections = data["gene_sections"]
    by_gene = {row["gene"]: row for row in sections}

    assert set(by_gene) == {"ERBB2", "TP53"}
    assert "受体酪氨酸激酶" in by_gene["ERBB2"]["intro"]
    assert "用药评估相关" in by_gene["ERBB2"]["mutation_analysis"]
    assert by_gene["TP53"]["c_hgvs"] == "c.821T>A"
    assert by_gene["TP53"]["p_hgvs"] == "p.V274D"

    rendered = output.read_text(encoding="utf-8")
    assert "source_id" not in rendered
    assert "content_hash" not in rendered
    assert "SRC00" not in rendered
    assert "该样本检出" not in rendered
    assert "突变丰度" not in rendered
    assert "待审核" not in rendered
    assert "不通过" not in rendered


def test_build_draft_skips_existing_overlay_keys(tmp_path):
    module = _load_module()
    review_xlsx = tmp_path / "review.xlsx"
    existing = tmp_path / "existing.yaml"
    output = tmp_path / "draft.yaml"
    _write_review_workbook(review_xlsx)
    _write_overlay(existing, [{"gene": "ERBB2", "intro": "existing"}])

    result = module.build_draft(
        review_xlsx=review_xlsx,
        existing_overlay=existing,
        output=output,
        apply=False,
        approve_all=False,
    )

    assert result["skipped_existing"] == 1
    data = yaml.safe_load(output.read_text(encoding="utf-8"))
    assert [row["gene"] for row in data["gene_sections"]] == ["TP53"]


def test_gap_base_candidate_text_is_not_promoted_without_reviewed_rewrite(tmp_path):
    module = _load_module()
    review_xlsx = tmp_path / "gap_review.xlsx"
    existing = tmp_path / "existing.yaml"
    output = tmp_path / "draft.yaml"
    _write_gap_review_workbook(review_xlsx)
    _write_overlay(existing)

    result = module.build_draft(
        review_xlsx=review_xlsx,
        existing_overlay=existing,
        output=output,
        apply=False,
        approve_all=False,
    )

    assert result["approved_variant_rows"] == 1
    assert result["new_sections"] == 1
    data = yaml.safe_load(output.read_text(encoding="utf-8"))
    assert [row["gene"] for row in data["gene_sections"]] == ["KMT2A"]
    rendered = output.read_text(encoding="utf-8")
    assert "ERBB2" not in rendered
    assert "当前研究热点" not in rendered
    assert "审核后文本可入库" not in rendered
