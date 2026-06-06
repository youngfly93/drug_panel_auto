from __future__ import annotations

import importlib.util
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "scripts" / "harvest_crc_part3_candidates.py"


def _load_module():
    spec = importlib.util.spec_from_file_location("crc_part3_harvest", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _write_docx(path: Path, paragraphs: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body = []
    for paragraph in paragraphs:
        body.append(f"<w:p><w:r><w:t>{paragraph}</w:t></w:r></w:p>")
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body)
        + "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("word/document.xml", xml)


def _write_gap_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "待审核候选内容"
    ws.append(
        [
            "优先级",
            "基因",
            "cHGVS",
            "pHGVS",
            "等级",
            "当前状态",
            "候选基因简介",
            "候选变异解析",
        ]
    )
    ws.append(
        [
            "P1",
            "ERBB2",
            "c.1133C>T",
            "p.P378L",
            "Ⅲ类",
            "基础库通用内容",
            "",
            "",
        ]
    )
    wb.save(path)


def test_crc_part3_harvest_output_is_deidentified(tmp_path):
    module = _load_module()
    corpus = tmp_path / "肠癌"
    report = corpus / "测试患者-直肠癌-结直肠癌358基因+msi-project-case001-终版.docx"
    _write_docx(
        report,
        [
            "◆ ERBB2：c.1133C>T，p.P378L；12.3%",
            "ERBB2基因编码人表皮生长因子受体2蛋白，是受体酪氨酸激酶家族成员，参与细胞增殖和信号通路调控。",
            "该样本检出ERBB2基因c.2033G>A，p.R678Q错义突变，此突变在样本中的突变丰度为23.01%。",
            "ERBB2基因在结直肠癌中可见扩增或突变，相关异常可能与疾病发生发展及用药评估相关。",
        ],
    )
    gap = tmp_path / "gap.xlsx"
    output = tmp_path / "candidates.xlsx"
    _write_gap_workbook(gap)

    result = module.harvest(
        gap_xlsx=gap,
        corpus_dir=corpus,
        output=output,
        priorities={"P1"},
        max_gene_hits=5,
        max_exact_hits=5,
    )

    assert result["gap_candidates"] == 1
    assert result["docx_files"] == 1
    assert result["exact_candidates"] >= 1

    wb = load_workbook(output, data_only=True)
    values = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            values.extend(str(value) for value in row if value is not None)
    combined = "\n".join(values)

    assert "测试患者" not in combined
    assert "case001" not in combined
    assert report.name not in combined
    assert "source_id" in combined
    assert "content_hash" in combined
    assert "ERBB2" in combined

    gene_sheet_values = []
    for row in wb["历史基因级候选"].iter_rows(values_only=True):
        gene_sheet_values.extend(str(value) for value in row if value is not None)
    gene_sheet_text = "\n".join(gene_sheet_values)
    assert "该样本检出" not in gene_sheet_text
    assert "突变丰度" not in gene_sheet_text
    assert "c.2033G>A" not in gene_sheet_text
